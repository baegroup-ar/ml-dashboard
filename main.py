import os
import httpx
import secrets
import asyncio
import json
from urllib.parse import urlencode
from datetime import datetime, timedelta, timezone
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, HTTPException, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import URLSafeTimedSerializer, BadSignature
import bcrypt
from sqlalchemy import create_engine, text
from sqlalchemy.pool import StaticPool

ML_CLIENT_ID = os.environ["ML_CLIENT_ID"]
ML_CLIENT_SECRET = os.environ["ML_CLIENT_SECRET"]
APP_URL = os.environ.get("APP_URL", "http://localhost:8000")
SECRET_KEY = os.environ.get("SECRET_KEY", secrets.token_hex(32))
DATABASE_URL = os.environ["DATABASE_URL"].replace("postgres://", "postgresql+pg8000://").replace("postgresql://", "postgresql+pg8000://")
ADMIN_EMAIL = os.environ["ADMIN_EMAIL"]
ADMIN_PASSWORD = os.environ["ADMIN_PASSWORD"]
MASTER_EMAIL = os.environ.get("MASTER_EMAIL")
MASTER_PASSWORD = os.environ.get("MASTER_PASSWORD")
MASTER_NAME = os.environ.get("MASTER_NAME", "Maestro")

ML_AUTH_URL = "https://auth.mercadolibre.com.ar/authorization"
ML_TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
ML_API_URL = "https://api.mercadolibre.com"
SHIPPING_LOGIC_VERSION = "v34-logistic-type-in-orders"
PROMO_ITEM_SCAN_CACHE_TTL_SECONDS = 300
PROMO_ITEM_SCAN_CACHE: dict[str, dict] = {}
# Cache de resultados ya armados de /api/promociones/{acc}/{promo}/items.
# Permite que cambiar de pestaña o recargar sea instantaneo. Se invalida
# cuando se aplica o se quita un descuento en esa cuenta.
PROMO_ITEMS_RESULT_TTL_SECONDS = 120
PROMO_ITEMS_RESULT_CACHE: dict[str, dict] = {}
# Mínimo de descuento MÁS BAJO que ML reportó alguna vez por (promo, item).
# La API de ML para campañas FLEXIBLE_PERCENTAGE es eventualmente consistente:
# devuelve un `max_discounted_price` distinto entre llamadas (a veces 10%, a
# veces 5% para el MISMO item, segundos después). El "mínimo requerido para
# participar" es el descuento MÁS CHICO que ML aceptó en cualquier momento:
# si alguna vez dijo 5%, entonces 5% alcanza. Guardamos ese piso para que el
# valor no parpadee 5%↔10% y quede estable. Clave: f"{promotion_id}:{item_id}".
PROMO_MIN_PCT_SEEN: dict[str, float] = {}
_PROMO_FLOOR_LOADED = False


def _promo_floor_key(promotion_id, item_id) -> str:
    return f"{promotion_id}:{str(item_id or '').upper()}"


def _load_promo_floor_from_db() -> None:
    """Carga el piso histórico persistido (sobrevive redeploys)."""
    global _PROMO_FLOOR_LOADED
    try:
        rows = db_fetchall("SELECT promotion_id, item_id, min_pct FROM promo_min_floor")
        for r in rows:
            PROMO_MIN_PCT_SEEN[_promo_floor_key(r["promotion_id"], r["item_id"])] = float(r["min_pct"])
        _PROMO_FLOOR_LOADED = True
    except Exception:
        pass


def _record_promo_floor(promotion_id, item_id, pct: float) -> bool:
    """Fija el piso por item si `pct` es menor al guardado. Devuelve True si
    bajó/creó el piso (es decir, cambió). Persiste en DB (best-effort)."""
    if pct is None or pct <= 0:
        return False
    key = _promo_floor_key(promotion_id, item_id)
    prev = PROMO_MIN_PCT_SEEN.get(key)
    if prev is not None and pct >= prev:
        return False
    PROMO_MIN_PCT_SEEN[key] = pct
    try:
        db_execute(
            "INSERT INTO promo_min_floor (promotion_id, item_id, min_pct) "
            "VALUES (:p, :i, :m) ON CONFLICT (promotion_id, item_id) "
            "DO UPDATE SET min_pct = EXCLUDED.min_pct, updated_at = NOW() "
            "WHERE promo_min_floor.min_pct > EXCLUDED.min_pct",
            {"p": promotion_id, "i": str(item_id or "").upper(), "m": pct},
        )
    except Exception:
        pass
    return True


def _invalidate_promo_items_cache(account_id: int) -> None:
    prefix = f"{account_id}:"
    for k in [k for k in PROMO_ITEMS_RESULT_CACHE if k.startswith(prefix)]:
        PROMO_ITEMS_RESULT_CACHE.pop(k, None)


# Warmers en curso (1 por promo) para no spawnear duplicados.
# Promos cuyo mínimo por item ya convergió en este proceso. Evita repetir la
# convergencia sincrónica en cada carga: una vez fijado el piso (y persistido
# en DB), las cargas siguientes salen instantáneas.
PROMO_CONVERGED: set = set()

serializer = URLSafeTimedSerializer(SECRET_KEY)
engine = create_engine(DATABASE_URL, pool_pre_ping=True)


def clean_env_secret(value: Optional[str]) -> str:
    value = (value or "").strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
        value = value[1:-1].strip()
    return value


def master_credentials() -> tuple[str, str, str]:
    email = clean_env_secret(MASTER_EMAIL).lower()
    password = clean_env_secret(MASTER_PASSWORD)
    name = clean_env_secret(MASTER_NAME) or "Maestro"
    return email, password, name


def upsert_master_user(conn=None) -> bool:
    master_email, master_password, master_name = master_credentials()
    if not master_email or not master_password:
        return False

    master_hash = bcrypt.hashpw(master_password.encode(), bcrypt.gensalt()).decode()
    params = {"e": master_email, "h": master_hash, "n": master_name}
    query = text("""
        INSERT INTO users (email, password_hash, name, is_admin, is_master, role_label, permissions)
        VALUES (:e, :h, :n, TRUE, TRUE, 'Maestro', '[]'::JSONB)
        ON CONFLICT (email) DO UPDATE SET
            password_hash = EXCLUDED.password_hash,
            name = EXCLUDED.name,
            is_admin = TRUE,
            is_master = TRUE,
            role_label = 'Maestro',
            reset_token = NULL,
            reset_expires_at = NULL
    """)
    if conn is not None:
        conn.execute(query, params)
        return True
    with engine.connect() as c:
        c.execute(query, params)
        c.commit()
    return True


def init_db():
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                name TEXT NOT NULL,
                is_admin BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """))
        # Permisos por usuario + reset de contraseña sin que el admin la vea.
        conn.execute(text("ALTER TABLE users ALTER COLUMN password_hash DROP NOT NULL"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS role_label TEXT DEFAULT 'Colaborador'"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS permissions JSONB DEFAULT '[]'::JSONB"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS reset_token TEXT"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS reset_expires_at TIMESTAMP"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_master BOOLEAN DEFAULT FALSE"))
        # owner_id: el admin "dueño" de un colaborador. El colaborador ve las
        # cuentas de ML de su dueño. NULL para admins/master (cuentas propias).
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS owner_id INTEGER"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_users_reset_token ON users(reset_token)"))
        conn.execute(text("ALTER TABLE ml_accounts ADD COLUMN IF NOT EXISTS cache_fetched_from DATE DEFAULT NULL"))
        # Backfill: set cache_fetched_from from existing order_snapshot_cache so existing accounts
        # don't lose their "already fetched" coverage after this migration.
        conn.execute(text("""
            UPDATE ml_accounts ma SET cache_fetched_from = sub.min_date
            FROM (SELECT account_id, MIN(paid_date) AS min_date FROM order_snapshot_cache GROUP BY account_id) sub
            WHERE ma.id = sub.account_id AND ma.cache_fetched_from IS NULL
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS ml_accounts (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                ml_user_id TEXT NOT NULL,
                nickname TEXT NOT NULL,
                access_token TEXT NOT NULL,
                refresh_token TEXT NOT NULL,
                expires_at TIMESTAMP NOT NULL,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(user_id, ml_user_id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS shipment_cost_cache (
                shipping_id BIGINT PRIMARY KEY,
                cost NUMERIC(10,2) NOT NULL,
                buyer_cost NUMERIC(10,2) DEFAULT NULL,
                logistic_type VARCHAR(50) DEFAULT NULL,
                cached_at TIMESTAMP DEFAULT NOW()
            )
        """))
        conn.execute(text("ALTER TABLE shipment_cost_cache ADD COLUMN IF NOT EXISTS buyer_cost NUMERIC(10,2) DEFAULT NULL"))
        conn.execute(text("ALTER TABLE shipment_cost_cache ALTER COLUMN buyer_cost DROP NOT NULL"))
        conn.execute(text("ALTER TABLE shipment_cost_cache ADD COLUMN IF NOT EXISTS logistic_type VARCHAR(50) DEFAULT NULL"))
        conn.execute(text("ALTER TABLE shipment_cost_cache ADD COLUMN IF NOT EXISTS bonificacion NUMERIC(10,2) DEFAULT NULL"))
        conn.execute(text("ALTER TABLE shipment_cost_cache ADD COLUMN IF NOT EXISTS list_cost NUMERIC(10,2) DEFAULT NULL"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS order_snapshot_cache (
                account_id INTEGER NOT NULL REFERENCES ml_accounts(id) ON DELETE CASCADE,
                order_id TEXT NOT NULL,
                paid_date DATE NOT NULL,
                paid_time TEXT,
                payload JSONB NOT NULL,
                details_complete BOOLEAN DEFAULT FALSE,
                updated_at TIMESTAMP DEFAULT NOW(),
                PRIMARY KEY (account_id, order_id)
            )
        """))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_order_snapshot_cache_account_date ON order_snapshot_cache (account_id, paid_date)"))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS product_costs (
                account_id INTEGER NOT NULL REFERENCES ml_accounts(id) ON DELETE CASCADE,
                sku TEXT NOT NULL,
                cost NUMERIC(12,2) NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW(),
                PRIMARY KEY (account_id, sku)
            )
        """))
        # Migración a costos versionados (fecha de vigencia + tasa de IVA).
        conn.execute(text("ALTER TABLE product_costs ADD COLUMN IF NOT EXISTS valid_from DATE"))
        conn.execute(text("ALTER TABLE product_costs ADD COLUMN IF NOT EXISTS iva_included BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE product_costs ADD COLUMN IF NOT EXISTS iva_rate NUMERIC(5,2) DEFAULT 21"))
        conn.execute(text("UPDATE product_costs SET valid_from = CAST(updated_at AS DATE) WHERE valid_from IS NULL"))
        conn.execute(text("ALTER TABLE product_costs ALTER COLUMN valid_from SET NOT NULL"))
        # Cambiar PK a (account_id, sku, valid_from) si todavía no lo es.
        pk_cols = conn.execute(text("""
            SELECT string_agg(a.attname, ',' ORDER BY array_position(i.indkey, a.attnum))
            FROM pg_index i
            JOIN pg_attribute a ON a.attrelid = i.indrelid AND a.attnum = ANY(i.indkey)
            WHERE i.indrelid = 'product_costs'::regclass AND i.indisprimary
        """)).fetchone()
        if pk_cols and pk_cols[0] != "account_id,sku,valid_from":
            conn.execute(text("ALTER TABLE product_costs DROP CONSTRAINT product_costs_pkey"))
            conn.execute(text("ALTER TABLE product_costs ADD PRIMARY KEY (account_id, sku, valid_from)"))
        # Tabla de tarifas Flex (costo real que el vendedor paga a la
        # mensajería, por zona de entrega y con fecha de vigencia).
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS flex_tariffs (
                account_id INTEGER NOT NULL REFERENCES ml_accounts(id) ON DELETE CASCADE,
                zona VARCHAR(20) NOT NULL,
                tarifa NUMERIC(12,2) NOT NULL,
                iva_rate NUMERIC(5,2) DEFAULT 21,
                valid_from DATE NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW(),
                PRIMARY KEY (account_id, zona, valid_from)
            )
        """))
        # tarifa_ml: precio público que ML cobra por esa zona (a cargo del
        # comprador). Sirve de referencia para asociar cada shipment Flex
        # con la zona correcta (matcheamos contra shipping_option.list_cost).
        conn.execute(text("ALTER TABLE flex_tariffs ADD COLUMN IF NOT EXISTS tarifa_ml NUMERIC(12,2)"))
        # Base de descuentos por MLA (item_id de Mercado Libre). El usuario
        # mantiene su mapeo MLA → SKU → descuento en su Excel. El cruce con
        # las promociones se hace por MLA directamente (sin fallback por SKU).
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS product_discounts (
                account_id INTEGER NOT NULL REFERENCES ml_accounts(id) ON DELETE CASCADE,
                sku TEXT NOT NULL,
                discount_pct NUMERIC(5,2) NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW(),
                PRIMARY KEY (account_id, sku)
            )
        """))
        # Migración a MLA: si la tabla todavía no tiene columna mla, la
        # recreamos (no hay data crítica que conservar).
        mla_col = conn.execute(text("""
            SELECT 1 FROM information_schema.columns
            WHERE table_name='product_discounts' AND column_name='mla'
        """)).fetchone()
        if not mla_col:
            conn.execute(text("DROP TABLE IF EXISTS product_discounts CASCADE"))
            conn.execute(text("""
                CREATE TABLE product_discounts (
                    account_id INTEGER NOT NULL REFERENCES ml_accounts(id) ON DELETE CASCADE,
                    mla TEXT NOT NULL,
                    sku TEXT,
                    discount_pct NUMERIC(5,2) NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (account_id, mla)
                )
            """))
        # Piso histórico del mínimo de descuento por (promo, item). La API de
        # ML para campañas FLEXIBLE_PERCENTAGE es eventualmente consistente y
        # devuelve a veces el mínimo genérico (10%) y a veces el real (5%);
        # guardamos el más bajo visto para que el valor quede estable y
        # sobreviva redeploys.
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS promo_min_floor (
                promotion_id TEXT NOT NULL,
                item_id TEXT NOT NULL,
                min_pct NUMERIC(6,2) NOT NULL,
                updated_at TIMESTAMP DEFAULT NOW(),
                PRIMARY KEY (promotion_id, item_id)
            )
        """))
        # Migrar iva_included → iva_rate UNA VEZ (controlado por app_meta).
        iva_migrated = conn.execute(text(
            "SELECT 1 FROM app_meta WHERE key='iva_rate_migrated'"
        )).fetchone()
        if not iva_migrated:
            conn.execute(text("""
                UPDATE product_costs
                SET iva_rate = CASE WHEN COALESCE(iva_included, FALSE) THEN 0 ELSE 21 END
            """))
            conn.execute(text(
                "INSERT INTO app_meta (key, value) VALUES ('iva_rate_migrated', '1')"
                " ON CONFLICT (key) DO NOTHING"
            ))
        # Invalida el caché cuando cambia la lógica de cálculo de envío.
        current = conn.execute(text("SELECT value FROM app_meta WHERE key='shipping_logic_version'")).fetchone()
        if not current or current[0] != SHIPPING_LOGIC_VERSION:
            conn.execute(text("DELETE FROM shipment_cost_cache"))
            # También limpia el snapshot de órdenes para recalcular bonificaciones.
            conn.execute(text("DELETE FROM order_snapshot_cache"))
            conn.execute(text(
                "INSERT INTO app_meta (key, value) VALUES ('shipping_logic_version', :v)"
                " ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()"
            ), {"v": SHIPPING_LOGIC_VERSION})
        existing = conn.execute(text("SELECT id FROM users WHERE email=:e"), {"e": ADMIN_EMAIL}).fetchone()
        if not existing:
            hashed = bcrypt.hashpw(ADMIN_PASSWORD.encode(), bcrypt.gensalt()).decode()
            conn.execute(text(
                "INSERT INTO users (email, password_hash, name, is_admin) VALUES (:e,:h,:n,:a)"
            ), {"e": ADMIN_EMAIL, "h": hashed, "n": "Administrador", "a": True})
        upsert_master_user(conn)
        conn.commit()


@asynccontextmanager
async def lifespan(app):
    init_db()
    _load_promo_floor_from_db()
    yield


app = FastAPI(lifespan=lifespan)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


def db_fetchone(query, params=None):
    with engine.connect() as conn:
        result = conn.execute(text(query), params or {})
        row = result.fetchone()
        return dict(row._mapping) if row else None


def db_fetchall(query, params=None):
    with engine.connect() as conn:
        result = conn.execute(text(query), params or {})
        return [dict(r._mapping) for r in result.fetchall()]


def db_execute(query, params=None):
    with engine.connect() as conn:
        conn.execute(text(query), params or {})
        conn.commit()


def db_fetch_order_snapshots(account_id: int, date_from: str, date_to: str) -> list:
    rows = db_fetchall("""
        SELECT payload FROM order_snapshot_cache
        WHERE account_id=:aid AND paid_date BETWEEN :df AND :dt
        ORDER BY paid_date DESC, paid_time DESC
    """, {"aid": account_id, "df": date_from, "dt": date_to})
    orders = []
    for row in rows:
        payload = row["payload"]
        if isinstance(payload, str):
            payload = json.loads(payload)
        orders.append(payload)
    return orders


def db_save_order_snapshots(account_id: int, orders: list, details_complete: bool):
    if not orders:
        return
    params_list = [
        {
            "aid": account_id,
            "oid": str(order["id"]),
            "paid_date": order["fecha"],
            "paid_time": order.get("hora") or "",
            "payload": json.dumps(order),
            "complete": details_complete,
        }
        for order in orders
    ]
    # Una sola conexión + executemany en vez de N conexiones.
    with engine.connect() as conn:
        conn.execute(text("""
            INSERT INTO order_snapshot_cache
                (account_id, order_id, paid_date, paid_time, payload, details_complete, updated_at)
            VALUES (:aid, :oid, :paid_date, :paid_time, CAST(:payload AS JSONB), :complete, NOW())
            ON CONFLICT (account_id, order_id) DO UPDATE SET
                paid_date=EXCLUDED.paid_date,
                paid_time=EXCLUDED.paid_time,
                payload=EXCLUDED.payload,
                details_complete=EXCLUDED.details_complete,
                updated_at=NOW()
        """), params_list)
        conn.commit()


def db_replace_order_snapshots_for_range(account_id: int, date_from: str, date_to: str, orders: list, details_complete: bool):
    """Replace a fully-refreshed range atomically so old partial cache rows cannot survive."""
    params_list = [
        {
            "aid": account_id,
            "oid": str(order["id"]),
            "paid_date": order["fecha"],
            "paid_time": order.get("hora") or "",
            "payload": json.dumps(order),
            "complete": details_complete,
        }
        for order in orders
    ]
    with engine.connect() as conn:
        conn.execute(
            text(
                "DELETE FROM order_snapshot_cache"
                " WHERE account_id=:aid AND paid_date BETWEEN :df AND :dt"
            ),
            {"aid": account_id, "df": date_from, "dt": date_to},
        )
        if params_list:
            conn.execute(text("""
                INSERT INTO order_snapshot_cache
                    (account_id, order_id, paid_date, paid_time, payload, details_complete, updated_at)
                VALUES (:aid, :oid, :paid_date, :paid_time, CAST(:payload AS JSONB), :complete, NOW())
                ON CONFLICT (account_id, order_id) DO UPDATE SET
                    paid_date=EXCLUDED.paid_date,
                    paid_time=EXCLUDED.paid_time,
                    payload=EXCLUDED.payload,
                    details_complete=EXCLUDED.details_complete,
                    updated_at=NOW()
            """), params_list)
        conn.commit()


def build_dashboard_payload(orders: list, details_complete: bool):
    orders = sorted(orders, key=lambda x: (x.get("fecha") or "", x.get("hora") or ""), reverse=True)


    daily: dict = {}
    products: dict = {}
    for order in orders:
        if order.get("fecha") and order.get("estado") == "paid":
            f = order["fecha"]
            daily.setdefault(f, {"ventas": 0, "ingresos": 0, "ganancia": 0})
            daily[f]["ventas"] += 1
            daily[f]["ingresos"] += order.get("monto", 0)
            daily[f]["ganancia"] += order.get("ganancia", 0)
            for item in order.get("items", []):
                t = item.get("sku") or item.get("titulo", "Sin título")
                products.setdefault(t, {"cantidad": 0, "ingresos": 0})
                products[t]["cantidad"] += item.get("cantidad", 1)
                products[t]["ingresos"] += item.get("monto", 0)

    top = sorted(products.items(), key=lambda x: x[1]["ingresos"], reverse=True)[:5]
    return {
        "orders": orders,
        "daily": dict(sorted(daily.items())),
        "top_products": [{"nombre": k, **v} for k, v in top],
        "details_complete": details_complete,
        "ultima_actualizacion": datetime.utcnow().isoformat(),
    }


def get_session_user_id(request: Request) -> Optional[int]:
    token = request.cookies.get("session")
    if not token:
        return None
    try:
        return serializer.loads(token, max_age=86400 * 7)
    except BadSignature:
        return None


def get_user(user_id: int):
    return db_fetchone("SELECT * FROM users WHERE id=:id", {"id": user_id})


def is_master_user(user: Optional[dict]) -> bool:
    return bool(user and user.get("is_master"))


def is_admin_user(user: Optional[dict]) -> bool:
    return bool(user and (user.get("is_admin") or user.get("is_master")))


async def refresh_ml_token(account_id: int) -> Optional[str]:
    acc = db_fetchone("SELECT * FROM ml_accounts WHERE id=:id", {"id": account_id})
    if not acc:
        return None
    # Si los tokens están vacíos (cuenta desconectada/soft delete), no
    # hay forma de refrescarlos — el usuario tiene que reconectar OAuth.
    if not acc.get("refresh_token") or not acc.get("access_token"):
        return None
    if acc["expires_at"] and datetime.utcnow() < acc["expires_at"] - timedelta(minutes=5):
        return acc["access_token"]
    async with httpx.AsyncClient() as client:
        resp = await client.post(ML_TOKEN_URL, data={
            "grant_type": "refresh_token",
            "client_id": ML_CLIENT_ID,
            "client_secret": ML_CLIENT_SECRET,
            "refresh_token": acc["refresh_token"],
        })
        if resp.status_code != 200:
            return None
        tokens = resp.json()
        new_expires = datetime.utcnow() + timedelta(seconds=tokens["expires_in"])
        db_execute(
            "UPDATE ml_accounts SET access_token=:at, refresh_token=:rt, expires_at=:ea WHERE id=:id",
            {"at": tokens["access_token"], "rt": tokens.get("refresh_token", acc["refresh_token"]),
             "ea": new_expires, "id": account_id}
        )
        return tokens["access_token"]


def db_get_cached_shipping(shipping_ids: list) -> dict:
    if not shipping_ids:
        return {}
    result = {}
    chunk = 500
    for i in range(0, len(shipping_ids), chunk):
        batch = shipping_ids[i:i+chunk]
        placeholders = ",".join(f":id{j}" for j in range(len(batch)))
        params = {f"id{j}": sid for j, sid in enumerate(batch)}
        rows = db_fetchall(
            f"SELECT shipping_id, cost, buyer_cost, bonificacion, list_cost, logistic_type"
            f" FROM shipment_cost_cache WHERE shipping_id IN ({placeholders})",
            params,
        )
        for row in rows:
            if row["buyer_cost"] is not None and row["bonificacion"] is not None:
                result[row["shipping_id"]] = {
                    "seller": float(row["cost"]),
                    "buyer": float(row["buyer_cost"]),
                    "bonificacion": float(row["bonificacion"]),
                    "list_cost": float(row["list_cost"]) if row["list_cost"] is not None else 0.0,
                    "logistic_type": row["logistic_type"] or "",
                }
    return result


def db_save_shipping_costs(costs: dict):
    if not costs:
        return
    params_list = [
        {
            "sid": sid,
            "cost": float(c.get("seller", 0)),
            "buyer_cost": float(c.get("buyer", 0)),
            "bonif": float(c.get("bonificacion", 0)),
            "list_cost": float(c.get("list_cost", 0)),
            "lt": c.get("logistic_type", "") or "",
        }
        for sid, c in costs.items()
    ]
    # Una sola conexión + executemany en vez de N conexiones.
    with engine.connect() as conn:
        conn.execute(text(
            "INSERT INTO shipment_cost_cache (shipping_id, cost, buyer_cost, bonificacion, list_cost, logistic_type)"
            " VALUES (:sid, :cost, :buyer_cost, :bonif, :list_cost, :lt)"
            " ON CONFLICT (shipping_id) DO UPDATE SET"
            " cost = EXCLUDED.cost, buyer_cost = EXCLUDED.buyer_cost,"
            " bonificacion = EXCLUDED.bonificacion, list_cost = EXCLUDED.list_cost,"
            " logistic_type = EXCLUDED.logistic_type"
        ), params_list)
        conn.commit()


def amount_value(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, dict):
        for key in ("amount", "value", "cost"):
            if value.get(key) is not None:
                return amount_value(value.get(key))
        return 0.0
    if isinstance(value, list):
        return sum(amount_value(v) for v in value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def discount_total(discounts) -> float:
    total = 0.0
    for discount in discounts or []:
        total += amount_value(discount.get("promoted_amount"))
    return total


def order_ids_from_sales_info(row: dict) -> list:
    ids = []
    for sale in row.get("sales_info") or []:
        oid = sale.get("order_id")
        if oid is not None:
            ids.append(str(oid))
    if row.get("order_id") is not None:
        ids.append(str(row.get("order_id")))
    return ids


def billing_amount(row: dict, key: str, default=0.0) -> float:
    charge = row.get("charge_info") or {}
    return amount_value(charge.get(key, row.get(key, default)))


def billing_text(row: dict, key: str) -> str:
    charge = row.get("charge_info") or {}
    return str(charge.get(key, row.get(key, "")) or "").lower()


def apply_billing_row(target: dict, row: dict):
    charge = row.get("charge_info") or {}
    discount = row.get("discount_info") or {}
    shipping = row.get("shipping_info") or {}
    detail_type = billing_text(row, "detail_type").upper()
    sub_type = billing_text(row, "detail_sub_type").upper()
    concept_type = billing_text(row, "concept_type").upper()
    transaction_detail = billing_text(row, "transaction_detail")
    detail_amount = abs(billing_amount(row, "detail_amount"))
    amount_without_discount = abs(amount_value(discount.get("charge_amount_without_discount")))
    discount_amount = abs(amount_value(discount.get("discount_amount")))

    # Una fila BONUS NUNCA se suma a "envio" — sólo a "bonificacion".
    # Si no excluimos BONUS de is_shipping, su detail_amount se sumaba
    # también a envio y duplicaba el costo (ej. #2000016639025652).
    is_bonus = detail_type == "BONUS"
    is_shipping_concept = (
        concept_type == "SHIPPING"
        or sub_type in {"CXD", "CFF", "BXD", "BFF"}
        or "envío" in transaction_detail
        or "envio" in transaction_detail
        or "shipping" in transaction_detail
    )
    is_shipping = (not is_bonus) and is_shipping_concept
    is_sale_charge = (
        (not is_shipping) and (not is_bonus)
        and detail_type == "CHARGE"
        and (sub_type.startswith("CV") or "cargo por venta" in transaction_detail or "cargo por vender" in transaction_detail)
    )
    if is_sale_charge:
        target["comision"] += detail_amount
        target["has_comision"] = True
    if is_shipping:
        target["envio"] += amount_without_discount or detail_amount
        target["bonificacion"] += discount_amount
        target["has_shipping"] = True
        receiver_cost = amount_value(shipping.get("receiver_shipping_cost"))
        if receiver_cost > 0:
            target["ingreso_envio"] = max(target["ingreso_envio"], receiver_cost)
    elif is_bonus and is_shipping_concept:
        target["bonificacion"] += detail_amount
        target["has_shipping"] = True


def period_key_from_ml_date(value: str) -> str:
    if not value:
        return ""
    return f"{value[:7]}-01"


def flex_order_id(row: dict) -> str:
    shipping = row.get("shipping_info") or {}
    order = shipping.get("order") or {}
    oid = order.get("order_id") or row.get("order_id")
    return str(oid) if oid is not None else ""


def apply_flex_billing_row(target: dict, row: dict):
    charge = row.get("charge_info") or {}
    shipping = row.get("shipping_info") or {}
    detail_type = str(charge.get("detail_type") or row.get("detail_type") or "").upper()
    transaction_detail = str(charge.get("transaction_detail") or row.get("transaction_detail") or "").lower()
    amount = abs(amount_value(charge.get("detail_amount", row.get("detail_amount"))))
    receiver_cost = amount_value(shipping.get("receiver_shipping_cost"))
    if receiver_cost > 0:
        target["ingreso_envio"] = max(target["ingreso_envio"], receiver_cost)
    if not amount:
        return
    is_cancellation = "anulaci" in transaction_detail or "cancellation" in transaction_detail
    is_bonus = detail_type == "BONUS" or ("bonific" in transaction_detail and not is_cancellation)
    if is_bonus:
        target["bonificacion"] += amount
        target["has_flex"] = True
    elif is_cancellation or detail_type == "CHARGE":
        target["envio"] += amount
        target["has_flex_charge"] = True


async def fetch_flex_billing_by_order(client, headers, order_periods: dict) -> dict:
    if not order_periods:
        return {}
    parsed = {}
    period_map = {}
    for oid, key in order_periods.items():
        if key:
            period_map.setdefault(key, []).append(str(oid))

    document_types = ("BILL", "CREDIT_NOTE", "DEBIT_NOTE")
    for key, ids in period_map.items():
        unique_ids = list(dict.fromkeys(ids))
        for i in range(0, len(unique_ids), 100):
            batch = unique_ids[i:i+100]
            for document_type in document_types:
                params = {
                    "document_type": document_type,
                    "order_ids": ",".join(batch),
                    "limit": 1000,
                }
                r = await client.get(
                    f"{ML_API_URL}/billing/integration/periods/key/{key}/group/ML/flex/details",
                    headers=headers,
                    params=params,
                )
                if r.status_code not in (200, 206):
                    continue
                for row in r.json().get("results", []):
                    oid = flex_order_id(row)
                    if not oid:
                        continue
                    parsed.setdefault(oid, {
                        "envio": 0.0, "bonificacion": 0.0, "ingreso_envio": 0.0,
                        "has_flex": False, "has_flex_charge": False,
                    })
                    apply_flex_billing_row(parsed[oid], row)
    return {
        oid: {
            k: (round(v, 2) if isinstance(v, (int, float)) else v)
            for k, v in values.items()
        }
        for oid, values in parsed.items()
    }


async def fetch_billing_by_order(client, headers, seller_id, order_ids: list) -> dict:
    if not order_ids:
        return {}
    parsed = {}
    unique_ids = list(dict.fromkeys(str(oid) for oid in order_ids if oid))
    for i in range(0, len(unique_ids), 100):
        batch = unique_ids[i:i+100]
        params = {
            "order_ids": ",".join(batch),
            "seller_id": seller_id,
            "limit": 150,
        }
        r = await client.get(f"{ML_API_URL}/billing/integration/group/ML/order/details", headers=headers, params=params)
        if r.status_code not in (200, 206):
            continue
        data = r.json()
        for row in data.get("results", []):
            for oid in order_ids_from_sales_info(row):
                parsed.setdefault(oid, {
                    "comision": 0.0, "envio": 0.0, "bonificacion": 0.0, "ingreso_envio": 0.0,
                    "has_comision": False, "has_shipping": False,
                })
                apply_billing_row(parsed[oid], row)
    return {
        oid: {
            k: (round(v, 2) if isinstance(v, (int, float)) else v)
            for k, v in values.items()
        }
        for oid, values in parsed.items()
    }


def to_ar(dt_str: str) -> tuple:
    """Convierte datetime ISO de ML a fecha/hora Argentina (UTC-3).
    ML puede devolver -03:00 (ya local) o +00:00 (UTC) según el endpoint."""
    if not dt_str:
        return "", ""
    try:
        dt_naive = datetime.strptime(dt_str[:19], "%Y-%m-%dT%H:%M:%S")
        # Extraer offset saltando milisegundos opcionales: "...T12:16:00.000-03:00"
        tail = dt_str[19:].lstrip(".0123456789")  # quita ".000" si existe
        offset_min = 0
        if len(tail) >= 6 and tail[0] in ("+", "-"):
            sign = -1 if tail[0] == "-" else 1
            h, m = int(tail[1:3]), int(tail[4:6])
            offset_min = sign * (h * 60 + m)
        # dt_naive está en el timezone del offset → convertir a UTC → luego a AR (UTC-3)
        dt_ar = dt_naive - timedelta(minutes=offset_min) - timedelta(hours=3)
        return dt_ar.strftime("%Y-%m-%d"), dt_ar.strftime("%H:%M")
    except Exception:
        return dt_str[:10], ""


async def get_shipping_cost(client, shipping_id, headers) -> dict:
    """Lee el desglose de envío combinando /shipments/{id} y /shipments/{id}/costs.

    /costs sólo expone el costo NETO del vendedor (0 cuando el comprador paga),
    así que para que el panel coincida con lo que muestra ML para cada venta
    usamos /shipments para Ing./Costo Envío y /costs únicamente para la
    bonificación que ML otorga al vendedor.

    Reglas:
      - Flex (home_delivery): Ing. = cost del comprador, Costo = 0,
        Bonif. = senders[0].compensation o, si no viene, list_cost − cost.
      - Colecta/Full con comprador pagando envío (cost > 0):
        Ing. = Costo = cost (ML descuenta al vendedor el mismo monto).
      - Colecta/Full con Envío Gratis (cost = 0):
        Ing. = 0, Costo = senders[0].cost (descuento ya aplicado) o,
        si /costs no responde, 50% del list_cost.
    """
    # Resiliencia: si ML tira timeout / corta la conexión en un shipment, NO
    # queremos (a) que explote toda la request con un 500, ni (b) que el envío
    # quede en 0 (costo/ingreso de envío mal → totales mal en rangos grandes).
    # Con Semaphore(80) pegándole fuerte a ML, los timeouts esporádicos son
    # esperables, así que reintentamos con backoff (igual que el order search)
    # antes de rendirnos. Si tras los reintentos sigue fallando, devolvemos
    # None para NO cachear un 0 erróneo y reintentar en el próximo refresh.
    EMPTY = {"seller": 0.0, "buyer": 0.0, "bonificacion": 0.0}
    costs_headers = {**headers, "X-Costs-New": "true", "x-format-new": "true"}

    async def _ship_get(url, hdrs):
        """GET con reintentos sobre timeouts/429/5xx. Devuelve la respuesta,
        o None si tras 4 intentos no hubo un resultado utilizable."""
        for attempt in range(4):
            try:
                rp = await client.get(url, headers=hdrs)
            except Exception:
                if attempt < 3:
                    await asyncio.sleep(0.4 * (2 ** attempt))
                    continue
                return None
            if rp.status_code == 200:
                return rp
            # 404/410: el envío no existe / no es recuperable → respuesta válida.
            if rp.status_code in (429, 500, 502, 503, 504) and attempt < 3:
                await asyncio.sleep(0.4 * (2 ** attempt))
                continue
            return rp
        return None

    r_ship, r_costs = await asyncio.gather(
        _ship_get(f"{ML_API_URL}/shipments/{shipping_id}", headers),
        _ship_get(f"{ML_API_URL}/shipments/{shipping_id}/costs", costs_headers),
    )

    # /shipments falló de forma transitoria tras reintentos: no cachear (None).
    if r_ship is None:
        return None
    if r_ship.status_code != 200:
        return dict(EMPTY)

    ship = r_ship.json()
    so = ship.get("shipping_option") or {}
    logistic = ship.get("logistic") or {}
    logistic_type = ship.get("logistic_type") or logistic.get("type") or ""
    mode = ship.get("mode") or logistic.get("mode") or ""
    option_name = str(so.get("name") or so.get("shipping_method_type") or "").lower()
    cost = amount_value(so.get("cost"))
    list_cost = amount_value(so.get("list_cost"))
    base_cost = amount_value(so.get("base_cost") or ship.get("base_cost"))
    option_discount = so.get("discount") or {}
    option_discount_amount = amount_value(option_discount.get("promoted_amount"))
    option_discount_rate = amount_value(option_discount.get("rate"))
    is_flex = (
        logistic_type in {"self_service", "home_delivery"}
        or mode == "self_service"
        or "flex" in option_name
    )

    # Extraer datos del endpoint /costs si está disponible
    costs_sender_cost = 0.0
    costs_gross_amount = 0.0
    receiver_cost = 0.0
    sender_discount = 0.0
    sender_save = 0.0
    compensation = 0.0
    costs_responded = r_costs is not None and r_costs.status_code == 200
    if costs_responded:
        cd = r_costs.json()
        costs_gross_amount = amount_value(cd.get("gross_amount"))
        receiver = cd.get("receiver") or {}
        receiver_cost = amount_value(receiver.get("cost"))
        senders = cd.get("senders") or []
        if senders:
            s0 = senders[0]
            costs_sender_cost = amount_value(s0.get("cost"))
            sender_discount = discount_total(s0.get("discounts"))
            sender_save = amount_value(s0.get("save"))
            compensation = amount_value(s0.get("compensation"))

    buyer_cost_from_costs = receiver_cost if receiver_cost > 0 else cost

    def gross_candidate(net_cost: float, bonif: float = 0.0) -> float:
        if net_cost > 0 and bonif > 0:
            return net_cost + bonif
        if costs_gross_amount > 0:
            return costs_gross_amount
        if list_cost > 0:
            return list_cost
        if base_cost > 0:
            return base_cost
        if option_discount_amount > 0:
            return option_discount_amount
        return net_cost or cost

    def derive_bonif(net_cost: float, gross_cost: float) -> float:
        if compensation > 0:
            return compensation
        if sender_discount > 0:
            return sender_discount
        if option_discount_amount > 0:
            return option_discount_amount
        if gross_cost > net_cost > 0:
            return gross_cost - net_cost
        return 0.0

    # ── Fórmula unificada basada en lo que devuelve /shipments/{id}/costs ──
    # Trusteamos SÓLO los campos explícitos del API. La inferencia implícita
    # via gross_amount no es confiable porque incluye items adicionales
    # (fees, seguros, etc.) que inflan la bonificación.
    #
    #   Ing. Envío  = receiver.cost (lo que paga el comprador)
    #   Costo Envío:
    #     · Colecta / Full → senders.cost + receiver.cost (gross "Cargo por
    #       Envíos" que muestra el panel cuando el comprador aporta parte).
    #     · Flex u otros  → senders.cost directo (el vendedor no paga por ML).
    #   Bonificación = max(compensation, discounts.promoted_amount,
    #                      option.discount.promoted_amount).
    is_colecta = logistic_type in {"xd_drop_off", "drop_off", "cross_docking", "fulfillment"}

    buyer_cost = receiver_cost if receiver_cost > 0 else cost
    if is_colecta:
        seller_cost = costs_sender_cost + receiver_cost
    else:
        seller_cost = costs_sender_cost

    # senders[0].discounts[].promoted_amount tiene semántica distinta según
    # el caso: cuando el vendedor PAGA el envío (sender.cost > 0) es el
    # descuento aplicado sobre la tarifa de lista (ej. 50% off colecta),
    # NO una bonificación al vendedor. Cuando el vendedor NO paga
    # (sender.cost = 0, env. gratis bonificado), discounts[] sí representa
    # la bonificación de ML que cubre el envío.
    # Bonificación desde campos EXPLÍCITOS del API. El campo 'save' tiene
    # doble semántica cuando sender.cost > 0:
    #   - Flex paga: save << sender.cost (ej. $649 con sender ~$7.139)
    #     → save ES la bonificación de ML
    #   - Colecta paga: save ≈ sender.cost (ej. $7.470 = $7.470)
    #     → save es el "descuento de tarifa", NO bonificación
    # Heurística: usamos save sólo si es MENOR que sender.cost
    # (lo que indica que es un bonus real, no un offset de tarifa).
    if costs_sender_cost > 0:
        save_as_bonif = sender_save if 0 < sender_save < costs_sender_cost else 0.0
        # discount_total similar: si la suma equivale al sender.cost es
        # descuento de tarifa, no bonif.
        disc_as_bonif = sender_discount if 0 < sender_discount < costs_sender_cost else 0.0
        bonificacion = max(compensation, save_as_bonif, disc_as_bonif, option_discount_amount)
    else:
        # sender.cost = 0 → flex env gratis o colecta env gratis cubierta por ML.
        # Todos los campos representan bonificaciones reales aquí.
        bonificacion = max(compensation, sender_save, sender_discount, option_discount_amount)

    # Fallback al 50% del list_cost SÓLO cuando /costs no respondió
    # (ej. shipments cancelados sin info). Si /costs respondió y devolvió
    # sender.cost = 0, confiamos en eso: ML genuinamente no le cobró
    # envío al vendedor (caso #2000016669490932, panel sin línea de envío).
    if not costs_responded and seller_cost == 0 and bonificacion == 0 and buyer_cost == 0 and is_colecta:
        if option_discount_rate and list_cost > 0:
            seller_cost = list_cost * (1.0 - option_discount_rate)
        elif list_cost > 0 or base_cost > 0:
            seller_cost = (list_cost or base_cost) * 0.5

    return {
        "seller": round(seller_cost, 2),
        "buyer": round(buyer_cost, 2),
        "bonificacion": round(bonificacion, 2),
        "list_cost": round(list_cost, 2) if list_cost else 0.0,
        "logistic_type": logistic_type,
    }


# ── Auth ────────────────────────────────────────────────────────

# Ruta de cada pestaña (para mandar al usuario a la primera que tenga permitida).
PAGE_ROUTES = {
    "dashboard": "/dashboard",
    "costos": "/costos",
    "envios_flex": "/costos/envios-flex",
    "descuentos": "/descuentos",
    "ranking": "/ranking",
    "etiquetas": "/etiquetas",
}


def _landing_path(user):
    """Primera pestaña (en orden de PAGES) que el usuario puede ver, o None si no
    tiene ninguna. Evita el 403 cuando no tiene permiso de Dashboard."""
    perms = user_permissions(user)
    for key, _label in PAGES:
        if key in perms:
            return PAGE_ROUTES.get(key)
    return None


def _page_redirect(user, page):
    """Para handlers de PÁGINA (HTML): si el usuario no tiene la pestaña, devuelve
    un RedirectResponse a la primera que sí tenga (en vez de cortar con 403). Si
    no tiene ninguna, recién ahí 403."""
    if page in user_permissions(user):
        return None
    landing = _landing_path(user)
    if landing and landing != PAGE_ROUTES.get(page):
        return RedirectResponse(landing)
    raise HTTPException(403, "No tenés permiso para acceder a esta sección")


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    uid = get_session_user_id(request)
    if uid:
        return RedirectResponse(_landing_path(get_user(uid)) or "/dashboard")
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/", response_class=HTMLResponse)
async def do_login(request: Request, email: str = Form(...), password: str = Form(...)):
    login_email = email.lower().strip()
    master_email, _, _ = master_credentials()
    if master_email and login_email == master_email:
        upsert_master_user()
    user = db_fetchone("SELECT * FROM users WHERE email=:e", {"e": login_email})
    if not user or not user.get("password_hash") or not bcrypt.checkpw(password.encode(), user["password_hash"].encode()):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Email o contraseña incorrectos"})
    session_token = serializer.dumps(user["id"])
    response = RedirectResponse(_landing_path(user) or "/dashboard", status_code=303)
    response.set_cookie("session", session_token, max_age=86400 * 7, httponly=True)
    return response


@app.get("/logout")
async def logout():
    r = RedirectResponse("/")
    r.delete_cookie("session")
    return r


# ── Dashboard ───────────────────────────────────────────────────

def _accounts_owner_id(user: Optional[dict], user_id: int) -> int:
    """Dueño de las cuentas de ML que ve un usuario: si es colaborador (tiene
    owner_id), las del dueño; si no, las propias."""
    if user and user.get("owner_id"):
        return user["owner_id"]
    return user_id


def _account_for_user(account_id: int, user_id: int):
    """Devuelve la cuenta ML que el usuario (o su dueño) puede operar."""
    user = get_user(user_id)
    if not user:
        return None
    owner = _accounts_owner_id(user, user_id)
    return db_fetchone(
        "SELECT * FROM ml_accounts WHERE id=:id AND user_id=:uid",
        {"id": account_id, "uid": owner},
    )


def _cost_account_id_for(user: dict, account_id: int) -> int:
    """Admin users share costs/flex across all their accounts; master uses per-account costs."""
    if is_master_user(user):
        return account_id
    owner = _accounts_owner_id(user, user["id"])
    row = db_fetchone(
        "SELECT MIN(id) as min_id FROM ml_accounts WHERE user_id=:uid",
        {"uid": owner}
    )
    return (row["min_id"] if row and row["min_id"] else account_id)


def _invalidate_all_user_accounts_cache(user: dict, account_id: int):
    """For admin users (shared costs), invalidate cache for all their accounts. Master: just the one."""
    if is_master_user(user):
        invalidate_orders_cache_for_account(account_id)
        return
    owner = _accounts_owner_id(user, user["id"])
    accs = db_fetchall("SELECT id FROM ml_accounts WHERE user_id=:uid", {"uid": owner})
    for a in accs:
        invalidate_orders_cache_for_account(a["id"])


def get_visible_accounts(user_id: int, user: dict) -> list:
    """Cuentas ML que ve el usuario: las propias, o las del dueño si es
    colaborador. Los permisos solo controlan pestañas."""
    owner = _accounts_owner_id(user, user_id)
    return db_fetchall(
        "SELECT id, nickname, ml_user_id FROM ml_accounts"
        " WHERE user_id=:uid AND COALESCE(access_token, '') <> ''"
        " ORDER BY id",
        {"uid": owner},
    )


async def fetch_ml_nickname(client: httpx.AsyncClient, token: str, ml_user_id: str) -> Optional[str]:
    headers = {"Authorization": f"Bearer {token}"}
    for path in ("/users/me", f"/users/{ml_user_id}"):
        try:
            resp = await client.get(f"{ML_API_URL}{path}", headers=headers)
            if resp.status_code != 200:
                continue
            data = resp.json()
            nickname = (data.get("nickname") or "").strip()
            if nickname:
                return nickname
        except Exception:
            continue
    return None


async def refresh_visible_account_nicknames(accounts: list) -> list:
    """Completa apodos de cuentas viejas que quedaron guardadas con el ID numerico."""
    pending = [
        acc for acc in accounts
        if not str(acc.get("nickname") or "").strip()
        or str(acc.get("nickname") or "").strip() == str(acc.get("ml_user_id") or "").strip()
    ]
    if not pending:
        return accounts

    sem = asyncio.Semaphore(4)

    async with httpx.AsyncClient(timeout=5) as client:
        async def refresh_one(acc: dict) -> None:
            current = str(acc.get("nickname") or "").strip()
            ml_user_id = str(acc.get("ml_user_id") or "").strip()
            async with sem:
                token = await refresh_ml_token(acc["id"])
                if not token:
                    return
                nickname = await fetch_ml_nickname(client, token, ml_user_id)
            if not nickname or nickname == current:
                return
            db_execute(
                "UPDATE ml_accounts SET nickname=:nick WHERE id=:id",
                {"nick": nickname, "id": acc["id"]},
            )
            acc["nickname"] = nickname

        await asyncio.gather(*(refresh_one(acc) for acc in pending))
    return accounts


def can_access_account(account_id: int, user_id: int, user: dict) -> bool:
    """El usuario solo puede operar cuentas ML propias."""
    acc = db_fetchone(
        "SELECT id FROM ml_accounts WHERE id=:id AND user_id=:uid",
        {"id": account_id, "uid": user_id},
    )
    return acc is not None


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    _r = _page_redirect(user, "dashboard")
    if _r:
        return _r
    accounts = get_visible_accounts(user_id, user)
    accounts = await refresh_visible_account_nicknames(accounts)
    return templates.TemplateResponse("dashboard.html", {
        "request": request, "user": user, "accounts": accounts,
        "perms": user_permissions(user),
    })


# ── ML OAuth ────────────────────────────────────────────────────

@app.get("/ml/connect")
async def ml_connect(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    state = secrets.token_urlsafe(16)
    url = (f"{ML_AUTH_URL}?response_type=code&client_id={ML_CLIENT_ID}"
           f"&redirect_uri={APP_URL}/ml/callback&state={state}")
    r = RedirectResponse(url)
    r.set_cookie("oauth_state", state, max_age=600, httponly=True)
    return r


@app.get("/ml/callback")
async def ml_callback(request: Request, code: str, state: str):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    if request.cookies.get("oauth_state") != state:
        raise HTTPException(400, "Estado OAuth inválido")
    async with httpx.AsyncClient() as client:
        resp = await client.post(ML_TOKEN_URL, data={
            "grant_type": "authorization_code",
            "client_id": ML_CLIENT_ID,
            "client_secret": ML_CLIENT_SECRET,
            "code": code,
            "redirect_uri": f"{APP_URL}/ml/callback",
        })
        if resp.status_code != 200:
            raise HTTPException(400, f"Error ML: {resp.text}")
        tokens = resp.json()
        ml_user_id = str(tokens["user_id"])
        nickname = tokens.get("nickname") or await fetch_ml_nickname(client, tokens["access_token"], ml_user_id) or ml_user_id
    expires_at = datetime.utcnow() + timedelta(seconds=tokens["expires_in"])
    db_execute("""
        INSERT INTO ml_accounts (user_id, ml_user_id, nickname, access_token, refresh_token, expires_at)
        VALUES (:uid,:mlid,:nick,:at,:rt,:ea)
        ON CONFLICT (user_id, ml_user_id) DO UPDATE
        SET access_token=:at, refresh_token=:rt, expires_at=:ea, nickname=:nick
    """, {"uid": user_id, "mlid": ml_user_id, "nick": nickname,
          "at": tokens["access_token"], "rt": tokens["refresh_token"], "ea": expires_at})
    r = RedirectResponse("/dashboard", status_code=303)
    r.delete_cookie("oauth_state")
    return r


@app.post("/ml/disconnect/{account_id}")
async def ml_disconnect(request: Request, account_id: int):
    """Soft disconnect: limpia tokens pero conserva el registro y todos
    los datos asociados (costos, tarifas flex, descuentos, snapshots).
    Al reconectar (OAuth callback) hace UPSERT sobre el mismo account_id
    y todo sigue funcionando sin tener que volver a cargar nada."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    # Vaciar los tokens pero NO eliminar el registro
    db_execute(
        "UPDATE ml_accounts SET access_token='', refresh_token='',"
        " expires_at = NOW() - INTERVAL '1 day'"
        " WHERE id=:id AND user_id=:uid",
        {"id": account_id, "uid": user_id},
    )
    return RedirectResponse("/dashboard", status_code=303)


@app.post("/ml/delete/{account_id}")
async def ml_delete_account(request: Request, account_id: int):
    """Eliminación DEFINITIVA del registro y todos sus datos asociados
    (costos, tarifas flex, descuentos, snapshots de órdenes). Sólo para
    casos en los que el admin realmente quiere borrar todo."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    db_execute(
        "DELETE FROM ml_accounts WHERE id=:id AND user_id=:uid",
        {"id": account_id, "uid": user_id},
    )
    return RedirectResponse("/dashboard", status_code=303)


# ── Costos de mercadería (CMV) ──────────────────────────────────

def db_get_product_costs(account_id: int) -> dict:
    """Devuelve {sku_uppercase: [entries sorted by valid_from asc]}.
    Cada entry: {cost (sin IVA), iva_rate, valid_from (str ISO)}.
    """
    rows = db_fetchall(
        "SELECT sku, cost, iva_rate, valid_from FROM product_costs"
        " WHERE account_id = :aid ORDER BY sku, valid_from",
        {"aid": account_id},
    )
    out: dict = {}
    for r in rows:
        if not r.get("sku"):
            continue
        key = (r["sku"] or "").strip().upper()
        vf = r["valid_from"]
        out.setdefault(key, []).append({
            "cost": float(r["cost"]),
            "iva_rate": float(r["iva_rate"] if r["iva_rate"] is not None else 21),
            "valid_from": vf.isoformat() if hasattr(vf, "isoformat") else str(vf),
        })
    return out


import re as _re_costs
_CUOTA_SUFFIX_RE = _re_costs.compile(r"^(\d+C|CE|SI|SIN|\d+CE)$")


def _sku_fallbacks(sku: str):
    """Genera variantes del SKU stripeando sufijos típicos de cuotas
    (-3C, -6C, -12C, -CE, -SI, etc.) un segmento a la vez. Permite que
    SOP78-446-3C herede el costo de SOP78-446 si el primero no está cargado.
    """
    parts = sku.split("-")
    yielded = set()
    while len(parts) > 1 and _CUOTA_SUFFIX_RE.match(parts[-1]):
        parts = parts[:-1]
        candidate = "-".join(parts)
        if candidate not in yielded:
            yielded.add(candidate)
            yield candidate


def find_cost_for_date(versioned: dict, sku: str, sale_date: str) -> dict:
    """Devuelve la entry de costo más reciente con valid_from <= sale_date.
    Si el SKU exacto no está cargado, intenta con SKUs base (stripeando
    sufijos como -3C, -CE) — útil para variantes por cuotas."""
    if not sku or not sale_date:
        return None
    sku_norm = sku.strip().upper()
    entries = versioned.get(sku_norm)
    if not entries:
        for fallback in _sku_fallbacks(sku_norm):
            entries = versioned.get(fallback)
            if entries:
                break
    if not entries:
        return None
    selected = None
    for e in entries:
        if e["valid_from"] <= sale_date:
            selected = e
        else:
            break
    return selected


def cost_with_iva(entry: dict) -> float:
    """Convierte el costo (sin IVA) al equivalente con IVA aplicado según la
    tasa cargada (21%, 10.5%, 0 = exento, etc). El resultado es comparable
    con monto/comisión/envío que ya vienen con IVA en ML."""
    if not entry:
        return 0.0
    cost = float(entry.get("cost") or 0)
    rate = float(entry.get("iva_rate") or 0)
    return cost * (1 + rate / 100.0)


def db_save_product_costs(account_id: int, items: list):
    """Inserta/actualiza lista de {sku, cost, iva_rate, valid_from}."""
    if not items:
        return 0
    params = []
    today_iso = datetime.utcnow().date().isoformat()
    for it in items:
        sku = (it.get("sku") or "").strip()
        if not sku:
            continue
        try:
            cost = float(it.get("cost") or 0)
        except (TypeError, ValueError):
            continue
        if cost <= 0:
            continue
        vf = it.get("valid_from") or today_iso
        if hasattr(vf, "isoformat"):
            vf = vf.isoformat()
        try:
            iva_rate = float(it.get("iva_rate") if it.get("iva_rate") is not None else 21)
        except (TypeError, ValueError):
            iva_rate = 21.0
        params.append({
            "aid": account_id, "sku": sku, "cost": cost,
            "iva_rate": iva_rate, "vf": vf,
        })
    if not params:
        return 0
    with engine.connect() as conn:
        conn.execute(text(
            "INSERT INTO product_costs (account_id, sku, cost, iva_rate, valid_from, updated_at)"
            " VALUES (:aid, :sku, :cost, :iva_rate, :vf, NOW())"
            " ON CONFLICT (account_id, sku, valid_from) DO UPDATE SET"
            " cost = EXCLUDED.cost,"
            " iva_rate = EXCLUDED.iva_rate,"
            " updated_at = NOW()"
        ), params)
        conn.commit()
    return len(params)


def _parse_iva_rate_cell(value) -> float:
    """Devuelve la tasa de IVA (21, 10.5, 0, etc.) desde una celda.
    Acepta numérico, '21%', '10,5', 'Exento', 'No', '21 %', etc.
    Default si no se reconoce: 21."""
    if value is None or value == "":
        return 21.0
    if isinstance(value, bool):
        return 21.0 if value else 0.0
    if isinstance(value, (int, float)):
        # Si el valor es 0 o positivo razonable, usarlo directo
        v = float(value)
        return v if 0 <= v <= 100 else 21.0
    s = str(value).strip().lower().replace("%", "").replace(",", ".").strip()
    if not s:
        return 21.0
    if s in {"exento", "no", "sin iva", "sin", "0", "no aplica", "n/a"}:
        return 0.0
    if s in {"si", "sí", "yes", "y", "incluido"}:
        return 0.0  # ya incluye, sin agregar más
    try:
        v = float(s)
        return v if 0 <= v <= 100 else 21.0
    except ValueError:
        return 21.0


def _parse_date_cell(value):
    """Devuelve un string ISO (YYYY-MM-DD) o None."""
    if value is None or value == "":
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.date().isoformat() if hasattr(value, "date") else value.isoformat()
        except Exception:
            pass
    s = str(value).strip()
    if not s:
        return None
    # Probar formatos comunes
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _find_col(header: list, keywords: list) -> Optional[int]:
    for i, n in enumerate(header):
        for k in keywords:
            if k in n:
                return i
    return None


def parse_excel_costs(content: bytes) -> list:
    """Devuelve [{sku, cost, iva_included, valid_from}, ...] desde xlsx."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = None
    data_start = 0
    for idx, row in enumerate(rows[:5]):
        if row is None:
            continue
        normalized = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("sku" in n for n in normalized):
            header = normalized
            data_start = idx + 1
            break
    if not header:
        return []
    sku_idx = _find_col(header, ["sku"])
    cost_idx = _find_col(header, ["costo", "cost", "precio"])
    fecha_idx = _find_col(header, ["fecha", "date", "vigencia", "desde"])
    iva_idx = _find_col(header, ["iva"])
    if sku_idx is None or cost_idx is None:
        return []
    items = []
    today_iso = datetime.utcnow().date().isoformat()
    for row in rows[data_start:]:
        if row is None or len(row) <= max(sku_idx, cost_idx):
            continue
        sku = row[sku_idx]
        cost = row[cost_idx]
        if sku is None or cost is None:
            continue
        fecha = _parse_date_cell(row[fecha_idx]) if fecha_idx is not None and len(row) > fecha_idx else None
        iva_rate = _parse_iva_rate_cell(row[iva_idx]) if iva_idx is not None and len(row) > iva_idx else 21.0
        items.append({
            "sku": str(sku).strip(),
            "cost": cost,
            "valid_from": fecha or today_iso,
            "iva_rate": iva_rate,
        })
    return items


def parse_csv_costs(content: bytes) -> list:
    """Devuelve [{sku, cost, iva_included, valid_from}, ...] desde csv."""
    import csv, io
    text_content = content.decode("utf-8-sig", errors="ignore")
    sample = text_content[:2048]
    sep = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text_content), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []
    header = [(c or "").strip().lower() for c in rows[0]]
    sku_idx = _find_col(header, ["sku"])
    cost_idx = _find_col(header, ["costo", "cost", "precio"])
    fecha_idx = _find_col(header, ["fecha", "date", "vigencia", "desde"])
    iva_idx = _find_col(header, ["iva"])
    if sku_idx is None or cost_idx is None:
        return []
    items = []
    today_iso = datetime.utcnow().date().isoformat()
    for row in rows[1:]:
        if len(row) <= max(sku_idx, cost_idx):
            continue
        sku = (row[sku_idx] or "").strip()
        raw_cost = (row[cost_idx] or "").strip()
        if sep == ";":
            raw_cost = raw_cost.replace(".", "").replace(",", ".")
        else:
            raw_cost = raw_cost.replace(",", "")
        if not sku or not raw_cost:
            continue
        try:
            cost = float(raw_cost)
        except ValueError:
            continue
        fecha = _parse_date_cell(row[fecha_idx]) if fecha_idx is not None and len(row) > fecha_idx else None
        iva_rate = _parse_iva_rate_cell(row[iva_idx]) if iva_idx is not None and len(row) > iva_idx else 21.0
        items.append({
            "sku": sku, "cost": cost,
            "valid_from": fecha or today_iso,
            "iva_rate": iva_rate,
        })
    return items


@app.get("/costos", response_class=HTMLResponse)
async def costos_page(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    _r = _page_redirect(user, "costos")
    if _r:
        return _r
    accounts = get_visible_accounts(user_id, user)
    accounts = await refresh_visible_account_nicknames(accounts)
    return templates.TemplateResponse("costos.html", {
        "request": request, "user": user, "accounts": accounts,
        "perms": user_permissions(user),
    })


@app.get("/api/costos/template")
async def api_costos_template(request: Request):
    """Genera un Excel modelo (plantilla) con las columnas esperadas
    y un par de filas de ejemplo para que el usuario lo complete."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Costos"

    headers = ["SKU", "Costo", "Fecha", "IVA"]
    ws.append(headers)
    # Estilizar header
    header_fill = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas de ejemplo
    examples = [
        ["SOP68-443", 5000.00, "2025-01-15", 21],
        ["SOP78-446", 8500.50, "2025-01-15", 10.5],
        ["SOP22G-44T", 3200.00, "2025-02-01", 0],   # Exento
    ]
    for row in examples:
        ws.append(row)

    # Ancho de columnas
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instr = [
        ["Plantilla de costos de mercadería"],
        [""],
        ["Columnas:"],
        ["SKU", "Identificador único del producto (coincide con el seller_sku de ML)."],
        ["Costo", "Costo SIN IVA (el sistema le aplica la tasa de IVA automáticamente)."],
        ["Fecha", "Fecha de vigencia desde. Formato YYYY-MM-DD o DD/MM/YYYY. Si la dejás vacía toma hoy."],
        ["IVA", "Tasa de IVA: 21, 10.5, 0 (exento). También acepta 'Exento'. Default 21."],
        [""],
        ["Notas:"],
        ["", "Cada combinación SKU + Fecha es una versión histórica."],
        ["", "Las ventas toman el costo vigente más reciente con Fecha ≤ fecha de la venta."],
        ["", "Los costos son por cuenta de ML — cada cuenta tiene su propio listado."],
    ]
    for row in instr:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_costos.xlsx"'},
    )


@app.get("/api/costos/{account_id}")
async def api_costos_list(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    rows = db_fetchall(
        "SELECT sku, cost, iva_rate, valid_from, updated_at FROM product_costs"
        " WHERE account_id=:aid ORDER BY sku, valid_from DESC",
        {"aid": cost_aid},
    )
    return {
        "items": [
            {
                "sku": r["sku"],
                "cost": float(r["cost"]),
                "iva_rate": float(r["iva_rate"] if r["iva_rate"] is not None else 21),
                "valid_from": r["valid_from"].isoformat() if hasattr(r["valid_from"], "isoformat") else str(r["valid_from"]),
                "updated_at": r["updated_at"].isoformat() if r.get("updated_at") else None,
            }
            for r in rows
        ],
    }


@app.post("/api/costos/{account_id}")
async def api_costos_upsert(
    request: Request, account_id: int,
    sku: str = Form(...),
    cost: float = Form(...),
    valid_from: Optional[str] = Form(None),
    iva_rate: Optional[str] = Form(None),
):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    try:
        rate_val = float(iva_rate) if iva_rate is not None and iva_rate != "" else 21.0
    except ValueError:
        rate_val = 21.0
    saved = db_save_product_costs(cost_aid, [{
        "sku": sku, "cost": cost,
        "valid_from": valid_from or datetime.utcnow().date().isoformat(),
        "iva_rate": rate_val,
    }])
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True, "saved": saved}


@app.put("/api/costos/{account_id}/edit")
async def api_costos_edit(
    request: Request, account_id: int,
    old_sku: str = Form(...),
    old_valid_from: str = Form(...),
    sku: str = Form(...),
    cost: float = Form(...),
    valid_from: str = Form(...),
    iva_rate: Optional[str] = Form(None),
):
    """Edita una entry de costo: borra la vieja (old_sku, old_valid_from)
    e inserta la nueva. Sirve para corregir errores de carga incluyendo
    cambios de SKU o fecha."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    try:
        rate_val = float(iva_rate) if iva_rate is not None and iva_rate != "" else 21.0
    except ValueError:
        rate_val = 21.0
    # Borrar entry vieja
    db_execute(
        "DELETE FROM product_costs WHERE account_id=:aid AND sku=:sku AND valid_from=:vf",
        {"aid": cost_aid, "sku": old_sku, "vf": old_valid_from},
    )
    # Insertar entry nueva
    saved = db_save_product_costs(cost_aid, [{
        "sku": sku, "cost": cost,
        "valid_from": valid_from,
        "iva_rate": rate_val,
    }])
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True, "saved": saved}


@app.delete("/api/costos/{account_id}/{sku}")
async def api_costos_delete(request: Request, account_id: int, sku: str, valid_from: Optional[str] = None):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    if valid_from:
        db_execute(
            "DELETE FROM product_costs WHERE account_id=:aid AND sku=:sku AND valid_from=:vf",
            {"aid": cost_aid, "sku": sku, "vf": valid_from},
        )
    else:
        # Sin fecha: borra TODAS las versiones de ese SKU
        db_execute(
            "DELETE FROM product_costs WHERE account_id=:aid AND sku=:sku",
            {"aid": cost_aid, "sku": sku},
        )
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True}


@app.post("/api/costos/{account_id}/upload")
async def api_costos_upload(request: Request, account_id: int):
    from fastapi import UploadFile, File  # local import to keep top clean
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    form = await request.form()
    upload = form.get("file")
    if upload is None or not hasattr(upload, "read"):
        raise HTTPException(400, "Falta el archivo")
    content = await upload.read()
    filename = (getattr(upload, "filename", "") or "").lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            items = parse_excel_costs(content)
        else:
            items = parse_csv_costs(content)
    except Exception as e:
        raise HTTPException(400, f"No pude leer el archivo: {e}")
    saved = db_save_product_costs(cost_aid, items)
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True, "saved": saved, "rows_parsed": len(items)}


# Caché invalidator: cuando cambian los costos hay que recalcular las órdenes.
def invalidate_orders_cache_for_account(account_id: int):
    db_execute(
        "DELETE FROM order_snapshot_cache WHERE account_id = :aid",
        {"aid": account_id},
    )


# ── Tarifas Flex (costo real del envío Flex que paga el vendedor) ───

FLEX_ZONES = [
    ("cercana",    "Zonas cercanas"),
    ("media",      "Zonas de media distancia"),
    ("lejana",     "Zonas lejanas"),
    ("muy_lejana", "Zonas muy lejanas"),
]
FLEX_ZONE_KEYS = {z[0] for z in FLEX_ZONES}


@app.get("/costos/envios-flex", response_class=HTMLResponse)
async def envios_flex_page(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    _r = _page_redirect(user, "envios_flex")
    if _r:
        return _r
    accounts = get_visible_accounts(user_id, user)
    accounts = await refresh_visible_account_nicknames(accounts)
    return templates.TemplateResponse("envios_flex.html", {
        "request": request, "user": user, "accounts": accounts,
        "zones": FLEX_ZONES, "perms": user_permissions(user),
    })


@app.get("/api/flex-tariffs/template")
async def api_flex_template(request: Request):
    """Excel modelo con las 4 zonas y filas de ejemplo."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    wb = Workbook()
    ws = wb.active
    ws.title = "Tarifas Flex"
    ws.append(["Zona", "Tarifa Propia", "Tarifa ML", "Fecha", "IVA"])
    fill = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
    font = Font(bold=True)
    for col in range(1, 6):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    today_iso = datetime.utcnow().date().isoformat()
    # Filas de ejemplo con tarifas ML referenciales de Argentina.
    ml_reference = {"cercana": 4490, "media": 6490, "lejana": 8690, "muy_lejana": 9990}
    for key, label in FLEX_ZONES:
        ws.append([label, 0, ml_reference.get(key, 0), today_iso, 21])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws2 = wb.create_sheet("Instrucciones")
    rows = [
        ["Plantilla de tarifas Flex"],
        [""],
        ["Columnas:"],
        ["Zona", "Una de las 4: 'Zonas cercanas', 'Zonas de media distancia', 'Zonas lejanas', 'Zonas muy lejanas'."],
        ["Tarifa Propia", "Costo SIN IVA que le pagás a tu mensajería para esa zona."],
        ["Tarifa ML", "Tarifa pública que cobra ML para esa zona (referencia para identificar la zona de cada venta)."],
        ["Fecha", "Vigencia desde (YYYY-MM-DD o DD/MM/YYYY). Si está vacía, hoy."],
        ["IVA", "Tasa de IVA (21, 10.5, 0 = exento). Default 21."],
        [""],
        ["Cómo se usa:"],
        ["", "Cuando llega una venta Flex el sistema mira el list_cost del envío y busca la entry con Tarifa ML más cercana (vigente al momento de la venta). Usa la Tarifa Propia + IVA como Costo Envío real."],
        ["", "Si no hay match (zona no cargada o muy distinta), usa el cálculo basado en lo que devuelve ML."],
        ["", "Las tarifas son por cuenta de ML."],
    ]
    for r in rows:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 100
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_envios_flex.xlsx"'},
    )


@app.get("/api/flex-tariffs/{account_id}")
async def api_flex_list(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    rows = db_fetchall(
        "SELECT zona, tarifa, tarifa_ml, iva_rate, valid_from, updated_at FROM flex_tariffs"
        " WHERE account_id=:aid ORDER BY zona, valid_from DESC",
        {"aid": cost_aid},
    )
    return {
        "items": [
            {
                "zona": r["zona"],
                "tarifa": float(r["tarifa"]),
                "tarifa_ml": float(r["tarifa_ml"]) if r["tarifa_ml"] is not None else None,
                "iva_rate": float(r["iva_rate"] if r["iva_rate"] is not None else 21),
                "valid_from": r["valid_from"].isoformat() if hasattr(r["valid_from"], "isoformat") else str(r["valid_from"]),
                "updated_at": r["updated_at"].isoformat() if r.get("updated_at") else None,
            }
            for r in rows
        ],
    }


def _normalize_zone(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if s in FLEX_ZONE_KEYS:
        return s
    # Mapear desde el label visible
    for key, label in FLEX_ZONES:
        if s == label.lower():
            return key
    # Mapeos sueltos
    if "cercana" in s:
        return "cercana"
    if "muy" in s and "lejana" in s:
        return "muy_lejana"
    if "lejana" in s:
        return "lejana"
    if "media" in s or "medio" in s:
        return "media"
    return None


def _save_flex_tariffs(account_id: int, items: list) -> int:
    """Inserta/actualiza tarifas flex.
    items: {zona, tarifa, tarifa_ml, iva_rate, valid_from}."""
    if not items:
        return 0
    params = []
    today_iso = datetime.utcnow().date().isoformat()
    for it in items:
        zona = _normalize_zone(it.get("zona"))
        if zona is None:
            continue
        try:
            tarifa = float(it.get("tarifa") or 0)
        except (TypeError, ValueError):
            continue
        if tarifa < 0:
            continue
        vf = it.get("valid_from") or today_iso
        if hasattr(vf, "isoformat"):
            vf = vf.isoformat()
        try:
            iva_rate = float(it.get("iva_rate") if it.get("iva_rate") is not None else 21)
        except (TypeError, ValueError):
            iva_rate = 21.0
        try:
            tarifa_ml = float(it.get("tarifa_ml")) if it.get("tarifa_ml") not in (None, "", 0) else None
        except (TypeError, ValueError):
            tarifa_ml = None
        params.append({
            "aid": account_id, "zona": zona, "tarifa": tarifa,
            "iva": iva_rate, "vf": vf, "tml": tarifa_ml,
        })
    if not params:
        return 0
    with engine.connect() as conn:
        conn.execute(text(
            "INSERT INTO flex_tariffs (account_id, zona, tarifa, tarifa_ml, iva_rate, valid_from, updated_at)"
            " VALUES (:aid, :zona, :tarifa, :tml, :iva, :vf, NOW())"
            " ON CONFLICT (account_id, zona, valid_from) DO UPDATE SET"
            " tarifa = EXCLUDED.tarifa,"
            " tarifa_ml = EXCLUDED.tarifa_ml,"
            " iva_rate = EXCLUDED.iva_rate,"
            " updated_at = NOW()"
        ), params)
        conn.commit()
    return len(params)


def db_get_flex_tariffs(account_id: int) -> list:
    """Devuelve [{zona, tarifa, tarifa_ml, iva_rate, valid_from}, ...]
    ordenado por zona y fecha asc, para hacer lookup vigente al momento
    de la venta."""
    rows = db_fetchall(
        "SELECT zona, tarifa, tarifa_ml, iva_rate, valid_from FROM flex_tariffs"
        " WHERE account_id=:aid ORDER BY zona, valid_from",
        {"aid": account_id},
    )
    out = []
    for r in rows:
        vf = r["valid_from"]
        out.append({
            "zona": r["zona"],
            "tarifa": float(r["tarifa"]),
            "tarifa_ml": float(r["tarifa_ml"]) if r["tarifa_ml"] is not None else None,
            "iva_rate": float(r["iva_rate"] if r["iva_rate"] is not None else 21),
            "valid_from": vf.isoformat() if hasattr(vf, "isoformat") else str(vf),
        })
    return out


def match_flex_tariff(entries: list, list_cost: float, sale_date: str) -> dict:
    """Busca la entry de flex_tariffs con tarifa_ml más cercana al
    list_cost del shipment, vigente al sale_date (valid_from <= sale_date).
    Devuelve None si no encuentra match razonable (dif > 15%)."""
    if not entries or not list_cost or list_cost <= 0 or not sale_date:
        return None
    # Filtrar entries vigentes a la fecha de la venta, con tarifa_ml cargada.
    # Para cada zona conservar la entry más reciente con valid_from <= sale_date.
    latest_by_zone = {}
    for e in entries:
        if e.get("tarifa_ml") is None:
            continue
        if e["valid_from"] > sale_date:
            continue
        prev = latest_by_zone.get(e["zona"])
        if prev is None or e["valid_from"] > prev["valid_from"]:
            latest_by_zone[e["zona"]] = e
    if not latest_by_zone:
        return None
    # Mejor match por diferencia relativa.
    best = None
    best_diff = None
    for e in latest_by_zone.values():
        diff = abs(e["tarifa_ml"] - list_cost) / list_cost
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best = e
    # Tolerancia 15% (ej. variación por peso del envío)
    if best_diff is not None and best_diff <= 0.15:
        return best
    return None


def flex_cost_with_iva(entry: dict) -> float:
    """Convierte la tarifa propia (sin IVA) a un valor con IVA aplicado."""
    if not entry:
        return 0.0
    tarifa = float(entry.get("tarifa") or 0)
    rate = float(entry.get("iva_rate") or 0)
    return tarifa * (1 + rate / 100.0)


@app.post("/api/flex-tariffs/{account_id}")
async def api_flex_upsert(
    request: Request, account_id: int,
    zona: str = Form(...),
    tarifa: float = Form(...),
    tarifa_ml: Optional[str] = Form(None),
    valid_from: Optional[str] = Form(None),
    iva_rate: Optional[str] = Form(None),
):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    try:
        rate_val = float(iva_rate) if iva_rate is not None and iva_rate != "" else 21.0
    except ValueError:
        rate_val = 21.0
    try:
        tml_val = float(tarifa_ml) if tarifa_ml not in (None, "") else None
    except ValueError:
        tml_val = None
    saved = _save_flex_tariffs(cost_aid, [{
        "zona": zona, "tarifa": tarifa, "tarifa_ml": tml_val,
        "valid_from": valid_from or datetime.utcnow().date().isoformat(),
        "iva_rate": rate_val,
    }])
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True, "saved": saved}


@app.put("/api/flex-tariffs/{account_id}/edit")
async def api_flex_edit(
    request: Request, account_id: int,
    old_zona: str = Form(...),
    old_valid_from: str = Form(...),
    zona: str = Form(...),
    tarifa: float = Form(...),
    valid_from: str = Form(...),
    tarifa_ml: Optional[str] = Form(None),
    iva_rate: Optional[str] = Form(None),
):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    try:
        rate_val = float(iva_rate) if iva_rate is not None and iva_rate != "" else 21.0
    except ValueError:
        rate_val = 21.0
    try:
        tml_val = float(tarifa_ml) if tarifa_ml not in (None, "") else None
    except ValueError:
        tml_val = None
    db_execute(
        "DELETE FROM flex_tariffs WHERE account_id=:aid AND zona=:zona AND valid_from=:vf",
        {"aid": cost_aid, "zona": old_zona, "vf": old_valid_from},
    )
    saved = _save_flex_tariffs(cost_aid, [{
        "zona": zona, "tarifa": tarifa, "tarifa_ml": tml_val,
        "valid_from": valid_from, "iva_rate": rate_val,
    }])
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True, "saved": saved}


@app.delete("/api/flex-tariffs/{account_id}/{zona}")
async def api_flex_delete(request: Request, account_id: int, zona: str, valid_from: Optional[str] = None):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    if valid_from:
        db_execute(
            "DELETE FROM flex_tariffs WHERE account_id=:aid AND zona=:zona AND valid_from=:vf",
            {"aid": cost_aid, "zona": zona, "vf": valid_from},
        )
    else:
        db_execute(
            "DELETE FROM flex_tariffs WHERE account_id=:aid AND zona=:zona",
            {"aid": cost_aid, "zona": zona},
        )
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True}


@app.post("/api/flex-tariffs/{account_id}/upload")
async def api_flex_upload(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    user = get_user(user_id)
    cost_aid = _cost_account_id_for(user, account_id)
    form = await request.form()
    upload = form.get("file")
    if upload is None or not hasattr(upload, "read"):
        raise HTTPException(400, "Falta el archivo")
    content = await upload.read()
    filename = (getattr(upload, "filename", "") or "").lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            items = _parse_excel_flex(content)
        else:
            items = _parse_csv_flex(content)
    except Exception as e:
        raise HTTPException(400, f"No pude leer el archivo: {e}")
    saved = _save_flex_tariffs(cost_aid, items)
    _invalidate_all_user_accounts_cache(user, account_id)
    return {"ok": True, "saved": saved, "rows_parsed": len(items)}


def _parse_excel_flex(content: bytes) -> list:
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = None
    data_start = 0
    for idx, row in enumerate(rows[:5]):
        if row is None:
            continue
        normalized = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("zona" in n for n in normalized):
            header = normalized
            data_start = idx + 1
            break
    if not header:
        return []
    zona_idx = _find_col(header, ["zona"])
    # 'tarifa propia' o sólo 'tarifa' = lo que paga el vendedor a su mensajería
    tarifa_idx = _find_col(header, ["tarifa propia", "tu tarifa", "mi tarifa", "costo propio"])
    if tarifa_idx is None:
        tarifa_idx = _find_col(header, ["tarifa", "costo", "cost", "precio"])
    # 'tarifa ml' = lo que ML cobra públicamente (referencia para matchear zona)
    tarifa_ml_idx = _find_col(header, ["tarifa ml", "ml", "referencia", "publica", "pública"])
    fecha_idx = _find_col(header, ["fecha", "date", "vigencia", "desde"])
    iva_idx = _find_col(header, ["iva"])
    if zona_idx is None or tarifa_idx is None:
        return []
    items = []
    today_iso = datetime.utcnow().date().isoformat()
    for row in rows[data_start:]:
        if row is None or len(row) <= max(zona_idx, tarifa_idx):
            continue
        zona = row[zona_idx]
        tarifa = row[tarifa_idx]
        if zona is None or tarifa is None:
            continue
        fecha = _parse_date_cell(row[fecha_idx]) if fecha_idx is not None and len(row) > fecha_idx else None
        iva_rate = _parse_iva_rate_cell(row[iva_idx]) if iva_idx is not None and len(row) > iva_idx else 21.0
        tarifa_ml = None
        if tarifa_ml_idx is not None and len(row) > tarifa_ml_idx and row[tarifa_ml_idx] not in (None, ""):
            try:
                tarifa_ml = float(row[tarifa_ml_idx])
            except (TypeError, ValueError):
                tarifa_ml = None
        items.append({
            "zona": str(zona), "tarifa": tarifa, "tarifa_ml": tarifa_ml,
            "valid_from": fecha or today_iso, "iva_rate": iva_rate,
        })
    return items


def _parse_csv_flex(content: bytes) -> list:
    import csv, io
    text_content = content.decode("utf-8-sig", errors="ignore")
    sample = text_content[:2048]
    sep = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text_content), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []
    header = [(c or "").strip().lower() for c in rows[0]]
    zona_idx = _find_col(header, ["zona"])
    tarifa_idx = _find_col(header, ["tarifa propia", "tu tarifa", "mi tarifa", "costo propio"])
    if tarifa_idx is None:
        tarifa_idx = _find_col(header, ["tarifa", "costo", "cost", "precio"])
    tarifa_ml_idx = _find_col(header, ["tarifa ml", "ml", "referencia", "publica", "pública"])
    fecha_idx = _find_col(header, ["fecha", "date", "vigencia", "desde"])
    iva_idx = _find_col(header, ["iva"])
    if zona_idx is None or tarifa_idx is None:
        return []
    items = []
    today_iso = datetime.utcnow().date().isoformat()
    def _to_float(v):
        if v is None: return None
        s = str(v).strip()
        if not s: return None
        s = s.replace(".", "").replace(",", ".") if sep == ";" else s.replace(",", "")
        try: return float(s)
        except ValueError: return None
    for row in rows[1:]:
        if len(row) <= max(zona_idx, tarifa_idx):
            continue
        zona = (row[zona_idx] or "").strip()
        tarifa = _to_float(row[tarifa_idx])
        if not zona or tarifa is None:
            continue
        fecha = _parse_date_cell(row[fecha_idx]) if fecha_idx is not None and len(row) > fecha_idx else None
        iva_rate = _parse_iva_rate_cell(row[iva_idx]) if iva_idx is not None and len(row) > iva_idx else 21.0
        tarifa_ml = _to_float(row[tarifa_ml_idx]) if tarifa_ml_idx is not None and len(row) > tarifa_ml_idx else None
        items.append({
            "zona": zona, "tarifa": tarifa, "tarifa_ml": tarifa_ml,
            "valid_from": fecha or today_iso, "iva_rate": iva_rate,
        })
    return items


# ── Descuentos / Promociones ───────────────────────────────────────

@app.get("/descuentos", response_class=HTMLResponse)
async def descuentos_page(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    _r = _page_redirect(user, "descuentos")
    if _r:
        return _r
    accounts = get_visible_accounts(user_id, user)
    accounts = await refresh_visible_account_nicknames(accounts)
    return templates.TemplateResponse("descuentos.html", {
        "request": request, "user": user, "accounts": accounts,
        "perms": user_permissions(user),
    })


@app.get("/ranking", response_class=HTMLResponse)
async def ranking_page(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    _r = _page_redirect(user, "ranking")
    if _r:
        return _r
    accounts = get_visible_accounts(user_id, user)
    accounts = await refresh_visible_account_nicknames(accounts)
    return templates.TemplateResponse("ranking.html", {
        "request": request, "user": user, "accounts": accounts,
        "perms": user_permissions(user),
    })


@app.get("/api/ranking/{account_id}")
async def api_ranking(request: Request, account_id: int,
                      date_from: Optional[str] = None,
                      date_to: Optional[str] = None):
    """Agrupa las ventas del período por SKU. Por cada SKU:
      - unidades vendidas, ventas (cantidad), facturación
      - comisión, ingreso envío, bonif, costo envío, CMV (ponderados por
        la fracción del monto del item dentro de su orden)
      - ganancia, margen ponderado (ganancia / facturación)
    Ordenado por facturación descendente."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "ranking")
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)

    # Reutilizamos los snapshots cacheados de órdenes para no volver a
    # pegar a ML. Las órdenes ya vienen con ítems y cálculos resueltos.
    today = datetime.utcnow().date()
    df = date_from or str(today - timedelta(days=365))
    dt = date_to or str(today)
    if (datetime.strptime(dt, "%Y-%m-%d") - datetime.strptime(df, "%Y-%m-%d")).days > 365:
        df = str(datetime.strptime(dt, "%Y-%m-%d").date() - timedelta(days=365))

    cached_orders = db_fetch_order_snapshots(account_id, df, dt)
    if not cached_orders:
        return {"items": [], "period": {"from": df, "to": dt},
                "info": "No hay datos cacheados para este período. Cargá el Dashboard primero para generar el caché."}

    # Sólo ventas pagadas cuentan para ranking
    paid = [o for o in cached_orders if o.get("estado") == "paid"]

    rank: dict = {}
    for order in paid:
        items = order.get("items") or []
        if not items:
            continue
        order_monto = sum(float(i.get("monto", 0) or 0) for i in items) or float(order.get("monto", 0) or 0)
        order_envio = float(order.get("envio", 0) or 0)
        order_ing = float(order.get("ingreso_envio", 0) or 0)
        order_bonif = float(order.get("bonificacion", 0) or 0)
        for item in items:
            sku = (item.get("sku") or "").strip()
            if not sku:
                sku = (item.get("titulo") or "Sin SKU")[:40]
            qty = int(item.get("cantidad", 1) or 1)
            item_monto = float(item.get("monto", 0) or 0)
            item_com = float(item.get("comision", 0) or 0)
            item_cmv = float(item.get("cmv", 0) or 0)
            # Prorratear envío, ingreso envío y bonif del pedido en base
            # al peso del item dentro del monto total del pedido.
            share = (item_monto / order_monto) if order_monto > 0 else (1.0 / len(items))
            item_envio = order_envio * share
            item_ing = order_ing * share
            item_bonif = order_bonif * share
            rec = rank.setdefault(sku, {
                "sku": sku, "titulo": item.get("titulo") or "",
                "ventas": 0, "unidades": 0,
                "facturacion": 0.0, "comision": 0.0,
                "ingreso_envio": 0.0, "bonificacion": 0.0,
                "envio": 0.0, "cmv": 0.0, "ganancia": 0.0,
            })
            rec["ventas"] += 1
            rec["unidades"] += qty
            rec["facturacion"] += item_monto
            rec["comision"] += item_com
            rec["ingreso_envio"] += item_ing
            rec["bonificacion"] += item_bonif
            rec["envio"] += item_envio
            rec["cmv"] += item_cmv
            rec["ganancia"] += (
                item_monto + item_ing + item_bonif
                - item_com - item_envio - item_cmv
            )

    out = []
    for r in rank.values():
        # Redondeos
        r["facturacion"] = round(r["facturacion"], 2)
        r["comision"] = round(r["comision"], 2)
        r["ingreso_envio"] = round(r["ingreso_envio"], 2)
        r["bonificacion"] = round(r["bonificacion"], 2)
        r["envio"] = round(r["envio"], 2)
        r["cmv"] = round(r["cmv"], 2)
        r["ganancia"] = round(r["ganancia"], 2)
        r["margen_pct"] = round((r["ganancia"] / r["facturacion"]) * 100, 2) if r["facturacion"] > 0 else 0
        out.append(r)
    out.sort(key=lambda x: x["facturacion"], reverse=True)
    return {"items": out, "period": {"from": df, "to": dt}}


# ════════════════════════ ETIQUETAS (envíos a despachar) ════════════════════════
# Trae los envíos LISTOS PARA DESPACHAR de una cuenta, arma una fila por envío
# (SKU, unidades, código postal, destinatario) y permite:
#   - imprimir las etiquetas listas (PDF que devuelve ML), ordenadas por SKU y
#     filtrables por código postal;
#   - descargar un resumen de unidades por SKU en PDF para armar los paquetes.
# Las etiquetas NO las generamos: las da ML vía /shipment_labels (response_type=pdf).

# Cache liviano por cuenta: que pedir el PDF de etiquetas o el de resumen no
# tenga que volver a bajar todo de ML.
ETIQUETAS_CACHE: dict = {}
ETIQUETAS_CACHE_TTL_SECONDS = 300

# Traducción del tipo logístico de ML al nombre que usa el vendedor.
ETIQUETAS_LOGISTIC_LABELS = {
    "cross_docking": "Colecta",
    "self_service": "Flex",
    "drop_off": "Agencia",
    "xd_drop_off": "Agencia",
    "fulfillment": "Full",
}
# Tipo logístico → grupo (como agrupa el panel "Envíos de hoy" de ML).
ETIQUETAS_GROUP_FOR_LOGISTIC = {
    "cross_docking": ("colecta", "Colecta"),
    "self_service": ("flex", "Flex"),
    "drop_off": ("agencia", "Agencia"),
    "xd_drop_off": ("agencia", "Agencia"),
}
# Zonas CERCANAS para Flex: CABA (rango 1000-1499) + estos códigos postales
# EXACTOS (Vicente López, San Martín, San Isidro, 3 de Febrero, según la lista
# que pasó el vendedor). TODO lo que no esté acá → zona LEJANA. Se usa el CP
# exacto (no rangos) porque un CP suelto de un partido lejano puede caer dentro
# del rango de uno cercano (ej: 1627 no es cercano aunque esté entre 1602-1638).
ETIQUETAS_CERCANA_CPS = {
    # Vicente López
    1602, 1603, 1604, 1605, 1606, 1636, 1637, 1638,
    # San Isidro
    1607, 1609, 1640, 1641, 1642, 1643,
    # San Martín
    1650, 1653, 1654, 1655, 1657, 1672, 1676,
    # 3 de Febrero
    1674, 1675, 1678, 1682, 1683, 1684, 1687, 1702,
}


def _zona_distancia(zip_code) -> str:
    """cercana / lejana según el CP. CABA (1000-1499) y la lista exacta de CPs
    cercanos = cercana; cualquier otro CP válido = lejana. '' si no hay CP."""
    digits = "".join(ch for ch in str(zip_code or "") if ch.isdigit())[:4]
    if not digits:
        return ""
    z = int(digits)
    if 1000 <= z <= 1499:  # CABA
        return "cercana"
    if z in ETIQUETAS_CERCANA_CPS:
        return "cercana"
    return "lejana"
# Buckets que SÍ hay que despachar/armar (los que entran al resumen por SKU y
# quedan seleccionados por defecto). Canceladas y reprogramadas no se arman.
ETIQUETAS_PACKABLE_BUCKETS = {"listas", "etiquetas", "demoradas"}
# Tipos logísticos que SÍ imprimen etiqueta del vendedor. Todo lo demás (Full /
# fulfillment, "a acordar"/ME1, o sin tipo) se excluye: no necesitamos Full ni
# basura vieja en esta pantalla.
ETIQUETAS_ALLOWED_LOGISTICS = {
    "cross_docking", "self_service", "drop_off", "xd_drop_off",
}


def _dispatch_date_ar(sh) -> str:
    """Fecha límite de despacho del envío (YYYY-MM-DD, ya en hora AR porque ML
    la devuelve con offset -03:00). Se usa para separar 'Envíos de hoy' de
    'Próximos días': hoy o vencido = se despacha hoy; futuro = se excluye."""
    if not isinstance(sh, dict):
        return ""
    lt = sh.get("lead_time") or {}
    candidates = [
        lt.get("estimated_handling_limit"),
        sh.get("estimated_handling_limit"),
        lt.get("estimated_delivery_limit"),
    ]
    for c in candidates:
        if isinstance(c, dict) and c.get("date"):
            return str(c["date"])[:10]
        if isinstance(c, str) and c:
            return c[:10]
    return ""


def _classify_shipment_bucket(status, substatus):
    """Clasifica un envío en el mismo sub-estado que muestra el panel de ML.
    Devuelve (bucket_key, etiqueta). Los strings de substatus de ML pueden
    variar; este es el único lugar a tocar si algún envío cae en el bucket
    equivocado (cada fila expone su status/substatus crudo para verificar)."""
    st = (status or "").lower()
    sub = (substatus or "").lower()
    if st in ("cancelled", "canceled"):
        return ("canceladas", "Canceladas. No despachar")
    if st in ("shipped", "delivered", "not_delivered"):
        return ("en_camino", "En camino")
    if sub in ("reprogrammed", "rescheduled", "delayed_reprogrammed",
               "waiting_for_carrier_authorization", "buyer_rescheduled"):
        return ("reprogramadas", "Reprogramadas")
    if sub in ("delayed", "stale", "overdue", "delivery_failed"):
        return ("demoradas", "Demoradas. Despachar")
    if sub in ("ready_to_print", "invoice_pending", "regrouping",
               "printing", "ready_to_print_pending"):
        return ("etiquetas", "Etiquetas por imprimir")
    # printed, ready_to_ship, picked_up, in_packing_list, etc.
    return ("listas", "Listas para despachar")


def _addr_name(v):
    """receiver_address.state/city puede venir como dict {name} o como string."""
    if isinstance(v, dict):
        return (v.get("name") or "").strip()
    return str(v or "").strip()


async def _orders_search_all(client, headers, base, max_pages=60):
    """Pagina /orders/search con los filtros dados y devuelve todas las órdenes."""
    out = []
    offset = 0
    for _ in range(max_pages):
        try:
            r = await client.get(
                f"{ML_API_URL}/orders/search",
                headers=headers,
                params={**base, "offset": offset, "limit": 50},
            )
        except Exception:
            break
        if r.status_code != 200:
            break
        data = r.json()
        res = data.get("results", []) if isinstance(data, dict) else []
        if not res:
            break
        out.extend(res)
        offset += len(res)
        total = (data.get("paging") or {}).get("total", 0)
        if offset >= total or len(res) < 50:
            break
    return out


# ── Horario de corte de colecta (para separar "hoy" de "mañana") ──────────────
# ML calcula "Envíos de hoy / Próximos días" con el horario de la colecta: lo que
# queda listo después del corte del día va a la colecta de mañana. Ese horario NO
# está en la API (se renderiza en el HTML del panel de ML), así que lo guardamos
# editable por cuenta. Default = lo que mostraba el panel de TIENDA BAE.
AR_TZ = timezone(timedelta(hours=-3))
ETQ_WEEKDAY_NAMES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
# Defaults por tipo logístico (lo que mostraban los paneles de TIENDA BAE).
# Colecta (cross_docking): cambia semana a semana → se edita a mano.
# Flex (self_service): lo configura el vendedor y no cambia salvo que lo cambie.
ETQ_DEFAULTS = {
    "cross_docking": {
        0: {"cutoff": "13:30", "ventana": "14:00 a 16:00"},
        1: {"cutoff": "13:30", "ventana": "14:00 a 16:00"},
        2: {"cutoff": "13:30", "ventana": "14:00 a 16:00"},
        3: {"cutoff": "13:30", "ventana": "14:00 a 16:00"},
        4: {"cutoff": "13:30", "ventana": "14:00 a 16:00"},
        5: {"cutoff": "07:30", "ventana": "09:30 a 11:30"},
        6: {"cutoff": "", "ventana": ""},
    },
    "self_service": {
        0: {"cutoff": "13:00", "ventana": "16:00 a 21:00"},
        1: {"cutoff": "13:00", "ventana": "16:00 a 21:00"},
        2: {"cutoff": "13:00", "ventana": "16:00 a 21:00"},
        3: {"cutoff": "13:00", "ventana": "16:00 a 21:00"},
        4: {"cutoff": "13:00", "ventana": "16:00 a 21:00"},
        5: {"cutoff": "12:00", "ventana": "09:00 a 21:00"},
        6: {"cutoff": "", "ventana": ""},
    },
}
ETQ_CUTOFF_LOGISTICS = ("cross_docking", "self_service")


def _ensure_cutoff_table():
    try:
        db_execute(
            "CREATE TABLE IF NOT EXISTS etq_cutoff ("
            "account_id INTEGER NOT NULL, logistic TEXT NOT NULL, weekday INTEGER NOT NULL, "
            "cutoff TEXT, ventana TEXT, PRIMARY KEY (account_id, logistic, weekday))",
            {},
        )
    except Exception:
        pass


def _get_cutoffs(account_id, logistic) -> dict:
    """Devuelve {weekday(0=Lun): {cutoff, ventana}} para (cuenta, logística). Si no
    hay nada guardado, devuelve los defaults (sin persistir)."""
    _ensure_cutoff_table()
    base = ETQ_DEFAULTS.get(logistic, {})
    result = {k: dict(v) for k, v in base.items()}
    try:
        rows = db_fetchall(
            "SELECT weekday, cutoff, ventana FROM etq_cutoff WHERE account_id=:a AND logistic=:l",
            {"a": account_id, "l": logistic},
        )
        for r in rows:
            result[int(r["weekday"])] = {"cutoff": (r["cutoff"] or ""), "ventana": (r["ventana"] or "")}
    except Exception:
        pass
    return result


def _save_cutoffs(account_id, logistic, schedule: dict):
    _ensure_cutoff_table()
    for wd in range(7):
        item = schedule.get(wd) or schedule.get(str(wd)) or {}
        db_execute(
            "INSERT INTO etq_cutoff (account_id, logistic, weekday, cutoff, ventana) "
            "VALUES (:a, :l, :w, :c, :v) ON CONFLICT (account_id, logistic, weekday) "
            "DO UPDATE SET cutoff=EXCLUDED.cutoff, ventana=EXCLUDED.ventana",
            {"a": account_id, "l": logistic, "w": wd,
             "c": str(item.get("cutoff") or "").strip(),
             "v": str(item.get("ventana") or "").strip()},
        )


def _cutoff_minutes(cutoff: str):
    try:
        hh, mm = str(cutoff).split(":")
        return int(hh) * 60 + int(mm)
    except Exception:
        return None


def _ready_after_cutoff(ready_iso: str, cutoffs: dict, today_ar: str) -> bool:
    """True si el envío quedó listo HOY pero después del corte del día → va a la
    colecta/despacho de mañana (no es para despachar hoy)."""
    if not ready_iso:
        return False
    try:
        dt = datetime.fromisoformat(ready_iso)
    except Exception:
        return False
    if dt.tzinfo is None:
        return False
    ar = dt.astimezone(AR_TZ)
    if ar.date().isoformat() != today_ar:
        return False  # listo otro día: ya está esperando, es para despachar
    cm = _cutoff_minutes((cutoffs.get(ar.weekday()) or {}).get("cutoff") or "")
    if cm is None:
        return False
    return (ar.hour * 60 + ar.minute) > cm


def _ready_ar_parts(ready_iso: str):
    """(fecha DD/MM, hora HH:MM) en hora AR de cuando el envío quedó listo."""
    if not ready_iso:
        return ("", "")
    try:
        dt = datetime.fromisoformat(ready_iso)
    except Exception:
        return ("", "")
    if dt.tzinfo is None:
        return ("", "")
    ar = dt.astimezone(AR_TZ)
    return (ar.strftime("%d/%m"), ar.strftime("%H:%M"))


def _etq_latin(s) -> str:
    """fpdf2 con fuentes core usa latin-1; reemplaza lo que no entre."""
    return str(s or "").encode("latin-1", "replace").decode("latin-1")


def _etq_sku_sort_key(s):
    """Ordena por SKU; los vacíos van al final."""
    s = (s or "").strip().upper()
    return (s == "", s)


async def _fetch_pending_shipments(account_id, acc, token) -> dict:
    """Baja los envíos de hoy de la cuenta (listos para despachar + canceladas
    recientes) y devuelve {str(shipment_id): fila}. Cada fila trae SKU(s),
    unidades, código postal, zona, destinatario, tipo logístico y el grupo
    (Colecta/Flex/...) y sub-estado (canceladas/reprogramadas/demoradas/
    etiquetas por imprimir/listas para despachar), igual que el panel de ML."""
    headers = {"Authorization": f"Bearer {token}"}
    seller = acc["ml_user_id"]
    now = datetime.utcnow()
    today_ar = (now - timedelta(hours=3)).date().isoformat()
    recent_from = (now - timedelta(days=2)).strftime("%Y-%m-%dT00:00:00.000-00:00")
    recent_to = now.strftime("%Y-%m-%dT23:59:59.000-00:00")
    today_from = f"{today_ar}T00:00:00.000-03:00"   # inicio de hoy en AR
    async with httpx.AsyncClient(timeout=60) as client:
        # Listos para despachar (etiquetas por imprimir / listas / reprogramadas
        # — se distinguen luego por substatus).
        ready = await _orders_search_all(client, headers, {
            "seller": seller, "shipping.status": "ready_to_ship", "sort": "date_desc",
        })
        # Canceladas recientes (las "No despachar" del día). Se acotan por
        # date_last_updated para no traer cancelaciones viejas.
        cancelled = await _orders_search_all(client, headers, {
            "seller": seller, "shipping.status": "cancelled", "sort": "date_desc",
            "order.date_last_updated.from": recent_from,
            "order.date_last_updated.to": recent_to,
        })
        # Despachados HOY ("En camino": la logística escaneó la etiqueta).
        shipped = await _orders_search_all(client, headers, {
            "seller": seller, "shipping.status": "shipped", "sort": "date_desc",
            "order.date_last_updated.from": today_from,
            "order.date_last_updated.to": recent_to,
        })
        orders = ready + cancelled + shipped

        # Agrupar ítems por shipment_id (un pack/envío puede tener varias órdenes)
        ship_items: dict = {}
        ship_orders: dict = {}   # sid -> {"order_ids": [...], "pack_id": ...}
        for o in orders:
            shp = o.get("shipping") or {}
            sid = shp.get("id")
            if not sid:
                continue
            bucket = ship_items.setdefault(sid, [])
            so = ship_orders.setdefault(sid, {"order_ids": [], "pack_id": None})
            oid = o.get("id")
            if oid and oid not in so["order_ids"]:
                so["order_ids"].append(oid)
            if o.get("pack_id"):
                so["pack_id"] = o.get("pack_id")
            for it in o.get("order_items", []):
                itm = it.get("item") or {}
                bucket.append({
                    "sku": (itm.get("seller_sku") or "").strip(),
                    "title": itm.get("title") or "",
                    "qty": int(it.get("quantity", 1) or 1),
                })

        # Detalle de cada shipment: zona/CP, tipo logístico, status/substatus,
        # fecha de despacho. Con reintentos para no perder envíos por 429/timeout
        # (si fallaba, quedaban sin tipo logístico y caían en "Otros").
        sem = asyncio.Semaphore(8)

        async def _get_ship(sid):
            async with sem:
                for attempt in range(3):
                    try:
                        rs = await client.get(f"{ML_API_URL}/shipments/{sid}", headers=headers)
                        if rs.status_code == 200:
                            return sid, rs.json()
                        if rs.status_code in (429, 500, 502, 503, 504):
                            await asyncio.sleep(0.5 * (2 ** attempt))
                            continue
                        return sid, None
                    except Exception:
                        await asyncio.sleep(0.5 * (2 ** attempt))
                return sid, None

        details = await asyncio.gather(*[_get_ship(sid) for sid in ship_items])

    cutoffs_by_logistic = {lg: _get_cutoffs(account_id, lg) for lg in ETQ_CUTOFF_LOGISTICS}
    diag = {"today_ar": today_ar, "fetched": len(details), "kept": 0,
            "excluded_future": 0, "excluded_after_cutoff": 0, "excluded_logistic": 0,
            "excluded_logistic_by_type": {}}
    rows: dict = {}
    for sid, sh in details:
        logistic_type = status = substatus = ""
        if isinstance(sh, dict):
            logistic_type = (sh.get("logistic_type")
                             or (sh.get("logistic") or {}).get("type") or "")
            status = sh.get("status") or ""
            substatus = sh.get("substatus") or ""
        # Excluir Full / "a acordar" / sin tipo: no imprimen etiqueta del vendedor.
        if logistic_type not in ETIQUETAS_ALLOWED_LOGISTICS:
            diag["excluded_logistic"] += 1
            k = logistic_type or "(sin tipo)"
            diag["excluded_logistic_by_type"][k] = diag["excluded_logistic_by_type"].get(k, 0) + 1
            continue
        # Excluir lo que se despacha a futuro (Próximos días). Hoy o vencido = hoy.
        dispatch_date = _dispatch_date_ar(sh)
        if dispatch_date and dispatch_date > today_ar:
            diag["excluded_future"] += 1
            continue
        # Excluir lo que quedó listo HOY después del corte del día: ese va a la
        # colecta/despacho de MAÑANA (mismo criterio que ML para "Próximos días").
        # Aplica a Colecta y Flex, cada uno con su propio horario de corte.
        ready_iso_full = ""
        if isinstance(sh, dict):
            ready_iso_full = str((sh.get("status_history") or {}).get("date_ready_to_ship") or "")
        cuts = cutoffs_by_logistic.get(logistic_type)
        if status == "ready_to_ship" and cuts and _ready_after_cutoff(ready_iso_full, cuts, today_ar):
            diag["excluded_after_cutoff"] += 1
            continue
        diag["kept"] += 1
        items = ship_items.get(sid, [])
        # Consolidar SKUs repetidos dentro del mismo envío
        agg: dict = {}
        for it in items:
            key = it["sku"] or it["title"] or "Sin SKU"
            e = agg.setdefault(key, {"sku": it["sku"], "title": it["title"], "qty": 0})
            e["qty"] += it["qty"]
        items_list = sorted(agg.values(), key=lambda x: _etq_sku_sort_key(x["sku"]))
        skus = [e["sku"] for e in items_list if e["sku"]]
        ra = (sh.get("receiver_address") or {}) if isinstance(sh, dict) else {}
        zip_code = str(ra.get("zip_code") or "").strip()
        receiver = (ra.get("receiver_name") or ra.get("name") or "").strip()
        zona = _addr_name(ra.get("state"))
        localidad = _addr_name(ra.get("city")) or _addr_name(ra.get("municipality"))
        group_key, group_label = ETIQUETAS_GROUP_FOR_LOGISTIC.get(
            logistic_type, ("otros", "Otros"))
        bucket_key, bucket_label = _classify_shipment_bucket(status, substatus)
        # Fecha en que el envío quedó listo (status_history), solo informativa.
        # NO la usamos para "Demoradas": casi todos los envíos quedan listos uno
        # o más días antes y siguen esperando la colecta de hoy — eso es "Listas
        # para despachar", no demorado. ML distingue las demoradas con la fecha/
        # hora programada de la colecta, que la API no expone. Por eso no lo
        # inventamos: lo que está listo va todo a "Listas para despachar".
        ready_date = ready_iso_full[:10]
        ready_fecha, ready_hora = _ready_ar_parts(ready_iso_full)
        so = ship_orders.get(sid, {})
        order_ids = [str(x) for x in (so.get("order_ids") or [])]
        pack_id = so.get("pack_id")
        # Número de venta como lo muestra ML: el pack si la venta es de varios
        # ítems, si no el id de la orden.
        venta = str(pack_id) if pack_id else (order_ids[0] if order_ids else str(sid))
        rows[str(sid)] = {
            "shipment_id": sid,
            "venta": venta,
            "order_ids": order_ids,
            "pack_id": str(pack_id) if pack_id else "",
            "sku_principal": skus[0] if skus else "",
            "skus": skus,
            "items": items_list,
            "total_units": sum(e["qty"] for e in items_list),
            "zip_code": zip_code,
            "zona": zona,
            "localidad": localidad,
            "distancia": _zona_distancia(zip_code),
            "receiver": receiver,
            "dispatch_date": dispatch_date,
            "ready_date": ready_date,
            "ready_fecha": ready_fecha,
            "ready_hora": ready_hora,
            "logistic_type": logistic_type,
            "logistic_label": ETIQUETAS_LOGISTIC_LABELS.get(logistic_type, logistic_type or "—"),
            "group_key": group_key,
            "group_label": group_label,
            "bucket_key": bucket_key,
            "bucket_label": bucket_label,
            "packable": bucket_key in ETIQUETAS_PACKABLE_BUCKETS,
            "raw_status": status,
            "raw_substatus": substatus,
        }
    return rows, diag


def _build_resumen_pdf(items, total_units, account_name="") -> bytes:
    """Arma el PDF de resumen de unidades por SKU con fpdf2."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Resumen de unidades por SKU", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(110, 110, 110)
    sub = f"Cuenta: {account_name or '-'}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    pdf.cell(0, 6, _etq_latin(sub), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)
    col_sku, col_qty = 150, 35
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(240, 240, 238)
    pdf.cell(col_sku, 8, "SKU", border=1, fill=True)
    pdf.cell(col_qty, 8, "Unidades", border=1, fill=True, align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    for e in items:
        label = e.get("sku") or e.get("title") or "Sin SKU"
        pdf.cell(col_sku, 7, _etq_latin(label), border=1)
        pdf.cell(col_qty, 7, str(e.get("units", 0)), border=1, align="R",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(248, 248, 246)
    pdf.cell(col_sku, 8, "TOTAL", border=1, fill=True)
    pdf.cell(col_qty, 8, str(total_units), border=1, align="R", fill=True,
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    return bytes(pdf.output())


@app.get("/etiquetas", response_class=HTMLResponse)
async def etiquetas_page(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    _r = _page_redirect(user, "etiquetas")
    if _r:
        return _r
    accounts = get_visible_accounts(user_id, user)
    accounts = await refresh_visible_account_nicknames(accounts)
    return templates.TemplateResponse("etiquetas.html", {
        "request": request, "user": user, "accounts": accounts,
        "perms": user_permissions(user),
    })


@app.get("/api/etiquetas/{account_id:int}")
async def api_etiquetas(request: Request, account_id: int):
    """Lista los envíos listos para despachar de la cuenta + resumen por SKU."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    rows, diag = await _fetch_pending_shipments(account_id, acc, token)
    ETIQUETAS_CACHE[account_id] = {"at": datetime.utcnow(), "rows": rows}
    shipments = sorted(rows.values(), key=lambda r: _etq_sku_sort_key(r["sku_principal"]))
    # Resumen por SKU: SOLO los envíos que hay que armar (listas/etiquetas/
    # demoradas). Canceladas y reprogramadas no se arman, así que no suman.
    agg: dict = {}
    for r in rows.values():
        if not r.get("packable"):
            continue
        for it in r["items"]:
            sku = it.get("sku") or it.get("title") or "Sin SKU"
            e = agg.setdefault(sku, {"sku": it.get("sku") or "", "title": it.get("title") or "", "units": 0})
            e["units"] += it["qty"]
    resumen = sorted(agg.values(), key=lambda x: _etq_sku_sort_key(x["sku"]))
    # Conteos por (grupo, sub-estado) para verificar contra el panel de ML.
    counts: dict = {}
    sub_debug: dict = {}
    for r in rows.values():
        gk, bk = r["group_key"], r["bucket_key"]
        counts.setdefault(gk, {}).setdefault(bk, 0)
        counts[gk][bk] += 1
        key = f"{r.get('logistic_type') or '-'} / {r.get('raw_status') or '-'} / {r.get('raw_substatus') or '-'}"
        sub_debug[key] = sub_debug.get(key, 0) + 1
    # Horario de corte de hoy por logística (para mostrarlo arriba del panel).
    today_wd = (datetime.utcnow() + timedelta(hours=-3)).weekday()

    def _hoy(lg):
        c = _get_cutoffs(account_id, lg).get(today_wd) or {}
        return {"dia": ETQ_WEEKDAY_NAMES[today_wd],
                "cutoff": c.get("cutoff") or "", "ventana": c.get("ventana") or ""}

    return {
        "shipments": shipments,
        "resumen": resumen,
        "total_units": sum(e["units"] for e in resumen),
        "total_shipments": len(shipments),
        "counts": counts,
        "cutoff_hoy": {"cross_docking": _hoy("cross_docking"), "self_service": _hoy("self_service")},
        "debug_substatus": sub_debug,
        "debug_filter": diag,
    }


@app.get("/api/etiquetas/{account_id:int}/cutoff_schedule")
async def api_etiquetas_get_schedule(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    if not _account_for_user(account_id, user_id):
        raise HTTPException(404)
    out = {}
    for lg in ETQ_CUTOFF_LOGISTICS:
        cutoffs = _get_cutoffs(account_id, lg)
        out[lg] = [
            {"weekday": wd, "dia": ETQ_WEEKDAY_NAMES[wd],
             "cutoff": cutoffs.get(wd, {}).get("cutoff", ""),
             "ventana": cutoffs.get(wd, {}).get("ventana", "")}
            for wd in range(7)
        ]
    return {"schedules": out}


@app.post("/api/etiquetas/{account_id:int}/cutoff_schedule")
async def api_etiquetas_save_schedule(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    if not _account_for_user(account_id, user_id):
        raise HTTPException(404)
    body = await request.json()
    schedules = body.get("schedules") or {}
    for lg in ETQ_CUTOFF_LOGISTICS:
        rows = schedules.get(lg)
        if not isinstance(rows, list):
            continue
        schedule = {}
        for item in rows:
            try:
                wd = int(item.get("weekday"))
            except (TypeError, ValueError):
                continue
            if 0 <= wd <= 6:
                schedule[wd] = {"cutoff": item.get("cutoff") or "", "ventana": item.get("ventana") or ""}
        _save_cutoffs(account_id, lg, schedule)
    ETIQUETAS_CACHE.pop(account_id, None)  # invalida cache: cambió el filtro
    return {"ok": True}


def _etq_filter_ids(account_id, ids: str) -> list:
    """Normaliza el parámetro ids (CSV) e intersecta con lo cacheado de la cuenta
    (cuando hay cache) para no mandar shipment_ids de otra cuenta a ML."""
    id_list = [s.strip() for s in (ids or "").split(",") if s.strip()]
    cache = ETIQUETAS_CACHE.get(account_id)
    rows = (cache or {}).get("rows") or {}
    if rows:
        if id_list:
            id_list = [s for s in id_list if s in rows]
        else:
            id_list = list(rows.keys())
    return id_list


# ── TODAS las cuentas: junta los envíos de todas y combina las etiquetas ──────
ETIQUETAS_ALL_CACHE: dict = {}


def _merge_pdfs(pdf_list) -> bytes:
    import io
    from pypdf import PdfReader, PdfWriter
    writer = PdfWriter()
    for b in pdf_list:
        reader = PdfReader(io.BytesIO(b))
        for page in reader.pages:
            writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _etiquetas_parts(rows: dict):
    """Arma (shipments ordenados, resumen por SKU, counts, sub_debug) desde un
    dict de filas. Compartido entre una cuenta y TODAS."""
    shipments = sorted(rows.values(), key=lambda r: _etq_sku_sort_key(r["sku_principal"]))
    agg: dict = {}
    for r in rows.values():
        if not r.get("packable"):
            continue
        for it in r["items"]:
            sku = it.get("sku") or it.get("title") or "Sin SKU"
            e = agg.setdefault(sku, {"sku": it.get("sku") or "", "title": it.get("title") or "", "units": 0})
            e["units"] += it["qty"]
    resumen = sorted(agg.values(), key=lambda x: _etq_sku_sort_key(x["sku"]))
    counts: dict = {}
    sub_debug: dict = {}
    for r in rows.values():
        gk, bk = r["group_key"], r["bucket_key"]
        counts.setdefault(gk, {}).setdefault(bk, 0)
        counts[gk][bk] += 1
        key = f"{r.get('logistic_type') or '-'} / {r.get('raw_status') or '-'} / {r.get('raw_substatus') or '-'}"
        sub_debug[key] = sub_debug.get(key, 0) + 1
    return shipments, resumen, counts, sub_debug


async def _fetch_all_accounts(user_id, user):
    """Baja los envíos de TODAS las cuentas visibles, etiquetando cada fila con su
    cuenta. Devuelve (rows_dict, info_por_cuenta)."""
    accounts = get_visible_accounts(user_id, user)
    all_rows: dict = {}
    info = []
    for acc in accounts:
        a = {"account_id": acc["id"], "nick": acc.get("nickname") or str(acc["id"]),
             "ok": False, "count": 0, "error": None}
        token = await refresh_ml_token(acc["id"])
        if not token:
            a["error"] = "sin token"
            info.append(a)
            continue
        try:
            rows, _diag = await _fetch_pending_shipments(acc["id"], acc, token)
        except Exception as e:
            a["error"] = str(e)[:140]
            info.append(a)
            continue
        for sid, row in rows.items():
            row["account_id"] = acc["id"]
            row["account_nick"] = acc.get("nickname") or str(acc["id"])
            all_rows[str(sid)] = row
        a["ok"] = True
        a["count"] = len(rows)
        info.append(a)
    return all_rows, info


@app.get("/api/etiquetas/all")
async def api_etiquetas_all(request: Request):
    """Envíos de TODAS las cuentas del usuario, juntos."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    all_rows, accounts_info = await _fetch_all_accounts(user_id, user)
    ETIQUETAS_ALL_CACHE[user_id] = {"at": datetime.utcnow(), "rows": all_rows}
    shipments, resumen, counts, sub_debug = _etiquetas_parts(all_rows)
    return {
        "shipments": shipments,
        "resumen": resumen,
        "total_units": sum(e["units"] for e in resumen),
        "total_shipments": len(shipments),
        "counts": counts,
        "cutoff_hoy": None,
        "accounts": accounts_info,
        "debug_substatus": sub_debug,
        "debug_filter": {"accounts": accounts_info},
    }


@app.post("/api/etiquetas/all/labels.pdf")
async def api_etiquetas_all_labels(request: Request):
    """Combina en UN PDF las etiquetas de los envíos seleccionados de todas las
    cuentas. GARANTÍA: si ML falla para alguna cuenta, NO devuelve un PDF parcial
    — corta con error indicando qué faltó, para no imprimir incompleto."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    body = await request.json()
    ids = [str(x).strip() for x in (body.get("ids") or []) if str(x).strip()]
    if not ids:
        raise HTTPException(400, "No hay envíos seleccionados.")
    rows = (ETIQUETAS_ALL_CACHE.get(user_id) or {}).get("rows") or {}
    by_acc: dict = {}
    unknown = []
    for sid in ids:
        row = rows.get(str(sid))
        if not row or not row.get("account_id"):
            unknown.append(str(sid))
            continue
        by_acc.setdefault(row["account_id"], []).append(str(sid))
    if unknown:
        return JSONResponse(status_code=409, content={
            "error": "Algunos envíos no están en la lista cargada. Recargá 'TODAS' y reintentá.",
            "missing_count": len(unknown), "missing": unknown[:50]})
    pdfs = []
    failures = []
    async with httpx.AsyncClient(timeout=180) as client:
        for acc_id, acc_ids in by_acc.items():
            token = await refresh_ml_token(acc_id)
            if not token:
                failures.append({"account_id": acc_id, "count": len(acc_ids), "reason": "sin token"})
                continue
            try:
                r = await client.get(
                    f"{ML_API_URL}/shipment_labels",
                    headers={"Authorization": f"Bearer {token}"},
                    params={"shipment_ids": ",".join(acc_ids), "response_type": "pdf"},
                )
            except Exception as e:
                failures.append({"account_id": acc_id, "count": len(acc_ids), "reason": str(e)[:120]})
                continue
            if r.status_code != 200 or not r.content:
                failures.append({"account_id": acc_id, "count": len(acc_ids),
                                 "reason": f"HTTP {r.status_code}: {r.text[:120]}"})
                continue
            pdfs.append(r.content)
    if failures:
        return JSONResponse(status_code=502, content={
            "error": "Faltaron etiquetas de algunas cuentas. NO se generó el PDF para que no imprimas incompleto. Reintentá.",
            "failed": failures})
    try:
        merged = _merge_pdfs(pdfs)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"No se pudieron combinar los PDF: {str(e)[:120]}"})
    return Response(content=merged, media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=etiquetas_todas.pdf"})


@app.post("/api/etiquetas/all/resumen.pdf")
async def api_etiquetas_all_resumen(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    body = await request.json()
    ids = [str(x).strip() for x in (body.get("ids") or []) if str(x).strip()]
    rows = (ETIQUETAS_ALL_CACHE.get(user_id) or {}).get("rows") or {}
    id_set = set(ids) if ids else set(rows.keys())
    agg: dict = {}
    for sid in id_set:
        row = rows.get(str(sid))
        if not row:
            continue
        for it in row["items"]:
            sku = it.get("sku") or it.get("title") or "Sin SKU"
            e = agg.setdefault(sku, {"sku": it.get("sku") or "", "title": it.get("title") or "", "units": 0})
            e["units"] += it["qty"]
    items = sorted(agg.values(), key=lambda x: _etq_sku_sort_key(x["sku"]))
    total = sum(e["units"] for e in items)
    pdf_bytes = _build_resumen_pdf(items, total, "Todas las cuentas")
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=resumen_todas.pdf"})


@app.get("/api/etiquetas/{account_id:int}/raw_shipment")
async def api_etiquetas_raw_shipment(request: Request, account_id: int, ref: str = ""):
    """Debug: vuelca los campos relevantes (sin datos del comprador) de un envío,
    buscado por N° de venta (order/pack id) o por shipment_id. Sirve para ubicar
    el campo de fecha de despacho y separar 'para hoy' de 'para mañana'."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    ref = (ref or "").strip()
    if not ref:
        raise HTTPException(400, "Falta el N° de venta o shipment_id (ref).")
    # Buscar el shipment_id en lo cargado (cache) por venta/pack/orden/shipment.
    rows = (ETIQUETAS_CACHE.get(account_id) or {}).get("rows") or {}
    sid = None
    for row in rows.values():
        if ref in (str(row.get("shipment_id")), str(row.get("venta")), str(row.get("pack_id"))) \
           or ref in [str(x) for x in (row.get("order_ids") or [])]:
            sid = row.get("shipment_id")
            break
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=30) as client:
        # Si no estaba en cache, intentamos resolver el shipment desde la orden.
        if not sid and ref.isdigit():
            try:
                ro = await client.get(f"{ML_API_URL}/orders/{ref}", headers=headers)
                if ro.status_code == 200:
                    sid = ((ro.json() or {}).get("shipping") or {}).get("id")
            except Exception:
                pass
            if not sid:
                sid = ref  # último recurso: tratar ref como shipment_id
        if not sid:
            raise HTTPException(404, "No encontré ese envío. Cargá la lista primero.")
        r = await client.get(f"{ML_API_URL}/shipments/{sid}", headers=headers)
    if r.status_code != 200:
        return JSONResponse({"shipment_id": sid, "error": r.status_code, "text": r.text[:400]})
    sh = r.json() if isinstance(r.json(), dict) else {}
    # Devolvemos el envío COMPLETO menos los datos personales del comprador/
    # vendedor, para poder ubicar cualquier campo de fecha de colecta/despacho.
    pii = {"receiver_address", "sender_address", "customer_id",
           "receiver_id", "sender_id", "comments", "shipping_items"}
    out = {k: v for k, v in sh.items() if k not in pii}
    return JSONResponse(out)


@app.get("/api/etiquetas/{account_id:int}/probe")
async def api_etiquetas_probe(request: Request, account_id: int, path: str = ""):
    """Debug: hace GET a un endpoint de la API de ML con el token de la cuenta y
    devuelve la respuesta. Solo proxea api.mercadolibre.com (no otros hosts).
    `{uid}` en el path se reemplaza por el ml_user_id de la cuenta."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    path = (path or "").strip().replace("{uid}", str(acc.get("ml_user_id") or ""))
    if not path.startswith("/"):
        raise HTTPException(400, "El path debe empezar con '/' (ej: /users/{uid}/shipping_preferences)")
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.get(f"{ML_API_URL}{path}", headers=headers)
    try:
        body = r.json()
    except Exception:
        body = r.text[:1500]
    return JSONResponse({"path": path, "status": r.status_code, "body": body})


@app.get("/api/etiquetas/{account_id:int}/labels.pdf")
async def api_etiquetas_labels(request: Request, account_id: int, ids: str = ""):
    """Devuelve el PDF combinado de etiquetas (lo genera ML), en el orden de ids."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    id_list = _etq_filter_ids(account_id, ids)
    if not id_list:
        raise HTTPException(400, "No hay envíos seleccionados para imprimir.")
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.get(
            f"{ML_API_URL}/shipment_labels",
            headers=headers,
            params={"shipment_ids": ",".join(id_list), "response_type": "pdf"},
        )
    if r.status_code != 200:
        raise HTTPException(502, f"ML no devolvió las etiquetas (HTTP {r.status_code}): {r.text[:200]}")
    return Response(
        content=r.content, media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=etiquetas.pdf"},
    )


@app.get("/api/etiquetas/{account_id:int}/resumen.pdf")
async def api_etiquetas_resumen(request: Request, account_id: int, ids: str = ""):
    """Genera el PDF de resumen de unidades por SKU (fpdf2) de los envíos dados."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    require_page(user, "etiquetas")
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    cache = ETIQUETAS_CACHE.get(account_id)
    rows = (cache or {}).get("rows") or {}
    id_list = _etq_filter_ids(account_id, ids)
    agg: dict = {}
    for sid in id_list:
        row = rows.get(str(sid))
        if not row:
            continue
        for it in row["items"]:
            sku = it.get("sku") or it.get("title") or "Sin SKU"
            e = agg.setdefault(sku, {"sku": it.get("sku") or "", "title": it.get("title") or "", "units": 0})
            e["units"] += it["qty"]
    items = sorted(agg.values(), key=lambda x: _etq_sku_sort_key(x["sku"]))
    total = sum(e["units"] for e in items)
    pdf_bytes = _build_resumen_pdf(items, total, acc.get("nickname"))
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=resumen_skus.pdf"},
    )


def _normalize_mla(value) -> str:
    """Devuelve el MLA en mayúsculas y sin espacios. Acepta también números
    sueltos asumiendo prefijo MLA."""
    if value is None:
        return ""
    s = str(value).strip().upper().replace(" ", "")
    if not s:
        return ""
    if s.isdigit():
        s = "MLA" + s
    return s


def db_save_product_discounts(account_id: int, items: list) -> int:
    if not items:
        return 0
    params = []
    for it in items:
        mla = _normalize_mla(it.get("mla"))
        if not mla:
            continue
        try:
            pct = float(it.get("discount_pct") or 0)
        except (TypeError, ValueError):
            continue
        if pct < 0 or pct > 100:
            continue
        sku = (it.get("sku") or "").strip().upper() or None
        params.append({"aid": account_id, "mla": mla, "sku": sku, "pct": pct})
    if not params:
        return 0
    with engine.connect() as conn:
        conn.execute(text(
            "INSERT INTO product_discounts (account_id, mla, sku, discount_pct, updated_at)"
            " VALUES (:aid, :mla, :sku, :pct, NOW())"
            " ON CONFLICT (account_id, mla) DO UPDATE SET"
            " sku = EXCLUDED.sku,"
            " discount_pct = EXCLUDED.discount_pct,"
            " updated_at = NOW()"
        ), params)
        conn.commit()
    return len(params)


@app.get("/api/descuentos/template")
async def api_descuentos_template(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    wb = Workbook()
    ws = wb.active
    ws.title = "Descuentos"
    ws.append(["MLA", "SKU", "Descuento %"])
    fill = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
    font = Font(bold=True)
    for col in range(1, 4):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    examples = [
        ["MLA1648019734", "SOP78-446", 20],
        ["MLA1650123456", "SOP78-446-3C", 20],
        ["MLA1650789012", "SOP68-443", 15],
    ]
    for r in examples:
        ws.append(r)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws2 = wb.create_sheet("Instrucciones")
    rows = [
        ["Plantilla de base de descuentos"],
        [""],
        ["Columnas:"],
        ["MLA", "Identificador de la publicación de Mercado Libre (ej. MLA1648019734)."],
        ["SKU", "Referencial — sirve para identificar visualmente la publicación. No se usa para matchear."],
        ["Descuento %", "Porcentaje a aplicar (0 a 100). Ej: 20 = 20% off."],
        [""],
        ["Cómo se usa:"],
        ["", "Al aplicar una promo de ML, el cruce se hace por MLA exacto."],
        ["", "Cada publicación (sea con cuotas o sin) tiene su propio MLA y su propia entrada."],
        ["", "Si tu descuento es menor al mínimo requerido por ML, te avisa."],
        ["", "Siempre vas a ver la vista previa antes de aplicar."],
    ]
    for r in rows:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 100
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_descuentos.xlsx"'},
    )


@app.get("/api/descuentos/{account_id}/export")
async def api_descuentos_export(request: Request, account_id: int):
    """Descarga la base de descuentos ACTUAL como .xlsx. Mismo formato
    que la plantilla, así podés modificar el archivo y volver a
    subirlo — el upload hace UPSERT por MLA, así que los % cambiados
    se actualizan y los MLAs nuevos se agregan."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    rows = db_fetchall(
        "SELECT mla, sku, discount_pct FROM product_discounts"
        " WHERE account_id=:aid ORDER BY mla",
        {"aid": account_id},
    )
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    wb = Workbook()
    ws = wb.active
    ws.title = "Descuentos"
    ws.append(["MLA", "SKU", "Descuento %"])
    fill = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
    font = Font(bold=True)
    for col in range(1, 4):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([
            r.get("mla") or "",
            r.get("sku") or "",
            float(r.get("discount_pct") or 0),
        ])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    # Hoja de instrucciones — recuerda al usuario cómo funciona el
    # re-upload y deja un aviso de que el formato debe mantenerse.
    ws2 = wb.create_sheet("Instrucciones")
    info = [
        ["Cómo modificar tu base desde Excel"],
        [""],
        ["1.", "Editá esta planilla: cambiá %s, agregá MLAs nuevos o sacá filas."],
        ["2.", "Mantené las columnas MLA / SKU / Descuento % (en ese orden)."],
        ["3.", "Subila desde la sección 'Base de descuentos' con 'Subir archivo'."],
        ["4.", "Las filas con MLA existente actualizan su %. Los MLAs nuevos se agregan."],
        ["5.", "Para eliminar un MLA, usá 'Eliminar' en la web o 'Eliminar TODO' + subir."],
    ]
    for r in info:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 90
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nick = (acc.get("nickname") or acc.get("ml_user_id") or "cuenta").replace(" ", "_")
    fname = f"descuentos_{nick}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/descuentos/{account_id}")
async def api_descuentos_list(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    rows = db_fetchall(
        "SELECT mla, sku, discount_pct, updated_at FROM product_discounts"
        " WHERE account_id=:aid ORDER BY mla",
        {"aid": account_id},
    )
    return {
        "items": [
            {
                "mla": r["mla"],
                "sku": r["sku"] or "",
                "discount_pct": float(r["discount_pct"]),
                "updated_at": r["updated_at"].isoformat() if r.get("updated_at") else None,
            }
            for r in rows
        ],
    }


@app.post("/api/descuentos/{account_id}")
async def api_descuentos_upsert(
    request: Request, account_id: int,
    mla: str = Form(...),
    discount_pct: float = Form(...),
    sku: Optional[str] = Form(None),
):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    saved = db_save_product_discounts(
        account_id, [{"mla": mla, "sku": sku or "", "discount_pct": discount_pct}],
    )
    return {"ok": True, "saved": saved}


@app.put("/api/descuentos/{account_id}/edit")
async def api_descuentos_edit(
    request: Request, account_id: int,
    old_mla: str = Form(...),
    mla: str = Form(...),
    discount_pct: float = Form(...),
    sku: Optional[str] = Form(None),
):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    db_execute(
        "DELETE FROM product_discounts WHERE account_id=:aid AND mla=:mla",
        {"aid": account_id, "mla": _normalize_mla(old_mla)},
    )
    saved = db_save_product_discounts(
        account_id, [{"mla": mla, "sku": sku or "", "discount_pct": discount_pct}],
    )
    return {"ok": True, "saved": saved}


@app.delete("/api/descuentos/{account_id}/{mla}")
async def api_descuentos_delete(request: Request, account_id: int, mla: str):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    db_execute(
        "DELETE FROM product_discounts WHERE account_id=:aid AND mla=:mla",
        {"aid": account_id, "mla": _normalize_mla(mla)},
    )
    return {"ok": True}


@app.post("/api/descuentos/{account_id}/bulk-delete")
async def api_descuentos_bulk_delete(request: Request, account_id: int):
    """Borra descuentos en masa. Body JSON:
      { "all": true }                 → borra TODOS los descuentos de la cuenta
      { "mlas": ["MLA1", "MLA2", ...] } → borra solo los MLAs indicados
    """
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    body = await request.json()
    if body.get("all"):
        before = db_fetchone(
            "SELECT COUNT(*) AS c FROM product_discounts WHERE account_id=:aid",
            {"aid": account_id},
        )
        db_execute(
            "DELETE FROM product_discounts WHERE account_id=:aid",
            {"aid": account_id},
        )
        return {"ok": True, "deleted": int(before["c"]) if before else 0}
    mlas = body.get("mlas") or []
    if not isinstance(mlas, list) or not mlas:
        raise HTTPException(400, "Faltan MLAs para borrar")
    normalized = [_normalize_mla(m) for m in mlas if m]
    if not normalized:
        return {"ok": True, "deleted": 0}
    # SQLAlchemy/pg8000 manejan IN con expansión de tupla — armamos el placeholder
    placeholders = ", ".join([f":m{i}" for i in range(len(normalized))])
    params = {"aid": account_id}
    for i, m in enumerate(normalized):
        params[f"m{i}"] = m
    db_execute(
        f"DELETE FROM product_discounts WHERE account_id=:aid AND mla IN ({placeholders})",
        params,
    )
    return {"ok": True, "deleted": len(normalized)}


@app.post("/api/descuentos/{account_id}/bulk-edit")
async def api_descuentos_bulk_edit(request: Request, account_id: int):
    """Actualiza el % de descuento en masa. Body JSON:
      { "mlas": ["MLA1", "MLA2", ...], "discount_pct": <0-100> }
    """
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    body = await request.json()
    mlas = body.get("mlas") or []
    try:
        pct = float(body.get("discount_pct"))
    except (TypeError, ValueError):
        raise HTTPException(400, "discount_pct inválido")
    if pct < 0 or pct > 100:
        raise HTTPException(400, "discount_pct fuera de rango (0-100)")
    if not isinstance(mlas, list) or not mlas:
        raise HTTPException(400, "Faltan MLAs para actualizar")
    normalized = [_normalize_mla(m) for m in mlas if m]
    if not normalized:
        return {"ok": True, "updated": 0}
    placeholders = ", ".join([f":m{i}" for i in range(len(normalized))])
    params = {"aid": account_id, "pct": pct}
    for i, m in enumerate(normalized):
        params[f"m{i}"] = m
    db_execute(
        f"UPDATE product_discounts SET discount_pct=:pct, updated_at=NOW()"
        f" WHERE account_id=:aid AND mla IN ({placeholders})",
        params,
    )
    return {"ok": True, "updated": len(normalized)}


@app.post("/api/descuentos/{account_id}/upload")
async def api_descuentos_upload(request: Request, account_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    form = await request.form()
    upload = form.get("file")
    if upload is None or not hasattr(upload, "read"):
        raise HTTPException(400, "Falta el archivo")
    content = await upload.read()
    filename = (getattr(upload, "filename", "") or "").lower()
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            items = _parse_excel_discounts(content)
        else:
            items = _parse_csv_discounts(content)
    except Exception as e:
        raise HTTPException(400, f"No pude leer el archivo: {e}")
    saved = db_save_product_discounts(account_id, items)
    return {"ok": True, "saved": saved, "rows_parsed": len(items)}


def _parse_excel_discounts(content: bytes) -> list:
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = None
    data_start = 0
    for idx, row in enumerate(rows[:5]):
        if row is None:
            continue
        normalized = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("mla" in n or "publicaci" in n or "item" in n for n in normalized):
            header = normalized
            data_start = idx + 1
            break
    if not header:
        return []
    mla_idx = _find_col(header, ["mla", "publicaci", "item_id", "item id", "id"])
    sku_idx = _find_col(header, ["sku"])
    pct_idx = _find_col(header, ["descuento", "discount", "off", "%"])
    if mla_idx is None or pct_idx is None:
        return []
    items = []
    for row in rows[data_start:]:
        if row is None or len(row) <= max(mla_idx, pct_idx):
            continue
        mla = row[mla_idx]
        pct = row[pct_idx]
        if mla is None or pct is None:
            continue
        try:
            pct_val = float(str(pct).replace("%", "").replace(",", ".").strip())
        except ValueError:
            continue
        sku = ""
        if sku_idx is not None and len(row) > sku_idx and row[sku_idx] is not None:
            sku = str(row[sku_idx]).strip()
        items.append({"mla": str(mla), "sku": sku, "discount_pct": pct_val})
    return items


def _parse_csv_discounts(content: bytes) -> list:
    import csv, io
    text_content = content.decode("utf-8-sig", errors="ignore")
    sample = text_content[:2048]
    sep = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text_content), delimiter=sep)
    rows = list(reader)
    if not rows:
        return []
    header = [(c or "").strip().lower() for c in rows[0]]
    mla_idx = _find_col(header, ["mla", "publicaci", "item_id", "item id", "id"])
    sku_idx = _find_col(header, ["sku"])
    pct_idx = _find_col(header, ["descuento", "discount", "off", "%"])
    if mla_idx is None or pct_idx is None:
        return []
    items = []
    for row in rows[1:]:
        if len(row) <= max(mla_idx, pct_idx):
            continue
        mla = (row[mla_idx] or "").strip()
        raw = (row[pct_idx] or "").strip().replace("%", "")
        if sep == ";":
            raw = raw.replace(",", ".")
        else:
            raw = raw.replace(",", ".")
        if not mla or not raw:
            continue
        try:
            pct = float(raw)
        except ValueError:
            continue
        sku = ""
        if sku_idx is not None and len(row) > sku_idx:
            sku = (row[sku_idx] or "").strip()
        items.append({"mla": mla, "sku": sku, "discount_pct": pct})
    return items


# Tipos de promociones que ML expone. Iteramos en paralelo porque
# /seller-promotions/promotions/search exige promotion_type.
PROMO_TYPES = [
    "DEAL",
    "PRICE_DISCOUNT",
    "MARKETPLACE_CAMPAIGN",
    "DOD",
    "LIGHTNING",
    "SMART",
    "VOLUME",
    "MULTI_BUY",
    "BANK_DEAL",
    "PRE_NEGOTIATED",
    "UNHEALTHY_STOCK",
    "PRICE_MATCHING",
    "PRICE_MATCHING_MELI_ALL",
]


def promo_offers_from_response(data) -> list:
    """Normaliza respuestas de /seller-promotions/items/{id} a lista de promos."""
    if isinstance(data, dict):
        offers = data.get("results") or data.get("offers") or []
        if not offers and (data.get("id") or data.get("promotion_id")):
            offers = [data]
        return offers if isinstance(offers, list) else []
    if isinstance(data, list):
        return data
    return []


def promo_offer_id(offer: dict):
    return (
        offer.get("promotion_id")
        or offer.get("campaign_id")
        or offer.get("id")
        or offer.get("offer_id")
    )


def promo_id_matches(a, b) -> bool:
    return str(a or "").strip().lower() == str(b or "").strip().lower()


def promo_candidate_offer_id(promo_obj: dict):
    """Devuelve el offer_id del CANDIDATO para aplicar una promo.
    En v2, ML identifica la oferta concreta del item con un `offer_id`
    (distinto del `id` de la promoción). Para promos co-fondeadas con ML
    (MARKETPLACE_CAMPAIGN / SMART / PRICE_DISCOUNT) hay que mandarlo o ML
    responde CANDIDATE_NOT_FOUND. Puede venir top-level o anidado en
    `offers[]`."""
    if not isinstance(promo_obj, dict):
        return None
    oid = promo_obj.get("offer_id") or promo_obj.get("ref_id")
    if oid:
        return oid
    offers = promo_obj.get("offers")
    if isinstance(offers, list):
        # Preferimos el candidato (status candidate/pending/vacío); si no,
        # cualquier oferta con id.
        for off in offers:
            if not isinstance(off, dict):
                continue
            st = (off.get("status") or "").lower()
            if st in ("candidate", "pending", ""):
                cand = off.get("offer_id") or off.get("id")
                if cand:
                    return cand
        for off in offers:
            if isinstance(off, dict):
                cand = off.get("offer_id") or off.get("id")
                if cand:
                    return cand
    return None


def promo_status_matches(offer_status, requested_status: str) -> bool:
    status_norm = (offer_status or "").lower()
    requested = (requested_status or "").lower()
    # ML usa: candidate = elegible (todavia no se sumo), pending/programmed =
    # ya se sumo pero la campaña no arranco (PROGRAMADA), started = activa.
    programmed_states = ("pending", "programmed", "scheduled")
    if requested == "started":
        return status_norm not in (("", "candidate") + programmed_states)
    if requested in programmed_states:
        return status_norm in programmed_states
    if requested == "candidate":
        # Elegibles = solo candidatos reales; los programados van a su pestaña.
        return status_norm in ("", "candidate")
    return True


async def fetch_all_seller_item_ids(client, headers, seller_id, debug: Optional[list] = None) -> list:
    """Lista todas las publicaciones activas del seller sin el cap de offset.

    ML documenta que, arriba de 1000 resultados, users/{id}/items/search debe
    paginarse con search_type=scan + scroll_id. Si scan no entrega scroll_id,
    caemos al offset tradicional hasta donde responda.
    """
    cache_key = str(seller_id)
    now = datetime.utcnow()
    cached = PROMO_ITEM_SCAN_CACHE.get(cache_key)
    if cached and (now - cached["at"]).total_seconds() < PROMO_ITEM_SCAN_CACHE_TTL_SECONDS:
        if debug is not None:
            debug.append(f"items/search cache hit: {len(cached['ids'])} items")
        return list(cached["ids"])

    def remember(ids: list) -> list:
        if ids:
            PROMO_ITEM_SCAN_CACHE[cache_key] = {"at": datetime.utcnow(), "ids": list(ids)}
        return list(ids)

    item_ids: list = []
    seen: set = set()
    scroll_id = None

    for _ in range(300):
        params = {"status": "active", "limit": 100, "search_type": "scan"}
        if scroll_id:
            params["scroll_id"] = scroll_id
        try:
            r = await client.get(
                f"{ML_API_URL}/users/{seller_id}/items/search",
                headers=headers,
                params=params,
            )
        except Exception as e:
            if debug is not None:
                debug.append(f"items/search scan: {str(e)[:200]}")
            break
        if r.status_code != 200:
            if debug is not None:
                debug.append(f"items/search scan: HTTP {r.status_code}: {r.text[:200]}")
            break
        data = r.json()
        results = data.get("results") or []
        if not results:
            return remember(item_ids)
        for iid in results:
            if iid not in seen:
                seen.add(iid)
                item_ids.append(iid)
        next_scroll = data.get("scroll_id")
        if not next_scroll:
            if debug is not None:
                debug.append("items/search scan no devolvio scroll_id; fallback offset")
            break
        scroll_id = next_scroll

    if item_ids and len(item_ids) >= 1000:
        return remember(item_ids)

    offset = 0
    for _ in range(300):
        try:
            r = await client.get(
                f"{ML_API_URL}/users/{seller_id}/items/search",
                headers=headers,
                params={"status": "active", "limit": 100, "offset": offset},
            )
        except Exception as e:
            if debug is not None:
                debug.append(f"items/search offset={offset}: {str(e)[:200]}")
            break
        if r.status_code != 200:
            if debug is not None:
                debug.append(f"items/search offset={offset}: HTTP {r.status_code}: {r.text[:200]}")
            break
        data = r.json()
        results = data.get("results") or []
        if not results:
            break
        for iid in results:
            if iid not in seen:
                seen.add(iid)
                item_ids.append(iid)
        if len(results) < 100:
            break
        offset += 100
    return remember(item_ids)


@app.get("/api/promociones/{account_id}")
async def api_promociones_list(request: Request, account_id: int):
    """Descubre las promociones inspeccionando el atributo 'offers' de
    cada publicación del vendedor. ML devuelve `/seller-promotions/
    promotions` con body vacío para muchas cuentas, así que esta es la
    forma confiable de saber qué promos están realmente disponibles
    o activas para este seller."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    seller_id = acc["ml_user_id"]

    debug = []
    async with httpx.AsyncClient(timeout=90) as client:
        # Paso 1: listar todos los items activos. Usamos scan+scroll para no
        # quedar capados por offset cuando hay más de 1000 publicaciones.
        all_item_ids = await fetch_all_seller_item_ids(client, headers, seller_id, debug)

        if not all_item_ids:
            return {"items": [], "errors": debug or ["No se encontraron publicaciones activas en la cuenta"]}

        # Paso 2: iterar cada item con /seller-promotions/items/{id} —
        # este endpoint devuelve TODAS las promos disponibles para ese
        # item (incluso si todavía no participa, en status="candidate").
        # Es la forma en que ML expone marketplace campaigns, deals,
        # price_discount, etc. (la misma que usa Zentor).
        promos: dict = {}
        sem = asyncio.Semaphore(10)

        async def fetch_item_promos(item_id):
            async with sem:
                try:
                    r = await client.get(
                        f"{ML_API_URL}/seller-promotions/items/{item_id}",
                        headers=headers,
                        params={"app_version": "v2"},
                    )
                    if r.status_code != 200:
                        return item_id, None, f"seller-promotions/items/{item_id}: HTTP {r.status_code}: {r.text[:120]}"
                    return item_id, r.json(), None
                except Exception as e:
                    return item_id, None, f"seller-promotions/items/{item_id}: {str(e)[:120]}"

        outcomes = await asyncio.gather(*[fetch_item_promos(i) for i in all_item_ids])

        for item_id, data, err in outcomes:
            if err:
                if len(debug) < 8:
                    debug.append(err)
                continue
            offers = promo_offers_from_response(data)

            for offer in offers:
                if not isinstance(offer, dict):
                    continue
                pid = promo_offer_id(offer)
                if not pid:
                    continue
                ptype = (offer.get("type") or offer.get("promotion_type")
                         or offer.get("campaign_type"))
                pstatus = offer.get("status")
                if pid not in promos:
                    promos[pid] = {
                        "id": pid,
                        "name": (offer.get("name") or offer.get("description")
                                 or offer.get("title") or str(pid)),
                        "type": ptype,
                        "status": pstatus,
                        "start_date": offer.get("start_date"),
                        "finish_date": offer.get("finish_date"),
                        "deadline_date": offer.get("deadline_date"),
                        "applicable_items": 0,
                        "participating_items": 0,
                    }
                promos[pid]["applicable_items"] += 1
                # Si el item ya participa (no es candidato), contarlo aparte
                if pstatus and pstatus.lower() not in ("candidate", "pending"):
                    promos[pid]["participating_items"] += 1
                # Si en una pasada llega el status "started", priorizarlo
                if pstatus == "started" and promos[pid].get("status") != "started":
                    promos[pid]["status"] = "started"

        # Paso 3: enriquecer vigencia. Muchas promos (sobre todo SMART y
        # algunas SELLER_CAMPAIGN) no traen start_date/finish_date a nivel
        # item-offer, así que consultamos el detalle global de la promo
        # /seller-promotions/promotions/{id} para completar las fechas.
        def _missing_dates(p):
            return not p.get("start_date") and not p.get("finish_date")

        async def fetch_promo_dates(pid, ptype):
            async with sem:
                try:
                    params = {"app_version": "v2"}
                    if ptype:
                        params["promotion_type"] = ptype
                    r = await client.get(
                        f"{ML_API_URL}/seller-promotions/promotions/{pid}",
                        headers=headers,
                        params=params,
                    )
                    if r.status_code != 200:
                        return pid, None
                    gdata = r.json()
                    return pid, (gdata if isinstance(gdata, dict) else None)
                except Exception:
                    return pid, None

        to_enrich = [(p["id"], p.get("type")) for p in promos.values() if _missing_dates(p)]
        if to_enrich:
            date_outcomes = await asyncio.gather(
                *[fetch_promo_dates(pid, ptype) for pid, ptype in to_enrich]
            )
            for pid, gdata in date_outcomes:
                if not gdata or pid not in promos:
                    continue
                for key in ("start_date", "finish_date", "deadline_date"):
                    if not promos[pid].get(key) and gdata.get(key):
                        promos[pid][key] = gdata.get(key)

    items = list(promos.values())
    items.sort(key=lambda x: (x.get("status") != "started", -x.get("applicable_items", 0), (x.get("name") or "").lower()))
    return {
        "items": items,
        "errors": debug,
        "items_scanned": len(all_item_ids),
    }


@app.get("/api/promociones/{account_id}/{promotion_id}/items")
async def api_promociones_items(
    request: Request, account_id: int, promotion_id: str,
    status: str = "candidate", promotion_type: Optional[str] = None,
    debug_item: Optional[str] = None,
):
    """Lista los items elegibles (candidate) o participando (started) de una promo,
    cruzados con la base de descuentos del vendedor."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    # Cache de resultados: cambiar de pestaña o recargar es instantaneo
    # mientras no se aplique/quite un descuento (lo cual invalida la cache).
    debug_item_norm = (debug_item or "").strip().upper()
    promo_type_key = (promotion_type or "").strip().upper()
    cache_key = f"{account_id}:{promotion_id}:{status}:{promo_type_key}:{debug_item_norm}"
    _now = datetime.utcnow()
    _cached = (PROMO_ITEMS_RESULT_CACHE.get(cache_key)
               if not debug_item_norm else None)
    if _cached and (_now - _cached["at"]).total_seconds() < PROMO_ITEMS_RESULT_TTL_SECONDS:
        return _cached["data"]
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    # Cargar base de descuentos local (mapeada por MLA, sin fallback)
    discount_rows = db_fetchall(
        "SELECT mla, sku, discount_pct FROM product_discounts WHERE account_id=:aid",
        {"aid": account_id},
    )
    discount_map = {
        r["mla"].upper(): {"pct": float(r["discount_pct"]), "sku": r["sku"] or ""}
        for r in discount_rows
    }

    async with httpx.AsyncClient(timeout=60) as client:
        # Detalle global de la promo (acá ML expone el default discount
        # para DEAL/SELLER_CAMPAIGN, las fechas, el meli_percentage para
        # SMART, etc.). Lo usamos como fallback cuando no aparece por
        # item.
        promo_global: dict = {}
        try:
            promo_url_params = {"app_version": "v2"}
            if promotion_type:
                promo_url_params["promotion_type"] = promotion_type
            rg = await client.get(
                f"{ML_API_URL}/seller-promotions/promotions/{promotion_id}",
                headers=headers,
                params=promo_url_params,
            )
            if rg.status_code == 200:
                gdata = rg.json()
                if isinstance(gdata, dict):
                    promo_global = gdata
        except Exception:
            promo_global = {}

        promo_page_limit = 49
        params = {"app_version": "v2", "limit": promo_page_limit}
        if status:
            params["status"] = status
        if promotion_type:
            params["promotion_type"] = promotion_type

        all_results = []
        seen_ids: set = set()
        expected_total = 0
        total_raw = 0
        source = "promotion_items"
        direct_error = None

        offset = 0
        duplicate_streak = 0
        max_iterations = 300
        for _ in range(max_iterations):
            try:
                r = await client.get(
                    f"{ML_API_URL}/seller-promotions/promotions/{promotion_id}/items",
                    headers=headers,
                    params={**params, "offset": offset},
                )
            except Exception as e:
                direct_error = f"ML error: {str(e)[:200]}"
                break
            if r.status_code != 200:
                direct_error = f"ML {r.status_code}: {r.text[:200]}"
                break
            data = r.json()
            results = data.get("results", []) if isinstance(data, dict) else []
            paging = data.get("paging") or {}
            if paging.get("total"):
                try:
                    expected_total = int(paging["total"])
                except (TypeError, ValueError):
                    pass
            if not results:
                break
            total_raw += len(results)
            new_in_batch = 0
            for it in results:
                iid = it.get("id") if isinstance(it, dict) else None
                if iid and iid not in seen_ids:
                    seen_ids.add(iid)
                    all_results.append(it)
                    new_in_batch += 1
            offset += len(results)
            if len(results) < promo_page_limit:
                break
            if expected_total > 0 and len(seen_ids) >= expected_total:
                break
            duplicate_streak = duplicate_streak + 1 if new_in_batch == 0 else 0
            if duplicate_streak >= 5:
                break

        # CONVERGENCIA RÁPIDA: el endpoint LISTA es eventualmente consistente —
        # para un mismo item devuelve a veces el mínimo genérico (10%) y a veces
        # el real de ESE item (5%). En UNA sola lectura solo "aciertan" algunos
        # items; el resto queda en 10% hasta que el usuario refresca varias
        # veces. Para no depender de la suerte, re-consultamos la lista unas
        # pocas veces dentro del mismo request y vamos fijando el PISO por item
        # (el descuento más chico = precio permitido más alto). Así, en una sola
        # carga, casi todos los items convergen a su mínimo real.
        def _update_floor_from_list_item(it):
            if not isinstance(it, dict):
                return
            iid = it.get("id")
            if not iid:
                return
            try:
                o = float(it.get("original_price"))
                p = float(it.get("max_discounted_price"))
            except (TypeError, ValueError):
                return
            if o <= 0 or p <= 0 or p >= o:
                return
            pct = round((1 - p / o) * 100, 2)
            if pct <= 0:
                return
            _record_promo_floor(promotion_id, iid, pct)

        # Primera pasada (la que ya teníamos) alimenta el piso.
        for it in all_results:
            _update_floor_from_list_item(it)

        # CONVERGENCIA SINCRÓNICA (primera carga de la promo en este proceso):
        # la lista de ML es eventualmente consistente — para un mismo item
        # devuelve a veces el mínimo genérico (10%) y a veces el real (5%), y la
        # variedad solo aparece en el TIEMPO porque ML cachea su respuesta unos
        # segundos. Antes esto se "convergía" en background y el usuario tenía
        # que refrescar varias veces hasta que aparecía el 5% — poco confiable.
        # Ahora poleamos la lista acá mismo, con pausas (para que el cache de ML
        # rote), fijando el PISO por item hasta que se estabiliza (sin cambios 2
        # pasadas seguidas) o se agota el presupuesto. Así la página sale BIEN de
        # una. Las cargas siguientes ya tienen el piso (en memoria + DB) y se
        # marcan como convergidas, por lo que salen instantáneas.
        converge_key = f"{account_id}:{promotion_id}:{status}:{promo_type_key}"
        if converge_key not in PROMO_CONVERGED and status in ("candidate", "started"):
            MAX_CONVERGE_PASSES = 8     # techo de pasadas
            CONVERGE_DELAY = 2.5        # s entre pasadas (> cache de ML)
            STABLE_NEEDED = 2           # pasadas seguidas sin cambios = convergió
            stable_passes = 0
            for _pass in range(MAX_CONVERGE_PASSES):
                await asyncio.sleep(CONVERGE_DELAY)
                changed = False
                p_offset = 0
                for _ in range(max_iterations):
                    try:
                        rp = await client.get(
                            f"{ML_API_URL}/seller-promotions/promotions/{promotion_id}/items",
                            headers=headers,
                            params={**params, "offset": p_offset},
                        )
                    except Exception:
                        break
                    if rp.status_code != 200:
                        break
                    pdata = rp.json()
                    presults = pdata.get("results", []) if isinstance(pdata, dict) else []
                    if not presults:
                        break
                    for it in presults:
                        if not isinstance(it, dict) or not it.get("id"):
                            continue
                        try:
                            o = float(it.get("original_price"))
                            p = float(it.get("max_discounted_price"))
                        except (TypeError, ValueError):
                            continue
                        if o <= 0 or p <= 0 or p >= o:
                            continue
                        pct = round((1 - p / o) * 100, 2)
                        if pct > 0 and _record_promo_floor(promotion_id, it.get("id"), pct):
                            changed = True
                    p_offset += len(presults)
                    if len(presults) < promo_page_limit:
                        break
                    if expected_total > 0 and p_offset >= expected_total:
                        break
                stable_passes = 0 if changed else stable_passes + 1
                if stable_passes >= STABLE_NEEDED:
                    break
            PROMO_CONVERGED.add(converge_key)

        # Enriquecer con SKU y nombre desde /items/{id}. Sin filtrar
        # attributes para que vengan también `attributes[]` y `variations[]`,
        # donde muchas publicaciones tienen el SELLER_SKU.
        async def enrich(item_id):
            try:
                ri = await client.get(
                    f"{ML_API_URL}/items/{item_id}",
                    headers=headers,
                    params={"attributes": "id,title,seller_sku,price,available_quantity,attributes,variations"},
                )
                if ri.status_code == 200:
                    return ri.json()
            except Exception:
                pass
            return None

        # Detalle completo de la promo por item — el endpoint que lista los
        # items de la promo viene minimalista (sin top_deal_price ni
        # discount_percentage). El endpoint por item devuelve TODAS las
        # promos disponibles para el item, así que filtramos por el
        # promotion_id actual y mergeamos los campos faltantes.
        def matching_promo_offer(data) -> Optional[dict]:
            for off in promo_offers_from_response(data):
                if not isinstance(off, dict):
                    continue
                pid = promo_offer_id(off)
                if promo_id_matches(pid, promotion_id) and promo_status_matches(off.get("status"), status):
                    return off
            return None

        async def fetch_item_promo_detail(item_id):
            try:
                # CLAVE: ML devuelve un `max_discounted_price` DISTINTO segun
                # los filtros. La query FILTRADA (con promotion_id + status +
                # promotion_type) devuelve el minimo REAL por item de ESA promo
                # (ej: 19950 = 5%), que es EXACTAMENTE lo que muestra el panel
                # de ML. La query SIN filtros devuelve un valor generico mas
                # agresivo (ej: 18900 = 10%) que NO coincide con el panel.
                # Por eso la filtrada va PRIMERO: en el merge, sus precios
                # mandan sobre la generica.
                exact_params = {"app_version": "v2"}
                if status in ("candidate", "started"):
                    exact_params["status"] = status
                if promotion_id:
                    exact_params["promotion_id"] = promotion_id
                if promotion_type:
                    exact_params["promotion_type"] = promotion_type
                query_variants = [exact_params]
                if exact_params != {"app_version": "v2"}:
                    query_variants.append({"app_version": "v2"})

                variants: list[dict] = []
                for params in query_variants:
                    ri = await client.get(
                        f"{ML_API_URL}/seller-promotions/items/{item_id}",
                        headers=headers,
                        params=params,
                    )
                    if ri.status_code != 200:
                        continue
                    match = matching_promo_offer(ri.json())
                    if isinstance(match, dict):
                        variants.append(match)

                if not variants:
                    return None

                merged = dict(variants[0])
                for variant in variants[1:]:
                    for k, v in variant.items():
                        if merged.get(k) in (None, "", 0) and v not in (None, ""):
                            merged[k] = v
                merged["_detail_variants"] = variants
                return merged
            except Exception:
                pass
            return None

        scan_item_count = 0
        scan_matched_count = 0
        promo_details_override = []
        detail_sem = asyncio.Semaphore(40)

        if all_results:
            async def fetch_existing_detail(iid):
                async with detail_sem:
                    return await fetch_item_promo_detail(iid)

            direct_ids = [it.get("id") for it in all_results if isinstance(it, dict) and it.get("id")]
            promo_details_override = await asyncio.gather(
                *[fetch_existing_detail(i) for i in direct_ids]
            )
            promo_details_override = [
                detail if isinstance(detail, dict) else {}
                for detail in promo_details_override
            ]
        else:
            source = "seller_items_scan"
            seller_item_ids = await fetch_all_seller_item_ids(client, headers, acc["ml_user_id"])
        if not all_results and seller_item_ids:
            scan_item_count = len(seller_item_ids)

            async def fetch_matching_detail(iid):
                async with detail_sem:
                    detail = await fetch_item_promo_detail(iid)
                if not isinstance(detail, dict):
                    return None
                pid = promo_offer_id(detail)
                if not promo_id_matches(pid, promotion_id):
                    return None
                if not promo_status_matches(detail.get("status"), status):
                    return None
                return {"id": iid}, detail

            pairs = await asyncio.gather(*[fetch_matching_detail(i) for i in seller_item_ids])
            matched = [p for p in pairs if p]
            all_results = [base for base, _ in matched]
            promo_details_override = [detail for _, detail in matched]
            scan_matched_count = len(matched)

        ids = [it.get("id") for it in all_results if it.get("id")]
        # Multiget en lotes de 20 — MUCHO mas rapido que 1 request por item
        # (de N requests a N/20). ML devuelve [{code, body}, ...].
        enriched_map: dict = {}
        mget_sem = asyncio.Semaphore(12)
        _attrs = "id,title,seller_sku,price,available_quantity,attributes,variations"

        async def fetch_batch(chunk):
            async with mget_sem:
                try:
                    rb = await client.get(
                        f"{ML_API_URL}/items",
                        headers=headers,
                        params={"ids": ",".join(chunk), "attributes": _attrs},
                    )
                    if rb.status_code != 200:
                        return
                    for entry in rb.json():
                        if not isinstance(entry, dict):
                            continue
                        body = entry.get("body") if entry.get("code") == 200 else None
                        if isinstance(body, dict) and body.get("id"):
                            enriched_map[body["id"]] = body
                except Exception:
                    pass

        chunks = [ids[i:i + 20] for i in range(0, len(ids), 20)]
        await asyncio.gather(*[fetch_batch(c) for c in chunks])
        enriched = [enriched_map.get(i) for i in ids]
        promo_details = promo_details_override

    def _extract_sku(info: Optional[dict]) -> str:
        """Saca el SELLER_SKU del item. ML lo expone en distintos lugares:
        - `seller_sku` top-level (legacy)
        - `attributes[]` con id='SELLER_SKU' (la mayoría de las publicaciones
          actuales lo guardan acá)
        - `variations[].seller_sku` o `variations[].attributes[]` con
          id='SELLER_SKU' (para items con variantes)."""
        if not info:
            return ""
        sku = (info.get("seller_sku") or "").strip()
        if sku:
            return sku
        for attr in (info.get("attributes") or []):
            if isinstance(attr, dict) and attr.get("id") in ("SELLER_SKU", "SELLER_CUSTOM_FIELD"):
                v = (attr.get("value_name") or attr.get("value") or "").strip()
                if v:
                    return v
        for var in (info.get("variations") or []):
            if not isinstance(var, dict):
                continue
            v = (var.get("seller_sku") or "").strip()
            if v:
                return v
            for attr in (var.get("attributes") or []):
                if isinstance(attr, dict) and attr.get("id") in ("SELLER_SKU", "SELLER_CUSTOM_FIELD"):
                    vv = (attr.get("value_name") or attr.get("value") or "").strip()
                    if vv:
                        return vv
        return ""

    def _pct_from_price(price, original) -> Optional[float]:
        """Convierte un precio post-descuento a porcentaje de descuento."""
        if price is None or original is None:
            return None
        try:
            p, o = float(price), float(original)
        except (TypeError, ValueError):
            return None
        if o <= 0 or p < 0 or p >= o:
            return None
        return round((1 - p / o) * 100, 2)

    def _pct_from_amount(amount, original) -> Optional[float]:
        """Convierte un monto de descuento a porcentaje sobre el precio base."""
        if amount is None or original is None:
            return None
        try:
            a, o = float(amount), float(original)
        except (TypeError, ValueError):
            return None
        if o <= 0 or a <= 0 or a >= o:
            return None
        return round(a / o * 100, 2)

    def _extract_min_discount_pct(promo_item: dict, original_price) -> Optional[float]:
        """% MÍNIMO de descuento que ML exige para participar.

        El mínimo REAL por item lo da `max_discounted_price`: el precio MÁS
        ALTO que ML permite cobrar con descuento. Convertido a % sobre el
        precio original = el descuento mínimo que tenés que hacer para entrar.
        Esto varía item por item, por eso es la fuente principal.

        OJO: NO usamos `discount_percentage`, `default_discount_percentage`,
        `seller_percentage`, etc. Esos son el descuento NOMINAL de la campaña
        (un valor fijo a nivel campaña) o el aporte/sugerido — NO el mínimo
        requerido por item. Usarlos hace que todos los items muestren el mismo
        número plano (ej: 10%), que fue justamente la regresión a corregir.
        """
        if not isinstance(promo_item, dict):
            return None

        # Precios post-descuento → mínimo requerido (fuente principal, por item)
        price_fields = (
            "max_discounted_price",
            "top_deal_price",
        )
        # Solo porcentajes que representan EXPLÍCITAMENTE un mínimo requerido.
        pct_fields = (
            "min_discount_percentage", "min_discount_pct",
            "minimum_discount_percentage",
            "minimum_percentage", "min_percentage",
            "required_discount_percentage", "required_percentage",
            "seller_min_discount_percentage",
            "min_seller_discount_percentage",
        )

        def pct_value(value) -> Optional[float]:
            if value is None:
                return None
            try:
                pct = float(value)
            except (TypeError, ValueError):
                return None
            if 0 < pct < 100:
                return round(pct, 2)
            return None

        def from_price_fields(obj: dict) -> Optional[float]:
            best = None
            for k in price_fields:
                pct = _pct_from_price(obj.get(k), original_price)
                if pct is not None and pct > 0:
                    best = pct if best is None else min(best, pct)
            return best

        def from_pct_fields(obj: dict) -> Optional[float]:
            for k in pct_fields:
                pct = pct_value(obj.get(k))
                if pct is not None:
                    return pct
            return None

        containers = [promo_item]
        for nest_key in ("discount_breakdown", "prices", "benefits",
                         "nudge", "rebate"):
            nested = promo_item.get(nest_key)
            if isinstance(nested, dict):
                containers.append(nested)
        for arr_key in ("offers", "discounts", "rebates"):
            arr = promo_item.get(arr_key)
            if isinstance(arr, list):
                containers.extend(o for o in arr if isinstance(o, dict))

        # 1) Mínimo real por precio (max_discounted_price)
        price_candidates = [p for p in (from_price_fields(c) for c in containers)
                            if p is not None]
        if price_candidates:
            return min(price_candidates)

        # 2) Fallback: % explícito de mínimo requerido
        for c in containers:
            pct = from_pct_fields(c)
            if pct is not None:
                return pct

        return None

    def _extract_seller_suggested_pct(
        promo_item: dict, original_price, ml_contribution_pct
    ) -> Optional[float]:
        """% MÍNIMO que ML pide al VENDEDOR para participar de la promo.
        NO la recomendación / `suggested_discounted_price` — ese campo
        es el descuento que ML te sugiere para vender mejor, no el
        umbral de aceptación. La data real que necesitamos es:

        1. SMART/co-fondeada: `seller_percentage` (lo que ML pide al
           vendedor). Ya viene neto, no hay que restar el aporte.
        2. DEAL/SELLER_CAMPAIGN: `max_discounted_price` (precio MÁS
           ALTO permitido con descuento) → convertir a % vs original
           da el descuento MÍNIMO aceptable. Si ML aporta algo, lo
           restamos para quedarnos con la parte del vendedor.
        """
        # 1. Campo directo del vendedor (SMART)
        for k in ("seller_percentage", "seller_min_discount_percentage",
                  "min_seller_discount_percentage",
                  "seller_discount_percentage",
                  "nudge_seller_percentage"):
            v = promo_item.get(k)
            if v is not None:
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                if fv > 0:
                    return round(fv, 2)

        def _as_seller_part(total_pct):
            if total_pct is None or total_pct <= 0:
                return None
            if ml_contribution_pct and ml_contribution_pct > 0:
                return round(max(0.0, total_pct - ml_contribution_pct), 2)
            return round(total_pct, 2)

        # 2. DEAL: precios → % MÍNIMOS (max_discounted_price = precio
        # más alto con descuento = mínimo descuento)
        for k in ("max_discounted_price", "top_deal_price",
                  "minimum_price", "min_price"):
            pct = _pct_from_price(promo_item.get(k), original_price)
            res = _as_seller_part(pct)
            if res is not None and res > 0:
                return res
        # 3. % mínimos explícitos
        for k in ("min_discount_percentage", "min_discount_pct",
                  "minimum_discount_percentage",
                  "default_discount_percentage",
                  "expected_discount_percentage",
                  "target_discount_percentage"):
            v = promo_item.get(k)
            if v is not None:
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                res = _as_seller_part(fv)
                if res is not None and res > 0:
                    return res
        # 4. Dentro de offers[] / discounts[] / rebates[]
        for arr_key in ("offers", "discounts", "rebates"):
            arr = promo_item.get(arr_key)
            if isinstance(arr, list):
                for off in arr:
                    if not isinstance(off, dict):
                        continue
                    # Seller directo dentro del offer
                    for pk in ("seller_percentage", "seller_discount_percentage",
                               "min_seller_discount_percentage"):
                        v = off.get(pk)
                        if v is not None:
                            try:
                                fv = float(v)
                            except (TypeError, ValueError):
                                continue
                            if fv > 0:
                                return round(fv, 2)
                    # Precios mínimos → totales
                    for pk in ("max_discounted_price", "top_deal_price"):
                        pct = _pct_from_price(off.get(pk), original_price)
                        res = _as_seller_part(pct)
                        if res is not None and res > 0:
                            return res
                    # % mínimos explícitos
                    for pk in ("min_discount_percentage",
                               "minimum_discount_percentage"):
                        v = off.get(pk)
                        if v is not None:
                            try:
                                fv = float(v)
                            except (TypeError, ValueError):
                                continue
                            res = _as_seller_part(fv)
                            if res is not None and res > 0:
                                return res
        # 5. Objetos anidados
        for nest_key in ("nudge", "discount_breakdown", "rebate", "prices"):
            nested = promo_item.get(nest_key)
            if isinstance(nested, dict):
                for k in ("seller_percentage", "seller_discount_percentage",
                          "min_seller_discount_percentage"):
                    v = nested.get(k)
                    if v is not None:
                        try:
                            fv = float(v)
                        except (TypeError, ValueError):
                            continue
                        if fv > 0:
                            return round(fv, 2)
                for k in ("min_discount_percentage",
                          "minimum_discount_percentage",
                          "minimum_percentage", "min_percentage"):
                    v = nested.get(k)
                    if v is not None:
                        try:
                            fv = float(v)
                        except (TypeError, ValueError):
                            continue
                        res = _as_seller_part(fv)
                        if res is not None and res > 0:
                            return res
        return None

    # Fallback global: si los items vienen sin info, tomamos lo que
    # diga la promo entera. Para DEAL puede que no aplique (el minimo
    # suele venir por item), pero para SMART puede traer meli_percentage global.
    default_min_pct = _extract_min_discount_pct(promo_global, None) if promo_global else None
    default_meli_pct = None
    if promo_global:
        for k in ("meli_percentage", "default_meli_percentage", "ml_percentage"):
            v = promo_global.get(k)
            if v is not None:
                try:
                    default_meli_pct = float(v)
                    break
                except (TypeError, ValueError):
                    continue
    items_out = []
    for promo_item, info, detail in zip(all_results, enriched, promo_details):
        item_id = promo_item.get("id")
        # Mergeamos detail (per-item endpoint) sobre promo_item.
        merged = dict(promo_item or {})
        if isinstance(detail, dict):
            for k, v in detail.items():
                if merged.get(k) in (None, "", 0) and v not in (None, ""):
                    merged[k] = v
        original_price = (merged.get("original_price")
                          or (info.get("price") if info else None)
                          or merged.get("regular_price")
                          or merged.get("price"))
        # Minimo requerido por ML para participar EN ESTA PROMO, por item.
        #
        # FUENTE CORRECTA: el detalle per-item consultado con FILTROS de la
        # promo (fetch_item_promo_detail ya pide promotion_id+status+type). Esa
        # query devuelve el `max_discounted_price` REAL por item de esta promo
        # (ej: 19950 = 5%), idéntico al panel de ML. El listado de la promo
        # (`merged`/`promo_item`) trae un valor generico distinto (ej: 18900 =
        # 10%) que NO coincide con el panel — por eso el detalle FILTRADO tiene
        # prioridad y el listado queda solo como respaldo.
        #
        # Es UNA sola lectura determinista (no min() sobre variantes), asi que
        # el mismo item siempre da el mismo numero.
        #
        # ROBUSTEZ: ML puede devolver el offer de la promo en cualquiera de
        # las variantes de query (filtrada / sin filtros) y, según el orden
        # del merge, el genérico (18900=10%) podría tapar al real (19950=5%).
        # Para que NUNCA gane el genérico, juntamos el % de TODAS las fuentes
        # (cada variante per-item + el merge + el listado) y nos quedamos con
        # el MÍNIMO: el "mínimo requerido para participar" es justamente el
        # descuento más chico aceptado (= precio permitido más alto). Así da
        # 5% aunque alguna query traiga el 10% genérico.
        min_sources = []
        if isinstance(detail, dict):
            for variant in (detail.get("_detail_variants") or []):
                if isinstance(variant, dict):
                    min_sources.append(variant)
            min_sources.append(detail)
        min_sources.append(merged)
        min_candidates = []
        for src in min_sources:
            if isinstance(src, dict) and src:
                p = _extract_min_discount_pct(src, original_price)
                if p is not None:
                    min_candidates.append(p)
        min_discount_pct = min(min_candidates) if min_candidates else None
        if min_discount_pct is None and default_min_pct is not None:
            min_discount_pct = default_min_pct
        # PISO HISTÓRICO: la API de ML parpadea (mismo item da 10% en una
        # llamada y 5% en la siguiente). El mínimo REAL para participar es el
        # más bajo que ML aceptó alguna vez. Guardamos ese piso por item y lo
        # usamos siempre, así el valor queda estable y no vuelve a subir.
        seen_key = _promo_floor_key(promotion_id, item_id)
        if min_discount_pct is not None:
            _record_promo_floor(promotion_id, item_id, min_discount_pct)
            min_discount_pct = PROMO_MIN_PCT_SEEN.get(seen_key, min_discount_pct)
        elif PROMO_MIN_PCT_SEEN.get(seen_key) is not None:
            min_discount_pct = PROMO_MIN_PCT_SEEN[seen_key]
        title = ""
        sku_from_ml = _extract_sku(info)
        if info:
            title = info.get("title") or ""
        # Cruce DIRECTO por MLA (item_id). Sin fallback de SKU/cuotas.
        match = discount_map.get((item_id or "").upper())
        loaded_pct = match["pct"] if match else None
        sku_base = match["sku"] if match else ""
        sku = sku_base or sku_from_ml
        # Aporte ML (co-fondeado). Necesario PRIMERO porque la
        # sugerencia del vendedor depende de él.
        ml_contribution_pct = None
        meli_pct = merged.get("meli_percentage")
        if meli_pct is not None:
            try:
                ml_contribution_pct = float(meli_pct)
            except (TypeError, ValueError):
                ml_contribution_pct = None
        elif merged.get("meli_amount") and original_price:
            try:
                ml_contribution_pct = round(
                    float(merged["meli_amount"]) / float(original_price) * 100, 2)
            except (TypeError, ValueError):
                ml_contribution_pct = None
        if ml_contribution_pct is None and default_meli_pct is not None:
            ml_contribution_pct = default_meli_pct
        # Compat legacy: estos campos existen en respuestas anteriores, pero
        # no representan el sugerido/optimo de ML. Los espejamos al minimo para
        # que ningun consumidor viejo vuelva a mostrar un recomendado como
        # requisito operativo.
        seller_suggested_pct = min_discount_pct
        suggested_pct = min_discount_pct
        # `seller_pct` = lo que se va a aplicar. Priorizamos lo que
        # cargó el vendedor; si no hay, usamos solamente el mínimo de ML
        # (min + 0.1). No usamos sugeridos/óptimos como fallback.
        if loaded_pct is not None:
            seller_pct = loaded_pct
        elif min_discount_pct is not None:
            seller_pct = round(min_discount_pct + 0.1, 2)
        else:
            seller_pct = 0.0
        # `final_pct` (descuento total que ve el comprador) = tu % + aporte ML
        final_pct = seller_pct + (ml_contribution_pct or 0.0)
        # GAP: cuánto le falta al VENDEDOR para llegar al mínimo de ML.
        # NO descontamos el aporte ML — el mínimo es lo que ML te pide
        # al vendedor.
        below_min = False
        gap_pct = None
        # Para el gap usamos sólo lo que el vendedor cargó (loaded_pct),
        # no la sugerencia. Si no cargaste nada, te falta TODO.
        check_pct = loaded_pct if loaded_pct is not None else 0.0
        if min_discount_pct is not None:
            required_pct = round(min_discount_pct + 0.1, 2)
            gap_pct = round(max(0.0, required_pct - check_pct), 2)
            if check_pct <= min_discount_pct:
                below_min = True
        final_price = None
        if original_price:
            try:
                final_price = round(float(original_price) * (1 - final_pct / 100.0), 2)
            except Exception:
                final_price = None
        items_out.append({
            "item_id": item_id,
            "mla": item_id,
            "sku": sku,
            "title": title,
            "original_price": float(original_price) if original_price else None,
            "min_discount_pct": min_discount_pct,
            "suggested_discount_pct": suggested_pct,
            "seller_suggested_pct": seller_suggested_pct,
            "loaded_discount_pct": loaded_pct,
            "seller_discount_pct": seller_pct,
            "ml_contribution_pct": ml_contribution_pct,
            "final_discount_pct": round(final_pct, 2),
            "gap_pct": gap_pct,
            "final_price": final_price,
            "below_min": below_min,
            "ml_contribution": ml_contribution_pct,
            "status": merged.get("status"),
            "promotion_type": merged.get("promotion_type") or merged.get("type"),
            # Para el apply: ML pide offer_id en SMART y top_deal_price
            # en DEAL. Lo dejamos disponible para el frontend.
            "offer_id": merged.get("offer_id") or merged.get("ref_id"),
        })
    # Para diagnóstico: devolvemos las keys del primer item que devolvió
    # ML tanto en la lista como en el endpoint per-item. Así detectamos
    # nombres de campos no contemplados.
    raw_sample = {
        "code_version": "sync-converge-v7",
        "promo_global_keys": sorted(list(promo_global.keys())) if isinstance(promo_global, dict) and promo_global else None,
        "promo_global_sample": ({k: promo_global[k] for k in list(promo_global.keys())[:30]}
                                if isinstance(promo_global, dict) and promo_global else None),
        "default_min_pct": default_min_pct,
        "default_meli_pct": default_meli_pct,
        # Diagnóstico de paginación
        "paging_expected_total": expected_total,  # lo que dice ML que hay
        "paging_returned_raw": total_raw,         # items totales con dupes
        "paging_unique": len(seen_ids),           # items únicos
        "scan_item_count": scan_item_count,
        "scan_matched_count": scan_matched_count,
        "source": source,
        "direct_error": direct_error,
    }
    if all_results:
        first_list = all_results[0]
        first_detail = promo_details[0] if promo_details else None
        raw_sample.update({
            "list_keys": sorted(list(first_list.keys())) if isinstance(first_list, dict) else None,
            "list_sample": ({k: first_list[k] for k in list(first_list.keys())[:30]}
                            if isinstance(first_list, dict) else None),
            "detail_keys": sorted(list(first_detail.keys())) if isinstance(first_detail, dict) else None,
            "detail_sample": ({k: first_detail[k] for k in list(first_detail.keys())[:30]}
                              if isinstance(first_detail, dict) else None),
        })
    if debug_item_norm:
        target_idx = next(
            (i for i, it in enumerate(all_results)
             if str((it or {}).get("id") or "").upper() == debug_item_norm),
            None,
        )
        target_processed = next(
            (it for it in items_out
             if str(it.get("item_id") or "").upper() == debug_item_norm),
            None,
        )
        target_list = all_results[target_idx] if target_idx is not None else None
        target_detail = promo_details[target_idx] if target_idx is not None and target_idx < len(promo_details) else None
        raw_sample.update({
            "debug_item": debug_item_norm,
            "debug_item_found": target_idx is not None,
            "debug_item_processed": target_processed,
            "debug_item_list_keys": sorted(list(target_list.keys())) if isinstance(target_list, dict) else None,
            "debug_item_list_raw": target_list if isinstance(target_list, dict) else None,
            "debug_item_detail_keys": sorted(list(target_detail.keys())) if isinstance(target_detail, dict) else None,
            "debug_item_detail_raw": target_detail if isinstance(target_detail, dict) else None,
        })
        # Desglose EXACTO de qué % calcula cada fuente para este item, con el
        # precio original usado. Así, con UN solo paste, se ve si el deploy es
        # min-across-sources-v2 y qué valor da cada origen (filtrada vs genérica).
        try:
            tgt_orig = None
            tgt_info = enriched[target_idx] if (target_idx is not None and target_idx < len(enriched)) else None
            if isinstance(target_detail, dict):
                tgt_orig = (target_detail.get("original_price")
                            or (tgt_info.get("price") if tgt_info else None)
                            or target_detail.get("regular_price")
                            or target_detail.get("price"))
            elif tgt_info:
                tgt_orig = tgt_info.get("price")
            src_breakdown = []
            srcs = []
            if isinstance(target_detail, dict):
                for vi, variant in enumerate(target_detail.get("_detail_variants") or []):
                    if isinstance(variant, dict):
                        srcs.append((f"detail_variant_{vi}", variant))
                srcs.append(("detail_merged", target_detail))
            if isinstance(target_list, dict):
                srcs.append(("list", target_list))
            for name, src in srcs:
                src_breakdown.append({
                    "source": name,
                    "max_discounted_price": src.get("max_discounted_price"),
                    "top_deal_price": src.get("top_deal_price"),
                    "original_price": src.get("original_price"),
                    "pct": _extract_min_discount_pct(src, tgt_orig),
                })
            raw_sample["debug_item_min_sources"] = {
                "original_price_used": tgt_orig,
                "computed_min_pct": (min([b["pct"] for b in src_breakdown if b["pct"] is not None])
                                     if any(b["pct"] is not None for b in src_breakdown) else None),
                "min_seen_floor": PROMO_MIN_PCT_SEEN.get(f"{promotion_id}:{debug_item_norm}"),
                "sources": src_breakdown,
            }
        except Exception as ee:
            raw_sample["debug_item_min_sources_error"] = str(ee)[:200]
        # Dump COMPLETO y SIN TOCAR de lo que ML responde para este item, en
        # todas las variantes de query. Sin matching, sin merge: la respuesta
        # cruda tal cual. Asi vemos TODOS los campos/ofertas que ML expone
        # (incluido el minimo real por item que pueda no estar en max_discounted_price).
        try:
            async with httpx.AsyncClient(timeout=30) as dbg_client:
                dbg_queries = {
                    "nofilter": {"app_version": "v2"},
                    "filtered": {"app_version": "v2", "status": status,
                                 "promotion_id": promotion_id,
                                 "promotion_type": promotion_type},
                    "plain": {},
                }
                dbg_out = {}
                for name, qp in dbg_queries.items():
                    qp_clean = {k: v for k, v in qp.items() if v}
                    try:
                        rr = await dbg_client.get(
                            f"{ML_API_URL}/seller-promotions/items/{debug_item_norm}",
                            headers=headers, params=qp_clean)
                        dbg_out[name] = {
                            "status_code": rr.status_code,
                            "params": qp_clean,
                            "body": (rr.json() if rr.status_code == 200 else rr.text[:500]),
                        }
                    except Exception as ee:
                        dbg_out[name] = {"error": str(ee)[:200], "params": qp_clean}
                raw_sample["debug_item_ml_raw_responses"] = dbg_out
        except Exception as ee:
            raw_sample["debug_item_ml_raw_responses_error"] = str(ee)[:200]
    result = {
        "items": items_out,
        "discount_base_count": len(discount_map),
        "raw_sample": raw_sample,
    }
    PROMO_ITEMS_RESULT_CACHE[cache_key] = {"at": datetime.utcnow(), "data": result}
    return result


@app.post("/api/promociones/{account_id}/{promotion_id}/apply")
async def api_promociones_apply(
    request: Request, account_id: int, promotion_id: str,
):
    """Aplica los descuentos a los items seleccionados.
    Body JSON: { items: [{item_id, discount_pct, promotion_type?}, ...] }
    """
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    body = await request.json()
    items = body.get("items") or []
    if not items:
        raise HTTPException(400, "Faltan items para aplicar")
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    results = []
    async with httpx.AsyncClient(timeout=60) as client:
        # ML bloquea temporalmente ofertas cuando procesa cambios masivos.
        # Concurrencia baja + retries reduce HTTP 423 LockedEntityException.
        sem = asyncio.Semaphore(3)

        async def resolve_offer_id(iid, ptype):
            """Consulta en vivo el offer_id del candidato para esta promo.
            Necesario para promos co-fondeadas con ML: si no se manda el
            offer_id correcto, ML responde CANDIDATE_NOT_FOUND."""
            try:
                params = {"app_version": "v2", "promotion_id": promotion_id}
                if ptype:
                    params["promotion_type"] = ptype
                r = await client.get(
                    f"{ML_API_URL}/seller-promotions/items/{iid}",
                    headers=headers,
                    params=params,
                )
                if r.status_code != 200:
                    return None
                data = r.json()
                for off in promo_offers_from_response(data):
                    if isinstance(off, dict) and promo_id_matches(promo_offer_id(off), promotion_id):
                        return promo_candidate_offer_id(off)
            except Exception:
                pass
            return None

        async def apply_one(it):
            iid = it.get("item_id")
            ptype = (it.get("promotion_type") or "DEAL").upper()
            offer_id = it.get("offer_id")
            original_price = it.get("original_price")
            # `discount_pct` desde el frontend = parte del VENDEDOR
            # (loaded_pct), no el total. Para DEAL/PRICE_DISCOUNT
            # se convierte a precio final (top_deal_price /
            # deal_price). Para SMART no se manda — se acepta el
            # candidate offer_id.
            try:
                pct = float(it.get("discount_pct") or 0)
            except (TypeError, ValueError):
                pct = 0
            try:
                min_pct = float(it.get("min_discount_pct")) if it.get("min_discount_pct") is not None else None
            except (TypeError, ValueError):
                min_pct = None
            if min_pct is not None and pct <= min_pct:
                return {
                    "item_id": iid,
                    "ok": False,
                    "status": 0,
                    "error": f"Descuento insuficiente: MercadoLibre exige mas de {min_pct:g}%. Cargá un porcentaje mayor antes de aplicar.",
                    "payload": None,
                    "attempts": 0,
                }
            deal_price = None
            if original_price and pct > 0:
                try:
                    deal_price = round(float(original_price) * (1 - pct / 100.0), 2)
                except (TypeError, ValueError):
                    deal_price = None
            # Cada tipo de promo tiene su shape de payload distinta.
            # IMPORTANTE: la API v2 unificada usa SIEMPRE `promotion_id`
            # (no `deal_id` como decían los docs viejos de Traditional
            # Campaigns). Si no se manda, ML responde
            # 400 "Promotion id is required".
            # OJO: `top_deal_price` NO es un alias de deal_price — es el
            # precio EXCLUSIVO Meli+ (loyalty). Si lo mandás igual al
            # deal_price, ML rechaza con LOYALTY_PRICE_DIFF_TOO_LOW.
            # Por ahora NO lo enviamos (el descuento loyalty es opcional).
            # Promos co-fondeadas con ML (donde ML aporta un %) se aplican
            # aceptando la oferta candidata: ML exige el offer_id correcto.
            # SMART/MARKETPLACE no llevan deal_price (el precio lo fija la
            # oferta de ML); DEAL/SELLER_CAMPAIGN/PRICE_DISCOUNT sí.
            if ptype in ("SMART", "MARKETPLACE_CAMPAIGN"):
                payload = {
                    "promotion_id": promotion_id,
                    "promotion_type": ptype,
                    "offer_id": offer_id,
                }
            elif ptype == "SELLER_CAMPAIGN":
                payload = {
                    "promotion_id": promotion_id,
                    "promotion_type": "SELLER_CAMPAIGN",
                    "deal_price": deal_price,
                }
            elif ptype == "PRICE_DISCOUNT":
                payload = {
                    "promotion_id": promotion_id,
                    "promotion_type": "PRICE_DISCOUNT",
                    "deal_price": deal_price,
                }
            elif ptype == "DEAL":
                payload = {
                    "promotion_id": promotion_id,
                    "promotion_type": "DEAL",
                    "deal_price": deal_price,
                }
            else:
                # Fallback (LIGHTNING, VOLUME, DOD, etc.) — mandamos
                # precio final si lo tenemos.
                payload = {
                    "promotion_id": promotion_id,
                    "promotion_type": ptype,
                    "deal_price": deal_price,
                }
            # El offer_id identifica la oferta concreta del item; mandarlo
            # cuando lo tengamos ayuda a ML a ubicar el candidato y evita
            # CANDIDATE_NOT_FOUND en cualquier tipo de promo.
            if offer_id:
                payload["offer_id"] = offer_id
            # Limpiamos campos None
            payload = {k: v for k, v in payload.items() if v is not None}
            async with sem:
                last_status = 0
                last_error = None
                offer_id_refreshed = False
                for attempt in range(5):
                    try:
                        r = await client.post(
                            f"{ML_API_URL}/seller-promotions/items/{iid}",
                            headers=headers,
                            params={"app_version": "v2"},
                            json=payload,
                        )
                        if r.status_code in (200, 201, 204):
                            return {
                                "item_id": iid,
                                "ok": True,
                                "status": r.status_code,
                                "error": None,
                                "payload": payload,
                                "attempts": attempt + 1,
                            }
                        last_status = r.status_code
                        last_error = r.text[:500]
                        # Reintentos por status server-side / lock.
                        retriable = r.status_code in (423, 429, 500, 502, 503, 504)
                        # Algunos 400 de ML son TRANSITORIOS (ML está procesando
                        # internamente la publicación): reintentar también.
                        #  - OFFER_SIBLING_CREATION_IN_PROCESS: oferta hermana en
                        #    creación; al rato deja aplicar el precio.
                        #  - REST_CREDIBILITY_API_ERROR: microservicio de precios
                        #    de ML caído/ocupado, suele recuperarse.
                        if r.status_code == 400 and (
                            "OFFER_SIBLING_CREATION_IN_PROCESS" in last_error
                            or "REST_CREDIBILITY_API_ERROR" in last_error
                            or "Credibility Api" in last_error
                        ):
                            retriable = True
                        # CANDIDATE_NOT_FOUND: el offer_id que mandamos no
                        # corresponde (o falta). Lo resolvemos en vivo una vez
                        # y reintentamos con el candidato correcto.
                        if (r.status_code == 400 and "CANDIDATE_NOT_FOUND" in last_error
                                and not offer_id_refreshed):
                            offer_id_refreshed = True
                            fresh = await resolve_offer_id(iid, ptype)
                            if fresh and fresh != offer_id:
                                offer_id = fresh
                                payload["offer_id"] = fresh
                                await asyncio.sleep(0.5)
                                continue
                        if retriable and attempt < 4:
                            await asyncio.sleep(1.5 * (attempt + 1))
                            continue
                        break
                    except Exception as e:
                        last_status = 0
                        last_error = str(e)[:300]
                        if attempt < 4:
                            await asyncio.sleep(1.5 * (attempt + 1))
                            continue
                        break
                return {
                    "item_id": iid,
                    "ok": False,
                    "status": last_status,
                    "error": last_error,
                    "payload": payload,
                    "attempts": 5,
                }
        results = await asyncio.gather(*[apply_one(it) for it in items])
    _invalidate_promo_items_cache(account_id)
    return {"results": results, "ok": True}


@app.post("/api/promociones/{account_id}/{promotion_id}/remove")
async def api_promociones_remove(
    request: Request, account_id: int, promotion_id: str,
):
    """Quita (da de baja) los items seleccionados de una promoción.
    Body JSON: { items: [{item_id, promotion_type?}, ...] }
    ML expone esto como DELETE /seller-promotions/items/{id}."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    body = await request.json()
    items = body.get("items") or []
    if not items:
        raise HTTPException(400, "Faltan items para quitar")
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    results = []
    async with httpx.AsyncClient(timeout=60) as client:
        sem = asyncio.Semaphore(3)

        async def remove_one(it):
            iid = it.get("item_id")
            ptype = (it.get("promotion_type") or "DEAL").upper()
            params = {"app_version": "v2", "promotion_id": promotion_id,
                      "promotion_type": ptype}
            async with sem:
                last_status = 0
                last_error = None
                for attempt in range(5):
                    try:
                        r = await client.request(
                            "DELETE",
                            f"{ML_API_URL}/seller-promotions/items/{iid}",
                            headers=headers,
                            params=params,
                        )
                        if r.status_code in (200, 201, 204):
                            return {"item_id": iid, "ok": True,
                                    "status": r.status_code, "error": None}
                        last_status = r.status_code
                        last_error = r.text[:500]
                        if r.status_code in (423, 429, 500, 502, 503, 504) and attempt < 4:
                            await asyncio.sleep(1.5 * (attempt + 1))
                            continue
                        break
                    except Exception as e:
                        last_status = 0
                        last_error = str(e)[:300]
                        if attempt < 4:
                            await asyncio.sleep(1.5 * (attempt + 1))
                            continue
                        break
                return {"item_id": iid, "ok": False,
                        "status": last_status, "error": last_error}

        results = await asyncio.gather(*[remove_one(it) for it in items])
    _invalidate_promo_items_cache(account_id)
    return {"results": results, "ok": True}


# ── API datos ───────────────────────────────────────────────────

@app.get("/api/orders/{account_id}")
async def api_orders(request: Request, account_id: int,
                     date_from: Optional[str] = None, date_to: Optional[str] = None,
                     refresh: bool = False, fast: bool = False,
                     cache_only: bool = False):
    """Wrapper: captura excepciones y devuelve el traceback en el body (DEBUG).
    Antes un error tiraba 500 genérico ('Internal Server Error') sin detalle,
    y el front lo tragaba en silencio dejando el caché → 'no cambia nada'."""
    try:
        return await _api_orders_impl(request, account_id, date_from, date_to,
                                      refresh, fast, cache_only)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        try:
            print("ERROR api_orders:", tb, flush=True)
        except Exception:
            pass
        return JSONResponse(
            status_code=500,
            content={"error": str(e), "traceback": tb.splitlines()[-30:]},
        )


async def _api_orders_impl(request: Request, account_id: int,
                     date_from: Optional[str] = None, date_to: Optional[str] = None,
                     refresh: bool = False, fast: bool = False,
                     cache_only: bool = False):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)

    today = datetime.utcnow().date()
    df = date_from or str(today - timedelta(days=365))
    dt = date_to or str(today)
    # Limitar rango a 1 año máximo
    if (datetime.strptime(dt, "%Y-%m-%d") - datetime.strptime(df, "%Y-%m-%d")).days > 365:
        df = str(datetime.strptime(dt, "%Y-%m-%d").date() - timedelta(days=365))

    user = get_user(user_id)
    admin_orders_mode = bool(user and user.get("is_admin"))
    # Refrescar SIEMPRE recompone el rango completo contra ML (no solo admin).
    # Antes esto estaba gateado por is_admin: si la cuenta NO era admin, al
    # tocar "Actualizar" en 30 días solo se re-bajaban los últimos 7 días y el
    # resto salía del caché viejo (armado con date_closed) → el número no se
    # movía. Ahora cualquier refresh fuerza cff=None (rango completo) y reemplaza
    # el caché. Por eso "no cambiaba nada".
    admin_full_refresh = bool(refresh)
    # IMPORTANTE: buscamos por order.date_created (fecha real de la venta).
    # NO usar order.date_closed: ese campo solo se completa cuando ML "cierra"
    # la orden (finalizada/entregada), así que las ventas pagadas que todavía
    # están en proceso NO tienen date_closed y quedaban EXCLUIDAS del search →
    # faltaba ~9-17% de las ventas (peor en rangos recientes). date_created es
    # inmutable, siempre está, y es la fecha que usa ML en "Ventas del período".
    order_search_date_field = "order.date_created"

    # ── Cache lookup ──────────────────────────────────────────────
    # cache_fetched_from registra hasta dónde hemos bajado de ML para esta
    # cuenta. Si df >= cache_fetched_from, el rango pedido está completamente
    # cacheado. Si no, hay un hueco más viejo que tenemos que bajar de ML.
    cached_orders = db_fetch_order_snapshots(account_id, df, dt)
    cache_fetched_from = acc.get("cache_fetched_from")
    cff_iso = str(cache_fetched_from)[:10] if cache_fetched_from else None
    if cache_only:
        payload = build_dashboard_payload(cached_orders, details_complete=True)
        payload["cache_only"] = True
        payload["cache_empty"] = not bool(cached_orders)
        payload["cache_covers_older"] = cff_iso is not None and df >= cff_iso
        return payload
    # Si no hay órdenes en caché pero cff_iso está seteado, el caché está roto
    # (fetch previo falló y quedó marcado como "completo"). Forzar refetch
    # del rango completo en lugar de confiar en el cff_iso poisoned.
    if not cached_orders and cff_iso is not None:
        cff_iso = None
    # Para el administrador, refresh=1 significa recomponer el rango pedido
    # completo. No confiamos en cache_fetched_from porque puede haber quedado
    # adelantado por un fetch parcial anterior y dejar miles de ventas afuera.
    if admin_full_refresh:
        cff_iso = None
    cache_covers_older = cff_iso is not None and df >= cff_iso

    # Pure cache hit: rango cubierto + no se pidió refresh → devolver inmediato.
    if cached_orders and not refresh and cache_covers_older:
        return build_dashboard_payload(cached_orders, details_complete=True)

    # ── Determinar qué rangos hay que bajar de ML ─────────────────
    # Idea: NO re-bajar lo que ya tenemos en caché. Solo:
    #  (a) El hueco más viejo [df, cff_iso - 1d] si df < cff_iso.
    #  (b) Los últimos RECENT_REFRESH_DAYS días si refresh=True (datos recientes
    #      pueden cambiar de estado: pending→paid, paid→refund, etc).
    RECENT_REFRESH_DAYS = 7
    recent_cutoff_iso = (today - timedelta(days=RECENT_REFRESH_DAYS)).isoformat()

    fetch_ranges: list = []  # list of (df_iso, dt_iso)
    if cff_iso is None:
        # No hay caché todavía: bajar el rango completo
        fetch_ranges.append((df, dt))
    else:
        if df < cff_iso:
            # Hueco viejo a llenar
            gap_end = (datetime.strptime(cff_iso, "%Y-%m-%d").date() - timedelta(days=1)).isoformat()
            fetch_ranges.append((df, gap_end))
        if refresh:
            recent_start = max(recent_cutoff_iso, df)
            if recent_start <= dt:
                fetch_ranges.append((recent_start, dt))

    if not fetch_ranges:
        # Defensa: no debería pasar dado el early-return, pero si llegamos
        # acá con caché y sin nada que bajar, devolver el caché tal cual.
        return build_dashboard_payload(cached_orders, details_complete=True)

    token = await refresh_ml_token(account_id)
    if not token:
        # Si tenemos caché pero no token, igual servimos el caché.
        if cached_orders:
            return build_dashboard_payload(cached_orders, details_complete=True)
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}

    base_search = {
        "seller": acc["ml_user_id"],
        "sort": "date_desc",
        "limit": 50,
    }

    # Lista compartida para errores que vuelven a ML (429, 500, etc.). Si algo
    # falla, lo loggeamos y lo devolvemos en el payload para que se vea en el UI.
    fetch_errors: list = []
    # Counters de diagnóstico para entender qué pasó del lado de ML.
    diag = {"chunks_fetched": 0, "raw_ml_orders": 0, "split_events": 0, "retries": 0,
            "ml_reported_total": 0, "days_over_cap": 0}

    # Threshold para forzar split del chunk. Bajado a 900 con margen ante
    # variabilidad de paging.total que devuelve ML (a veces reporta el cap).
    SPLIT_THRESHOLD = 900

    # Pool de conexiones amplio: con cientos de envíos + búsqueda de órdenes en
    # paralelo, el default de httpx (max_connections=100) se satura y genera
    # timeouts en cascada en rangos grandes (7/30 días). Lo subimos.
    http_limits = httpx.Limits(max_connections=200, max_keepalive_connections=50)
    # Cap global de concurrencia para /orders/search. La recursión binaria de
    # chunks puede disparar 100+ requests simultáneas y ML responde 429; al
    # agotar reintentos perdíamos chunks enteros (miles de ventas) → totales mal.
    order_search_sem = asyncio.Semaphore(20)
    async with httpx.AsyncClient(timeout=180, limits=http_limits) as client:
        async def _get_with_retry(params):
            """GET a /orders/search con reintentos sobre 429/5xx/timeouts.
            ML rate-limitea bastante agresivo cuando recursamos en paralelo,
            y silenciar esos errores nos cuesta cientos/miles de ventas."""
            last_err = None
            for attempt in range(4):
                try:
                    async with order_search_sem:
                        rp = await client.get(f"{ML_API_URL}/orders/search", headers=headers, params=params)
                except Exception as e:
                    last_err = f"exc {str(e)[:120]}"
                    if attempt < 3:
                        diag["retries"] += 1
                        await asyncio.sleep(0.5 * (2 ** attempt))
                        continue
                    return None, last_err
                if rp.status_code == 200:
                    return rp, None
                if rp.status_code in (429, 500, 502, 503, 504) and attempt < 3:
                    diag["retries"] += 1
                    # backoff exponencial: 0.5s, 1s, 2s, 4s. ML respira y volvemos.
                    await asyncio.sleep(0.5 * (2 ** attempt))
                    continue
                return rp, f"HTTP {rp.status_code}: {rp.text[:150]}"
            return None, last_err or "unknown"

        async def fetch_chunk_recursive(start_d, end_d, depth=0):
            """Trae todas las órdenes en [start_d, end_d]. Si el chunk pega el
            tope que /orders/search expone (max 1000 paginables, paging.total
            puede reportar más), parte el rango por la mitad y recursa.

            Buscamos por order.date_created (fecha real de la venta, siempre
            presente). Para no admin se mantiene date_last_updated."""
            from_utc = datetime.combine(start_d, datetime.min.time()) + timedelta(hours=3) - timedelta(hours=24)
            to_utc = datetime.combine(end_d, datetime.min.time()) + timedelta(hours=27)
            chunk_params = {
                **base_search,
                f"{order_search_date_field}.from": from_utc.strftime("%Y-%m-%dT%H:%M:%S.000-00:00"),
                f"{order_search_date_field}.to":   to_utc.strftime("%Y-%m-%dT%H:%M:%S.000-00:00"),
            }
            r, err = await _get_with_retry({**chunk_params, "offset": 0})
            if r is None or r.status_code != 200:
                fetch_errors.append(f"[{start_d}..{end_d}] {err}")
                return []
            data = r.json()
            results = list(data.get("results", []))
            total = data.get("paging", {}).get("total", 0)

            # Si pegamos el tope y todavía hay días para partir, recursamos.
            # max_depth=10 alcanza para partir 1 año hasta chunks de ~1 día.
            if total >= SPLIT_THRESHOLD and (end_d - start_d).days >= 1 and depth < 10:
                diag["split_events"] += 1
                span = (end_d - start_d).days
                mid = start_d + timedelta(days=span // 2)
                left, right = await asyncio.gather(
                    fetch_chunk_recursive(start_d, mid, depth + 1),
                    fetch_chunk_recursive(mid + timedelta(days=1), end_d, depth + 1),
                )
                return left + right

            # Hoja: paginar hasta total (cap duro de ML en 1000 paginables).
            diag["chunks_fetched"] += 1
            # Ground-truth: lo que ML dice que existe en esta hoja. Sumado sobre
            # todas las hojas = total que ML reporta para la ventana buscada.
            # Si raw_ml_orders < ml_reported_total, perdimos órdenes (chunk con
            # error o un día con >1000 que no se pudo partir).
            diag["ml_reported_total"] = diag.get("ml_reported_total", 0) + total
            if total > 1000 and (end_d - start_d).days < 1:
                diag["days_over_cap"] = diag.get("days_over_cap", 0) + 1
            capped = min(total, 1000)
            offsets = list(range(50, capped, 50))
            if offsets:
                sem = asyncio.Semaphore(8)
                async def get_page(off):
                    async with sem:
                        rp, perr = await _get_with_retry({**chunk_params, "offset": off})
                        if rp is None or rp.status_code != 200:
                            fetch_errors.append(f"[{start_d}..{end_d}] offset {off} {perr}")
                            return []
                        return rp.json().get("results", [])
                pages = await asyncio.gather(*[get_page(off) for off in offsets])
                for page in pages:
                    results.extend(page)
            diag["raw_ml_orders"] += len(results)
            return results

        async def fetch_one_range(range_df_iso, range_dt_iso):
            """Particiona un rango en chunks iniciales de 30 días y los baja
            en paralelo. fetch_chunk_recursive maneja el split si alguno pega el tope."""
            range_df_d = datetime.strptime(range_df_iso, "%Y-%m-%d").date()
            range_dt_d = datetime.strptime(range_dt_iso, "%Y-%m-%d").date()
            chunks = []
            cur = range_df_d
            while cur <= range_dt_d:
                chunk_end = min(cur + timedelta(days=29), range_dt_d)
                chunks.append((cur, chunk_end))
                cur = chunk_end + timedelta(days=1)
            chunk_lists = await asyncio.gather(*[fetch_chunk_recursive(s, e) for s, e in chunks])
            out = []
            for c in chunk_lists:
                out.extend(c)
            return out

        # Todos los rangos a bajar, en paralelo.
        per_range = await asyncio.gather(*[fetch_one_range(rd, rt) for rd, rt in fetch_ranges])
        all_results: list = []
        for r in per_range:
            all_results.extend(r)

        # Deduplicar por id (rangos pueden solaparse 24h por el offset UTC↔AR).
        seen: set = set()
        deduped: list = []
        for o in all_results:
            oid = o.get("id")
            if oid in seen:
                continue
            seen.add(oid)
            deduped.append(o)
        all_results = deduped

        # Costos de envío: primero desde cache, luego API solo para los que faltan.
        # NOTA: ya no consultamos /billing/integration/... — agregaba latencia
        # sustancial (loops sequenciales internos) y todo lo que necesitamos
        # (Ing./Bonif./Costo Envío) lo da /shipments/{id}/costs.
        all_sids = [(o.get("shipping") or {}).get("id") for o in all_results]
        cost_cache: dict = {}
        billing_by_order: dict = {}
        flex_billing_by_order: dict = {}
        new_shipping_costs: dict = {}
        if not fast:
            unique_sids = [sid for sid in dict.fromkeys(s for s in all_sids if s)]
            cost_cache = db_get_cached_shipping(unique_sids)
            uncached = [sid for sid in unique_sids if sid not in cost_cache]

            if uncached:
                # Cada tarea abre 2 conexiones (/shipments + /costs), así que
                # 80 → hasta 160 conexiones concurrentes contra ML, que es lo
                # que disparaba los timeouts masivos en rangos grandes (7/30
                # días) y dejaba envíos en 0. Bajado a 50 (≈100 conexiones) +
                # reintentos en get_shipping_cost = mucho más confiable.
                ship_sem = asyncio.Semaphore(50)
                async def fetch_ship(sid):
                    async with ship_sem:
                        try:
                            return sid, await get_shipping_cost(client, sid, headers)
                        except Exception:
                            # Un envío que falla NO debe tumbar toda la request.
                            return sid, None
                fetched = await asyncio.gather(*[fetch_ship(sid) for sid in uncached])
                # Solo cacheamos/usamos los que devolvieron datos (None = falla
                # transitoria → se reintenta en el próximo refresh, sin cachear 0).
                new_shipping_costs = {sid: c for sid, c in fetched if c is not None}
                cost_cache.update(new_shipping_costs)

    empty_ship = {"seller": 0.0, "buyer": 0.0, "bonificacion": 0.0}

    # Cargar costos versionados de mercadería (CMV por SKU con fechas).
    # Para admin (no master) los costos son compartidos entre cuentas → _cost_account_id_for.
    cost_aid = _cost_account_id_for(user, account_id)
    versioned_costs = db_get_product_costs(cost_aid)
    # Cargar tarifas Flex propias del vendedor (para reemplazar el costo de
    # envío de las ventas flex con lo que el vendedor le paga a su mensajería).
    flex_tariffs = db_get_flex_tariffs(cost_aid)

    # ── Paso 1: construir lista raw por orden individual ──────────
    raw_list = []
    for o, sid in zip(all_results, all_sids):
        a = float(o.get("total_amount", 0))
        estado = o.get("status", "")
        payments = o.get("payments", [])
        # Bucket por date_created (la MISMA fecha por la que buscamos en ML).
        # Antes el path no-admin bucketeaba por fecha de pago (date_approved),
        # distinta del campo de búsqueda → una orden creada en el rango pero
        # aprobada otro día quedaba fuera del bucket y se descartaba (línea de
        # filtro `df <= fecha <= dt`). Unificar search+bucket evita ese descarte
        # y hace que el período coincida con "Ventas" de ML (que usa date_created).
        sale_date_raw = o.get("date_created") or o.get("date_closed", "")
        fecha, hora = to_ar(sale_date_raw)
        if fecha and not (df <= fecha <= dt):
            continue
        order_id = str(o.get("id"))
        billing = billing_by_order.get(order_id, {})
        flex_billing = flex_billing_by_order.get(order_id, {})
        comision = round(sum(
            float(i.get("sale_fee", 0)) * int(i.get("quantity", 1))
            for i in o.get("order_items", [])
        ), 2)
        ship_info = cost_cache.get(sid, empty_ship) if sid else empty_ship
        if billing.get("has_shipping"):
            envio = billing.get("envio", 0)
            ingreso_envio = billing.get("ingreso_envio", 0)
            bonificacion = billing.get("bonificacion", 0)
        else:
            envio = ship_info["seller"]
            ingreso_envio = ship_info["buyer"]
            bonificacion = ship_info.get("bonificacion", 0)
        if flex_billing.get("has_flex") or flex_billing.get("has_flex_charge"):
            bonificacion = flex_billing.get("bonificacion", bonificacion)
            ingreso_envio = flex_billing.get("ingreso_envio", ingreso_envio) or ingreso_envio
            if flex_billing.get("has_flex_charge"):
                envio = flex_billing.get("envio", envio)
            elif abs(envio - bonificacion) < 0.01:
                envio = 0.0
        if billing.get("has_comision"):
            comision = billing.get("comision", comision)
        refund_amount = round(sum(
            float(p.get("transaction_amount_refunded") or 0)
            for p in payments
        ), 2)
        # Override Flex: si la venta es Flex y tenemos tarifa propia cargada
        # que matchea con la tarifa ML del shipment, usamos la tarifa propia
        # como Costo Envío (lo que el vendedor le paga a su mensajería).
        # Buscamos un valor "tarifa ML" para matchear:
        #  - list_cost de shipping_option (preferido; flex paga lo expone)
        #  - bonificación (en env. gratis es exactamente la tarifa ML)
        #  - ingreso del comprador (último recurso)
        logistic_type = ship_info.get("logistic_type", "") if isinstance(ship_info, dict) else ""
        list_cost = ship_info.get("list_cost", 0) if isinstance(ship_info, dict) else 0
        if logistic_type in {"home_delivery", "self_service"} and flex_tariffs:
            match_ref = list_cost
            if match_ref <= 0:
                match_ref = bonificacion if bonificacion > 0 else (ingreso_envio if ingreso_envio > 0 else 0)
            if match_ref > 0:
                matched = match_flex_tariff(flex_tariffs, match_ref, fecha)
                if matched:
                    envio = flex_cost_with_iva(matched)
        items = []
        order_cmv = 0.0
        for i in o.get("order_items", []):
            sku = (i.get("item", {}).get("seller_sku") or "").strip()
            qty = int(i.get("quantity", 1))
            # Buscar el costo vigente al momento de la venta (más reciente
            # entry con valid_from <= fecha de la venta). Si el costo está
            # cargado sin IVA, le aplicamos 21% para que sea comparable con
            # monto/comisión que vienen con IVA incluido.
            cost_entry = find_cost_for_date(versioned_costs, sku, fecha) if sku else None
            unit_cost = cost_with_iva(cost_entry)
            item_cmv = round(unit_cost * qty, 2)
            order_cmv += item_cmv
            items.append({
                "sku": sku,
                "titulo": i.get("item", {}).get("title", "?"),
                "monto": round(float(i.get("unit_price", 0)) * qty, 2),
                "comision": round(float(i.get("sale_fee", 0)) * qty, 2),
                "cantidad": qty,
                "cmv_unit": round(unit_cost, 2),
                "cmv": item_cmv,
            })
        raw_list.append({
            "id": o.get("id"),
            "pack_id": o.get("pack_id"),
            "order_count": 1,
            "fecha": fecha,
            "hora": hora,
            "monto": round(a, 2),
            "comision": comision,
            "envio": round(envio, 2),
            "shipping_buyer": round(ingreso_envio, 2),
            "bonificacion": round(bonificacion, 2),
            "coupon_amt": round(float((o.get("coupon") or {}).get("amount", 0)), 2),
            "refund_amount": refund_amount,
            "cmv": round(order_cmv, 2),
            "logistic_type": ship_info.get("logistic_type", "") if isinstance(ship_info, dict) else "",
            "estado": estado,
            "items": items,
        })

    # ── Paso 2: agrupar packs (mismo pack_id → un solo envío) ─────
    orders = []
    pack_map: dict = {}
    for raw in raw_list:
        pid = raw["pack_id"]
        if pid:
            if pid not in pack_map:
                pack_map[pid] = {
                    "id": pid, "venta_id": pid,
                    "fecha": raw["fecha"], "hora": raw["hora"],
                    "monto": 0.0, "comision": 0.0,
                    "envio": raw["envio"],            # un envío por pack
                    "shipping_buyer": raw["shipping_buyer"],
                    "bonificacion": raw["bonificacion"],
                    "coupon_total": 0.0,
                    "refund_amount": 0.0,
                    "order_count": 0,
                    "cmv": 0.0,
                    "logistic_type": raw.get("logistic_type", ""),
                    "estado": raw["estado"],
                    "is_pack": True, "items": [],
                }
            p = pack_map[pid]
            p["monto"] = round(p["monto"] + raw["monto"], 2)
            p["comision"] = round(p["comision"] + raw["comision"], 2)
            p["coupon_total"] = round(p["coupon_total"] + raw["coupon_amt"], 2)
            p["refund_amount"] = round(p["refund_amount"] + raw.get("refund_amount", 0), 2)
            p["order_count"] += raw.get("order_count", 1)
            p["cmv"] = round(p["cmv"] + raw["cmv"], 2)
            p["items"].extend(raw["items"])
        else:
            # Ing. Envío = sólo lo que paga el comprador por envío.
            ingreso_envio = raw["shipping_buyer"]
            # Bonif = SÓLO la bonificación de envío que recibe el vendedor.
            # El cupón ML NO se cuenta: lo aporta ML para que el comprador
            # pague menos, pero NO es ingreso del vendedor.
            sku_col = " / ".join(i["sku"] or i["titulo"] for i in raw["items"])
            ganancia = round(
                raw["monto"] + ingreso_envio + raw["bonificacion"]
                - raw["comision"] - raw["envio"] - raw["cmv"], 2
            )
            orders.append({
                "id": raw["id"], "venta_id": raw["id"],
                "order_count": raw.get("order_count", 1),
                "fecha": raw["fecha"], "hora": raw["hora"],
                "producto": sku_col,
                "monto": raw["monto"], "comision": raw["comision"],
                "ingreso_envio": ingreso_envio,
                "bonificacion": raw["bonificacion"],
                "envio": raw["envio"],
                "refund_amount": raw.get("refund_amount", 0),
                "cmv": raw["cmv"],
                "ganancia": ganancia,
                "logistic_type": raw.get("logistic_type", ""),
                "estado": raw["estado"],
                "is_pack": False, "items": raw["items"],
            })

    for pid, p in pack_map.items():
        ingreso_envio = p["shipping_buyer"]
        p["ingreso_envio"] = ingreso_envio
        # Cupón ML NO se suma: lo aporta ML para el comprador, no es ingreso del vendedor.
        p["ganancia"] = round(
            p["monto"] + ingreso_envio + p["bonificacion"]
            - p["comision"] - p["envio"] - p["cmv"], 2
        )
        if len(p["items"]) == 1:
            p["is_pack"] = False
            p["producto"] = p["items"][0]["sku"] or p["items"][0]["titulo"]
        else:
            pack_units = sum(item.get("cantidad", 1) for item in p["items"])
            p["producto"] = f"Paquete x{pack_units}"
        del p["shipping_buyer"], p["coupon_total"]
        orders.append(p)

    orders.sort(key=lambda x: (x.get("fecha") or "", x.get("hora") or ""), reverse=True)

    # ── Merge con el caché ──────────────────────────────────────
    # `orders` contiene solo las órdenes que bajamos de ML en esta llamada
    # (rangos delta). Las órdenes ya cacheadas que estaban dentro de [df, dt]
    # se mergean acá, priorizando las recién bajadas si coinciden por id.
    ml_ids = {o.get("id") for o in orders if o.get("id") is not None}
    cached_for_merge = [] if admin_full_refresh and not fetch_errors else cached_orders
    merged = list(orders) + [o for o in cached_for_merge if o.get("id") not in ml_ids]
    # Filtro defensivo por rango (debería ser redundante porque ya filtramos)
    merged = [o for o in merged if not o.get("fecha") or df <= o["fecha"] <= dt]
    merged.sort(key=lambda x: (x.get("fecha") or "", x.get("hora") or ""), reverse=True)

    if not fast:
        # Persistir caches en background para no demorar la respuesta al cliente.
        if new_shipping_costs:
            asyncio.create_task(asyncio.to_thread(db_save_shipping_costs, new_shipping_costs))
        if admin_full_refresh and not fetch_errors:
            asyncio.create_task(
                asyncio.to_thread(db_replace_order_snapshots_for_range, account_id, df, dt, orders, True)
            )
        elif orders:
            asyncio.create_task(
                asyncio.to_thread(db_save_order_snapshots, account_id, orders, True)
            )
        # Solo avanzar cache_fetched_from si NO hubo errores. Si alguna chunk
        # falló, queremos volver a intentar la próxima vez en lugar de quedar
        # con un caché parcial considerado "completo". Esto fixea el caso de
        # cuentas que terminaban mostrando 0 órdenes para siempre porque la
        # primera fetch tuvo errores y guardamos cff = df igual.
        if not fetch_errors and fetch_ranges:
            min_fetched_df = min(rd for rd, _ in fetch_ranges)
            # IMPORTANTE: usar CAST(:df AS DATE), NO ':df::DATE'. El '::' de
            # Postgres rompe el parser de parámetros nombrados de SQLAlchemy y
            # tiraba 500 ("syntax error at or near :") en CADA refresh → el
            # frontend lo tragaba y volvía al caché ("nunca actualiza").
            db_execute(
                "UPDATE ml_accounts SET cache_fetched_from ="
                " LEAST(COALESCE(cache_fetched_from, CAST('9999-12-31' AS DATE)),"
                " CAST(:df AS DATE)) WHERE id=:id",
                {"df": min_fetched_df, "id": account_id}
            )

    payload = build_dashboard_payload(merged, details_complete=not fast)
    # Diagnostics: cuántos chunks bajamos, cuántas órdenes vio ML, cuántas
    # quedaron después de dedup y filtro de fecha. Para debug del usuario.
    diag["after_dedup"] = len(all_results)
    diag["after_filter"] = len(orders)
    diag["from_cache"] = len([o for o in merged if o.get("id") not in ml_ids])
    diag["fetch_ranges"] = fetch_ranges
    diag["admin_full_refresh"] = admin_full_refresh
    diag["search_date_field"] = order_search_date_field
    # Completitud: ¿bajamos todo lo que ML dice que hay en la ventana buscada?
    # raw_ml_orders = lo que efectivamente trajimos; ml_reported_total = lo que
    # ML reporta (suma de paging.total por hoja). Si difieren o hubo errores,
    # el fetch quedó parcial y NO hay que confiar en los totales.
    diag["lost_orders"] = max(0, diag.get("ml_reported_total", 0) - diag["raw_ml_orders"])
    diag["complete"] = (not fetch_errors) and diag["lost_orders"] == 0

    # ── Breakdown para conciliar contra ML ────────────────────────
    # Permite ver EXACTAMENTE por qué difiere el conteo/monto del panel vs ML:
    #  - status_counts: cuántas órdenes (ya mergeadas, dentro del rango) hay en
    #    cada estado. ML "Ventas" normalmente cuenta solo 'paid'.
    #  - paid_orders / paid_units: conteo de ventas pagadas y unidades.
    #  - paid_amount: facturación bruta (suma de total_amount de las pagadas).
    #  - refund_total: reembolsos detectados (ML resta esto de su facturación).
    #  - paid_amount_net: paid_amount - refund_total (lo que debería matchear ML).
    status_counts: dict = {}
    paid_orders_n = 0
    paid_units = 0
    paid_amount = 0.0
    refund_total = 0.0
    for o in merged:
        st = o.get("estado") or "?"
        n = int(o.get("order_count", 1) or 1)
        status_counts[st] = status_counts.get(st, 0) + n
        refund_total += float(o.get("refund_amount", 0) or 0)
        if st == "paid":
            paid_orders_n += n
            paid_amount += float(o.get("monto", 0) or 0)
            paid_units += sum(int(i.get("cantidad", 1) or 1) for i in (o.get("items") or []))
    diag["status_counts"] = status_counts
    diag["paid_orders"] = paid_orders_n
    diag["paid_units"] = paid_units
    diag["paid_amount"] = round(paid_amount, 2)
    diag["refund_total"] = round(refund_total, 2)
    diag["paid_amount_net"] = round(paid_amount - refund_total, 2)
    payload["diag"] = diag
    if fetch_errors:
        # Devolver los errores en el payload para que el UI los pueda mostrar.
        # Capeamos a 10 para no inflar la respuesta.
        payload["fetch_errors"] = fetch_errors[:10]
    return payload


# Mapeo logistic_type → grupo visible (debe matchear con el del front)
_ENVIO_GROUPS = {
    "home_delivery": "flex", "self_service": "flex",
    "xd_drop_off": "colecta", "drop_off": "colecta",
    "cross_docking": "colecta",
    "fulfillment": "fulfillment",
    "pickup": "pickup",
}
_ENVIO_LABELS = {
    "home_delivery": "Flex", "self_service": "Flex",
    "xd_drop_off": "Colecta", "drop_off": "Colecta",
    "cross_docking": "Colecta",
    "fulfillment": "Full",
    "pickup": "Retiro en persona",
    "mercadoenvios": "Mercado Envíos",
    "mercado_envios_lite": "ME Lite",
}


@app.get("/api/orders-all")
async def api_orders_all(request: Request,
                         date_from: Optional[str] = None, date_to: Optional[str] = None,
                         refresh: bool = False, fast: bool = False,
                         cache_only: bool = False):
    """Vista agregada de TODAS las cuentas ML del usuario en un solo payload.

    Reusa `api_orders` por cuenta (mismo cálculo, caché, delta-fetch y
    refresh), corre las cuentas en paralelo y mergea las órdenes. Cada orden
    queda etiquetada con el nickname de su cuenta (`cuenta`) para poder
    distinguirla en la tabla. Si una cuenta falla, no rompe el resto.
    """
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    accs = db_fetchall(
        "SELECT id, nickname FROM ml_accounts WHERE user_id=:uid"
        " AND access_token IS NOT NULL AND access_token <> '' ORDER BY id",
        {"uid": user_id},
    )
    if not accs:
        payload = build_dashboard_payload([], details_complete=not fast)
        if cache_only:
            payload["cache_only"] = True
            payload["cache_empty"] = True
        return payload

    async def _safe(aid: int):
        try:
            return await _api_orders_impl(request, aid, date_from, date_to, refresh, fast, cache_only)
        except HTTPException:
            return None
        except Exception:
            return None

    results = await asyncio.gather(*[_safe(a["id"]) for a in accs])

    all_orders: list = []
    fetch_errors: list = []
    diag_accounts: dict = {}
    for a, p in zip(accs, results):
        if not p:
            fetch_errors.append(f"[{a['nickname']}] no se pudo cargar")
            continue
        for o in p.get("orders", []):
            o = dict(o)
            o["cuenta"] = a["nickname"]
            all_orders.append(o)
        for e in (p.get("fetch_errors") or []):
            fetch_errors.append(f"[{a['nickname']}] {e}")
        if p.get("diag"):
            diag_accounts[a["nickname"]] = p["diag"]

    payload = build_dashboard_payload(all_orders, details_complete=not fast)
    payload["diag"] = {"accounts": diag_accounts, "n_accounts": len(accs)}
    if cache_only:
        payload["cache_only"] = True
        # Vacío sólo si NINGUNA cuenta tiene órdenes cacheadas.
        payload["cache_empty"] = not any((p and p.get("orders")) for p in results)
    if fetch_errors:
        payload["fetch_errors"] = fetch_errors[:10]
    return payload


@app.get("/api/orders/{account_id}/export")
async def api_orders_export(
    request: Request, account_id: int,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    sku: Optional[str] = None, envio: Optional[str] = None,
):
    """Descarga las ventas del período como .xlsx. Respeta los mismos
    filtros del Dashboard (SKU y tipo de envío). Lee del caché — el
    Dashboard ya se encarga de refrescar contra ML."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = _account_for_user(account_id, user_id)
    if not acc:
        raise HTTPException(404)
    today = datetime.utcnow().date()
    df = date_from or str(today - timedelta(days=30))
    dt = date_to or str(today)
    cached_orders = db_fetch_order_snapshots(account_id, df, dt) or []

    # Filtros
    sku_q = (sku or "").strip().upper()
    envio_q = (envio or "").strip().lower()
    def _matches(o):
        if sku_q:
            items = o.get("items") or []
            if not any(sku_q in (i.get("sku") or "").upper() for i in items):
                return False
        if envio_q:
            grp = _ENVIO_GROUPS.get(o.get("logistic_type") or "", "otro")
            if grp != envio_q:
                return False
        return True
    orders = [o for o in cached_orders if _matches(o)]
    orders = sorted(orders, key=lambda x: (x.get("fecha") or "", x.get("hora") or ""), reverse=True)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    headers = [
        "Fecha", "Hora", "N° Venta", "Estado", "Producto", "SKU", "Envío",
        "Unidades", "Monto", "Comisión", "Ingreso Envío", "Bonificación",
        "Costo Envío", "CMV", "Ganancia neta", "Margen %",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for o in orders:
        items = o.get("items") or []
        sku_list = ", ".join(sorted({(i.get("sku") or "") for i in items if i.get("sku")}))
        producto = o.get("producto") or (items[0].get("titulo") if items else "")
        units = sum(int(i.get("cantidad") or 1) for i in items) if items else 0
        envio_label = _ENVIO_LABELS.get(o.get("logistic_type") or "", o.get("logistic_type") or "—")
        monto = float(o.get("monto") or 0)
        ganancia = float(o.get("ganancia") or 0)
        margen = (ganancia / monto * 100) if monto > 0 else 0
        ws.append([
            o.get("fecha") or "",
            o.get("hora") or "",
            str(o.get("venta_id") or o.get("id") or ""),
            o.get("estado") or "",
            producto,
            sku_list,
            envio_label,
            units,
            monto,
            float(o.get("comision") or 0),
            float(o.get("ingreso_envio") or 0),
            float(o.get("bonificacion") or 0),
            float(o.get("envio") or 0),
            float(o.get("cmv") or 0),
            ganancia,
            round(margen, 2),
        ])

    # Ancho de columnas + formato moneda básico
    widths = [12, 8, 16, 10, 40, 22, 12, 8, 14, 14, 14, 14, 14, 14, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=15):
        for cell in row:
            cell.number_format = '"$"#,##0.00'
    for row in ws.iter_rows(min_row=2, min_col=16, max_col=16):
        for cell in row:
            cell.number_format = '0.00"%"'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nick = (acc.get("nickname") or acc.get("ml_user_id") or "cuenta").replace(" ", "_")
    fname = f"ventas_{nick}_{df}_{dt}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


    # ── Paso 3: agregados diarios y top productos ─────────────────
    daily: dict = {}
    products: dict = {}
    for order in orders:
        if order.get("fecha") and order.get("estado") == "paid":
            f = order["fecha"]
            daily.setdefault(f, {"ventas": 0, "ingresos": 0, "ganancia": 0})
            daily[f]["ventas"] += 1
            daily[f]["ingresos"] += order["monto"]
            daily[f]["ganancia"] += order["ganancia"]
            for item in order.get("items", []):
                t = item.get("sku") or item.get("titulo", "Sin título")
                products.setdefault(t, {"cantidad": 0, "ingresos": 0})
                products[t]["cantidad"] += item.get("cantidad", 1)
                products[t]["ingresos"] += item.get("monto", 0)

    top = sorted(products.items(), key=lambda x: x[1]["ingresos"], reverse=True)[:5]
    return {
        "orders": orders,
        "daily": dict(sorted(daily.items())),
        "top_products": [{"nombre": k, **v} for k, v in top],
        "details_complete": not fast,
        "ultima_actualizacion": datetime.utcnow().isoformat(),
    }


# ── Admin / Permisos ────────────────────────────────────────────

# Páginas a las que se les puede dar/quitar permiso. Las keys son los
# slugs internos; los labels los muestra la UI. El admin SIEMPRE las ve.
PAGES = [
    ("dashboard",   "Dashboard"),
    ("costos",      "Costos (CMV)"),
    ("envios_flex", "Envíos Flex"),
    ("descuentos",  "Descuentos"),
    ("ranking",     "Ranking por SKU"),
    ("etiquetas",   "Etiquetas"),
]
PAGE_KEYS = {k for k, _ in PAGES}


def user_permissions(user: dict) -> set:
    """Devuelve el set de páginas que el usuario puede ver. Admin ve todas."""
    if not user:
        return set()
    if is_admin_user(user):
        return PAGE_KEYS
    perms = user.get("permissions")
    if isinstance(perms, str):
        try:
            perms = json.loads(perms)
        except Exception:
            perms = []
    return set(perms or [])


def require_page(user: dict, page: str):
    """Lanza 403 si el usuario no tiene acceso a esa página."""
    if page not in user_permissions(user):
        raise HTTPException(403, "No tenés permiso para acceder a esta sección")


def password_reset_email_content(name: str, reset_link: str) -> tuple[str, str]:
    text_body = (
        f"Hola {name},\n\n"
        f"Un administrador del Panel ML solicito que establezcas o renueves tu contrasena.\n\n"
        f"Hace click aca para hacerlo:\n{reset_link}\n\n"
        f"El link expira en 24 horas.\n\n"
        f"Si no esperabas este mail, ignoralo; tu contrasena actual sigue intacta."
    )
    html_body = (
        f"<p>Hola {name},</p>"
        "<p>Un administrador del Panel ML solicito que establezcas o renueves tu contrasena.</p>"
        f'<p><a href="{reset_link}">Establecer contrasena</a></p>'
        "<p>El link expira en 24 horas.</p>"
        "<p>Si no esperabas este mail, ignoralo; tu contrasena actual sigue intacta.</p>"
    )
    return text_body, html_body


def send_password_reset_email_resend(to_email: str, name: str, reset_link: str) -> tuple[bool, str]:
    api_key = (os.environ.get("RESEND_API_KEY") or "").strip()
    if not api_key:
        return False, "Falta RESEND_API_KEY"
    sender = (
        os.environ.get("RESEND_FROM")
        or os.environ.get("EMAIL_FROM")
        or os.environ.get("SMTP_FROM")
        or os.environ.get("SMTP_USER")
        or ""
    ).strip()
    if not sender:
        return False, "Falta RESEND_FROM o EMAIL_FROM"

    text_body, html_body = password_reset_email_content(name, reset_link)
    try:
        with httpx.Client(timeout=15) as client:
            resp = client.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {api_key}"},
                json={
                    "from": sender,
                    "to": [to_email],
                    "subject": "Panel ML - Establecer / renovar tu contrasena",
                    "text": text_body,
                    "html": html_body,
                },
            )
        if 200 <= resp.status_code < 300:
            return True, "Mail enviado por Resend"
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text
        return False, f"Error Resend {resp.status_code}: {str(detail)[:220]}"
    except Exception as e:
        return False, f"Error Resend: {str(e)[:180]}"


def send_password_reset_email(to_email: str, name: str, reset_link: str) -> tuple[bool, str]:
    """Envia el mail de reset. Prefiere API HTTPS; SMTP queda como fallback."""
    if (os.environ.get("RESEND_API_KEY") or "").strip():
        return send_password_reset_email_resend(to_email, name, reset_link)

    host = (os.environ.get("SMTP_HOST") or "").strip()
    port_raw = (os.environ.get("SMTP_PORT") or "587").strip()
    smtp_user = (os.environ.get("SMTP_USER") or "").strip()
    smtp_pass = (os.environ.get("SMTP_PASS") or "").replace(" ", "").strip()
    smtp_from = (os.environ.get("SMTP_FROM") or smtp_user).strip()
    missing = [
        key for key, value in {
            "SMTP_HOST": host,
            "SMTP_USER": smtp_user,
            "SMTP_PASS": smtp_pass,
            "SMTP_FROM": smtp_from,
        }.items()
        if not value
    ]
    if missing:
        return False, "Faltan variables SMTP: " + ", ".join(missing)
    try:
        import smtplib
        from email.message import EmailMessage
        msg = EmailMessage()
        msg["From"] = smtp_from
        msg["To"] = to_email
        msg["Subject"] = "Panel ML — Establecer / renovar tu contraseña"
        msg.set_content(
            f"Hola {name},\n\n"
            f"Un administrador del Panel ML solicitó que establezcas o renueves tu contraseña.\n\n"
            f"Hacé click acá para hacerlo:\n{reset_link}\n\n"
            f"El link expira en 24 horas.\n\n"
            f"Si no esperabas este mail, ignoralo — tu contraseña actual sigue intacta."
        )
        port = int(port_raw)
        with smtplib.SMTP(host, port, timeout=15) as s:
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)
        return True, "Mail enviado"
    except Exception as e:
        print(f"[email] Error enviando reset: {e}")
        detail = str(e)
        if "Network is unreachable" in detail or "Errno 101" in detail:
            detail += ". Railway bloquea SMTP en planes Free/Trial/Hobby; configura RESEND_API_KEY y RESEND_FROM para enviar por HTTPS."
        return False, f"Error SMTP: {detail[:260]}"


@app.get("/admin", response_class=HTMLResponse)
async def admin_panel(request: Request):
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    user = get_user(user_id)
    if not is_admin_user(user):
        raise HTTPException(403)
    users = db_fetchall("""
        SELECT id, email, name, is_admin, role_label, permissions, created_at,
               (password_hash IS NULL) AS pending_setup
        FROM users
        WHERE COALESCE(is_master, FALSE) = FALSE
        ORDER BY created_at DESC
    """)
    # Normalizar permissions a lista de strings
    for u in users:
        perms = u.get("permissions")
        if isinstance(perms, str):
            try:
                u["permissions"] = json.loads(perms)
            except Exception:
                u["permissions"] = []
        elif perms is None:
            u["permissions"] = []
    success = request.query_params.get("success")
    info = request.query_params.get("info")
    return templates.TemplateResponse("admin.html", {
        "request": request, "user": user, "users": users, "pages": PAGES,
        "error": None, "success": success, "info": info,
    })


@app.post("/admin/users/create", response_class=HTMLResponse)
async def admin_create_user(
    request: Request,
    name: str = Form(...), email: str = Form(...),
    role_label: Optional[str] = Form(None),
    permissions: list = Form(default=[]),
):
    """Crea un usuario sin contrasena inicial y genera un token de setup."""
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse("/")
    admin = get_user(user_id)
    if not is_admin_user(admin):
        raise HTTPException(403)

    # Validar permisos contra las páginas conocidas
    perms = [p for p in (permissions or []) if p in PAGE_KEYS]
    role = (role_label or "Colaborador").strip() or "Colaborador"

    token = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(hours=24)
    # El colaborador queda "bajo" el dueño de las cuentas del admin que lo crea.
    owner = _accounts_owner_id(admin, user_id)
    try:
        with engine.connect() as conn:
            conn.execute(
                text(
                    "INSERT INTO users (email, name, role_label, permissions,"
                    " reset_token, reset_expires_at, password_hash, owner_id)"
                    " VALUES (:e, :n, :r, CAST(:p AS JSONB), :t, :ex, NULL, :owner)"
                ),
                {"e": email.lower().strip(), "n": name, "r": role,
                 "p": json.dumps(perms), "t": token, "ex": expires, "owner": owner},
            )
            conn.commit()
    except Exception as e:
        err = str(e).lower()
        if "duplicate" in err or "unique" in err:
            error_msg = "Ese email ya está registrado"
        else:
            error_msg = f"Error al crear el usuario: {str(e)[:300]}"
        users = db_fetchall("""
            SELECT id, email, name, is_admin, role_label, permissions
            FROM users
            WHERE COALESCE(is_master, FALSE) = FALSE
            ORDER BY created_at DESC
        """)
        for u in users:
            perms = u.get("permissions")
            if isinstance(perms, str):
                try: u["permissions"] = json.loads(perms)
                except Exception: u["permissions"] = []
        return templates.TemplateResponse("admin.html", {
            "request": request, "user": admin, "users": users, "pages": PAGES,
            "error": error_msg, "success": None, "info": None,
        })

    reset_link = f"{APP_URL}/reset/{token}"
    sent, mail_detail = send_password_reset_email(email, name, reset_link)
    if sent:
        return RedirectResponse(
            "/admin?" + urlencode({"success": f"Usuario creado. Link de setup enviado a {email}"}),
            status_code=303,
        )
    else:
        msg = f"Usuario creado. {mail_detail}. Compartile este link al usuario (valido 24hs): {reset_link}"
        return RedirectResponse(
            "/admin?" + urlencode({"info": msg}),
            status_code=303,
        )


@app.post("/admin/users/{target_id}/permissions")
async def admin_update_permissions(
    request: Request, target_id: int,
    role_label: Optional[str] = Form(None),
    permissions: list = Form(default=[]),
):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    admin = get_user(user_id)
    if not is_admin_user(admin):
        raise HTTPException(403)
    perms = [p for p in (permissions or []) if p in PAGE_KEYS]
    role = (role_label or "Colaborador").strip() or "Colaborador"
    owner = _accounts_owner_id(admin, user_id)
    with engine.connect() as conn:
        conn.execute(
            text("""
                UPDATE users
                SET role_label=:r, permissions=CAST(:p AS JSONB)
                WHERE id=:id AND COALESCE(is_master, FALSE) = FALSE
            """),
            {"r": role, "p": json.dumps(perms), "id": target_id},
        )
        # Vincula al colaborador con su dueño (backfill de usuarios viejos sin
        # owner_id). No piso si ya tiene dueño ni si es admin.
        conn.execute(
            text("""
                UPDATE users SET owner_id=:owner
                WHERE id=:id AND owner_id IS NULL
                  AND COALESCE(is_admin, FALSE) = FALSE
                  AND COALESCE(is_master, FALSE) = FALSE
            """),
            {"owner": owner, "id": target_id},
        )
        conn.commit()
    return RedirectResponse(f"/admin?success=Permisos actualizados", status_code=303)


@app.post("/admin/users/{target_id}/reset-password")
async def admin_reset_password(request: Request, target_id: int):
    """Genera un token de reset y se lo manda al mail del usuario.
    El admin NO ve la contraseña en ningún momento."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    admin = get_user(user_id)
    if not is_admin_user(admin):
        raise HTTPException(403)
    target = db_fetchone(
        "SELECT id, email, name FROM users WHERE id=:id AND COALESCE(is_master, FALSE) = FALSE",
        {"id": target_id},
    )
    if not target:
        raise HTTPException(404)
    token = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(hours=24)
    db_execute(
        "UPDATE users SET reset_token=:t, reset_expires_at=:ex WHERE id=:id",
        {"t": token, "ex": expires, "id": target_id},
    )
    reset_link = f"{APP_URL}/reset/{token}"
    sent, mail_detail = send_password_reset_email(target["email"], target["name"], reset_link)
    if sent:
        msg = f"Mail de renovación enviado a {target['email']}"
    else:
        msg = f"{mail_detail}. Compartile este link (24hs): {reset_link}"
    return RedirectResponse("/admin?" + urlencode({"info": msg}), status_code=303)


@app.get("/reset/{token}", response_class=HTMLResponse)
async def reset_password_page(request: Request, token: str):
    """Página pública donde el usuario establece su nueva contraseña."""
    user = db_fetchone(
        "SELECT id, email, name FROM users WHERE reset_token=:t AND reset_expires_at > NOW()",
        {"t": token},
    )
    return templates.TemplateResponse("reset_password.html", {
        "request": request, "token": token,
        "user": user, "error": None, "done": False,
    })


@app.post("/reset/{token}", response_class=HTMLResponse)
async def reset_password_submit(
    request: Request, token: str,
    password: str = Form(...), password2: str = Form(...),
):
    user = db_fetchone(
        "SELECT id, email, name FROM users WHERE reset_token=:t AND reset_expires_at > NOW()",
        {"t": token},
    )
    if not user:
        return templates.TemplateResponse("reset_password.html", {
            "request": request, "token": token, "user": None,
            "error": "El link expiró o no es válido. Pedile al admin que genere uno nuevo.",
            "done": False,
        })
    if password != password2:
        return templates.TemplateResponse("reset_password.html", {
            "request": request, "token": token, "user": user,
            "error": "Las contraseñas no coinciden", "done": False,
        })
    if len(password) < 6:
        return templates.TemplateResponse("reset_password.html", {
            "request": request, "token": token, "user": user,
            "error": "La contraseña debe tener al menos 6 caracteres", "done": False,
        })
    hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    db_execute(
        "UPDATE users SET password_hash=:h, reset_token=NULL, reset_expires_at=NULL WHERE id=:id",
        {"h": hashed, "id": user["id"]},
    )
    return templates.TemplateResponse("reset_password.html", {
        "request": request, "token": token, "user": user,
        "error": None, "done": True,
    })


# ── Debug ───────────────────────────────────────────────────────

@app.get("/api/debug/item-full/{account_ref}")
async def debug_item_full(request: Request, account_ref: str):
    """Devuelve el JSON crudo del primer item activo para inspeccionar
    qué campos tiene (offers, discounts, marketplace_campaigns, etc.).
    Sirve para identificar dónde ML expone realmente las promos."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    acc = db_fetchone(
        "SELECT * FROM ml_accounts WHERE user_id=:uid AND (CAST(id AS TEXT)=:ref OR ml_user_id=:ref)",
        {"uid": user_id, "ref": account_ref},
    )
    if not acc:
        raise HTTPException(404)
    token = await refresh_ml_token(acc["id"])
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    seller_id = acc["ml_user_id"]
    async with httpx.AsyncClient(timeout=30) as client:
        # Tomar el primer item activo
        r = await client.get(
            f"{ML_API_URL}/users/{seller_id}/items/search",
            headers=headers,
            params={"status": "active", "limit": 1},
        )
        if r.status_code != 200:
            return {"step": "items/search", "status": r.status_code, "body": r.text[:500]}
        results = r.json().get("results", [])
        if not results:
            return {"step": "items/search", "status": 200, "body": "no results"}
        item_id = results[0]
        # Pedir el item completo SIN filtros para ver todos los campos
        r2 = await client.get(f"{ML_API_URL}/items/{item_id}", headers=headers)
        return {
            "item_id": item_id,
            "status": r2.status_code,
            "available_fields": sorted(list(r2.json().keys())) if r2.status_code == 200 else None,
            "full_item": r2.json() if r2.status_code == 200 else r2.text[:2000],
        }


@app.get("/api/debug/promos/{account_ref}")
async def debug_promos(request: Request, account_ref: str):
    """Prueba múltiples endpoints de promos y devuelve los resultados crudos.
    Acepta tanto el account_id interno (1, 2, ...) como el ml_user_id (1142561912)."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    # Buscar por id interno o por ml_user_id
    acc = db_fetchone(
        "SELECT * FROM ml_accounts WHERE user_id=:uid AND (CAST(id AS TEXT)=:ref OR ml_user_id=:ref)",
        {"uid": user_id, "ref": account_ref},
    )
    if not acc:
        raise HTTPException(404, f"No encontré cuenta con id o ml_user_id = {account_ref}")
    account_id = acc["id"]
    token = await refresh_ml_token(account_id)
    if not token:
        raise HTTPException(502)
    headers = {"Authorization": f"Bearer {token}"}
    seller_id = acc["ml_user_id"]
    attempts = [
        ("seller-promotions list (sin type)",
         f"{ML_API_URL}/seller-promotions/promotions",
         {"app_version": "v2"}),
        ("seller-promotions list (con seller_id)",
         f"{ML_API_URL}/seller-promotions/promotions",
         {"app_version": "v2", "seller_id": seller_id}),
        ("seller-promotions list (DEAL + seller_id)",
         f"{ML_API_URL}/seller-promotions/promotions",
         {"app_version": "v2", "promotion_type": "DEAL", "seller_id": seller_id}),
        ("items search (1 item con offers)",
         f"{ML_API_URL}/users/{seller_id}/items/search",
         {"limit": 1, "include_attributes": "offers"}),
    ]
    out = []
    first_item_id = None
    async with httpx.AsyncClient(timeout=30) as client:
        for label, url, params in attempts:
            try:
                r = await client.get(url, headers=headers, params=params)
                body = r.text[:2000]
                ct = r.headers.get("content-type", "")
                cl = r.headers.get("content-length", "")
                out.append({"label": label, "url": url, "params": params,
                            "status": r.status_code, "content_type": ct,
                            "content_length": cl, "body": body})
                # Capturar el primer item_id para probar /items/{id}?attributes=offers
                if not first_item_id and r.status_code == 200 and label.startswith("items search"):
                    try:
                        data = r.json()
                        results = data.get("results") or []
                        if results:
                            first_item_id = results[0] if isinstance(results[0], str) else results[0].get("id")
                    except Exception:
                        pass
            except Exception as e:
                out.append({"label": label, "url": url, "params": params,
                            "error": str(e)[:300]})

        # Probar también el item específico para ver sus offers/promos
        if first_item_id:
            extra = [
                (f"seller-promotions/items/{first_item_id} (v2)",
                 f"{ML_API_URL}/seller-promotions/items/{first_item_id}",
                 {"app_version": "v2"}),
                (f"marketplace/seller-promotions/items/{first_item_id} (header version v2)",
                 f"{ML_API_URL}/marketplace/seller-promotions/items/{first_item_id}",
                 {"user_id": seller_id}),
                (f"items/{first_item_id} (attributes=offers)",
                 f"{ML_API_URL}/items/{first_item_id}",
                 {"attributes": "offers"}),
                (f"items/{first_item_id} (full)",
                 f"{ML_API_URL}/items/{first_item_id}",
                 {}),
            ]
            for label, url, params in extra:
                try:
                    r = await client.get(url, headers=headers, params=params)
                    body = r.text[:2500]
                    out.append({"label": label, "url": url, "params": params,
                                "status": r.status_code,
                                "content_type": r.headers.get("content-type", ""),
                                "body": body})
                except Exception as e:
                    out.append({"label": label, "url": url, "params": params,
                                "error": str(e)[:300]})
    return {"first_item_tested": first_item_id, "attempts": out}


@app.get("/api/debug/order/{order_id}")
async def debug_order(request: Request, order_id: str):
    """Devuelve TODA la data cruda que devuelve ML para una orden puntual.

    Probá todos los endpoints relevantes para entender de dónde sacar
    la bonificación cuando no aparece en el panel:
      - /orders/{id}
      - /shipments/{shipping_id}
      - /shipments/{shipping_id}/costs
      - /billing/integration/group/ML/order/details (por seller_id)
      - /billing/integration/periods/key/{period}/group/ML/flex/details
      - /orders/{id}/discounts
      - /orders/{id}/feedback
    """
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    accs = db_fetchall("SELECT * FROM ml_accounts WHERE user_id=:uid ORDER BY id", {"uid": user_id})
    if not accs:
        raise HTTPException(404, "No tenés cuentas de ML conectadas")

    out: dict = {"order_id": order_id, "tries": []}
    for acc in accs:
        token = await refresh_ml_token(acc["id"])
        if not token:
            continue
        headers = {"Authorization": f"Bearer {token}"}
        costs_headers = {**headers, "X-Costs-New": "true", "x-format-new": "true"}
        async with httpx.AsyncClient(timeout=30) as client:
            entry: dict = {"account": acc["nickname"], "account_id": acc["id"]}
            r_order = await client.get(f"{ML_API_URL}/orders/{order_id}", headers=headers)
            entry["order_status"] = r_order.status_code
            if r_order.status_code != 200:
                entry["order_error"] = r_order.text[:500]
                out["tries"].append(entry)
                continue
            order = r_order.json()
            entry["order"] = order
            shipping_id = (order.get("shipping") or {}).get("id")
            entry["shipping_id"] = shipping_id

            tasks = {}
            if shipping_id:
                tasks["shipment"] = client.get(f"{ML_API_URL}/shipments/{shipping_id}", headers=headers)
                tasks["shipment_costs"] = client.get(f"{ML_API_URL}/shipments/{shipping_id}/costs", headers=costs_headers)
            tasks["billing"] = client.get(
                f"{ML_API_URL}/billing/integration/group/ML/order/details",
                headers=headers,
                params={"order_ids": order_id, "seller_id": acc["ml_user_id"], "limit": 150},
            )
            period_key = period_key_from_ml_date(order.get("date_created", ""))
            if period_key:
                tasks["flex_billing_BILL"] = client.get(
                    f"{ML_API_URL}/billing/integration/periods/key/{period_key}/group/ML/flex/details",
                    headers=headers,
                    params={"document_type": "BILL", "order_ids": order_id, "limit": 1000},
                )
                tasks["flex_billing_CREDIT"] = client.get(
                    f"{ML_API_URL}/billing/integration/periods/key/{period_key}/group/ML/flex/details",
                    headers=headers,
                    params={"document_type": "CREDIT_NOTE", "order_ids": order_id, "limit": 1000},
                )
                # También probar SIN document_type
                tasks["flex_billing_NO_TYPE"] = client.get(
                    f"{ML_API_URL}/billing/integration/periods/key/{period_key}/group/ML/flex/details",
                    headers=headers,
                    params={"order_ids": order_id, "limit": 1000},
                )
            tasks["discounts"] = client.get(f"{ML_API_URL}/orders/{order_id}/discounts", headers=headers)
            tasks["feedback"] = client.get(f"{ML_API_URL}/orders/{order_id}/feedback", headers=headers)

            results = await asyncio.gather(*tasks.values(), return_exceptions=True)
            for name, res in zip(tasks.keys(), results):
                if isinstance(res, Exception):
                    entry[name] = {"error": str(res)}
                else:
                    try:
                        entry[f"{name}_status"] = res.status_code
                        entry[name] = res.json() if res.status_code in (200, 206) else res.text[:1000]
                    except Exception as e:
                        entry[name] = {"parse_error": str(e), "text": res.text[:500]}

            out["tries"].append(entry)
            return out
    return out


@app.post("/admin/users/delete/{target_id}")
async def admin_delete_user(request: Request, target_id: int):
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(401)
    user = get_user(user_id)
    if not is_admin_user(user):
        raise HTTPException(403)
    if target_id == user_id:
        raise HTTPException(400, "No podés eliminarte a vos mismo")
    db_execute("DELETE FROM users WHERE id=:id AND COALESCE(is_master, FALSE) = FALSE", {"id": target_id})
    return RedirectResponse("/admin", status_code=303)
